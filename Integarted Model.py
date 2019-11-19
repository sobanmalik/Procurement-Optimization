#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#!/usr/bin/env python
# coding: utf-8

# In[1]:
import pandas as pd
from pulp import *
import csv
from math import *
import tkinter as tk
from tkinter import messagebox

class FullScreenApp(object):
    def __init__(self, master, **kwargs):
        self.master=master
        pad=3
        self._geom='450x600+0+0'
        master.geometry("{0}x{1}+0+0".format(
            master.winfo_screenwidth()-pad, master.winfo_screenheight()-pad))
        master.bind('<Escape>',self.toggle_geom)            
    def toggle_geom(self,event):
        geom=self.master.winfo_geometry()
        print(geom,self._geom)
        self.master.geometry(self._geom)
        self._geom=geom
#Interface build
window = tk.Tk()
app=FullScreenApp(window)
#window.geometry()#400x550
window.title("model")
window.configure(background='white')

#background
bg_image0 = tk.PhotoImage(file ="Pictures/edit.png")
#bg_image0 = tk.PhotoImage(file ="Pictures/P2_resized.png")
#bg_image = tk.PhotoImage(file ="Pictures/P2.png")
label_image0 = tk.Label (window,image = bg_image0)
label_image0.grid(row=0,column=0,columnspan=50, rowspan=50)#,columnspan=20, rowspan=20)
window.grid_columnconfigure(0, weight=1)
window.grid_columnconfigure(2, weight=1)


#creating spaces

label3=tk.Label(text='No. of Planned Days',bg='yellow')
label3.grid(column=1,row=14)
enter_days_input=tk.Text(master=window,height=1,width=5,relief="sunken")#,fill=both
enter_days_input.insert(tk.END, '30')
enter_days_input.grid(column=1,row=15)
#window.grid_columnconfigure(1, weight=0)
#HM Input P1
label4=tk.Label(text='Hot Metal Demand(P1)',bg='yellow')
label4.grid(column=0,row=14)
enter_hm_input_P1=tk.Text(master=window,height=1,width=10,relief="sunken")#,fill=both
#enter_hm_input_P1=tk.Text(master=window)
enter_hm_input_P1.insert(tk.END, '300000')
enter_hm_input_P1.grid(column=0,row=15)#,sticky="nsew")#, padx=0, pady=0)

#HM Input P2
label4=tk.Label(text='Hot Metal Demand(P2)',bg='yellow')
label4.grid(column=0,row=16)
enter_hm_input_P2=tk.Text(master=window,height=1,width=10,relief="sunken")#,fill=both
#enter_hm_input_P2=tk.Text(master=window)
enter_hm_input_P2.insert(tk.END, '220000')
enter_hm_input_P2.grid(column=0,row=17)#,sticky="nsew", padx=0, pady=0)
#labelsp3=tk.Label(text='',bg='brown')
#labelsp3.grid(row=9, column=1)

#DRI
label5=tk.Label(text='DRI Demand(P1)',bg='yellow')
label5.grid(column=2,row=14)
enter_dri_input_P1=tk.Text(master=window,height=1,width=10,relief="sunken")#,fill=both
enter_dri_input_P1.insert(tk.END, '75000')
enter_dri_input_P1.grid(column=2,row=15)#,sticky="nsew", padx=0, pady=0)

label6=tk.Label(text='DRI Demand(P2)',bg='yellow')
label6.grid(column=2,row=16)
enter_dri_input_P2=tk.Text(master=window,height=1,width=10,relief="sunken")#,fill=both
enter_dri_input_P2.insert(tk.END, '75000')
enter_dri_input_P2.grid(column=2,row=17)#,sticky="nsew", padx=0, pady=0)
labelsp3=tk.Label(text='Constraints',bg='yellow')
labelsp3.grid(row=16, column=1)

#iron ore constraints list
iron_ore_constraints =['Supply Fines','Supply Lumps','Siding','Logistics Road(P1 and P2)'
                       ,'Logistics Rail(P1 and P2)',
                       'Logistics Road(Barbil)','Logistics Rail(Barbil)','LTC Lower','LTC Upper']

io_constraint_dict = {} # dictionary to store all the IntVars
mb=tk.Menubutton ( window, text="Iron Ore Constraints",height=1,width=20, relief="raised",background='gray95' )
mb.menu  =  tk.Menu ( mb, tearoff = 0 )
mb["menu"]  =  mb.menu
mb.grid(column=1,row=17)#sticky="nsew", padx=0, pady=0)

for i in iron_ore_constraints:
    var = tk.IntVar()
    var.set(True)
    mb.menu.add_checkbutton(label=i, onvalue=True, offvalue=0, variable=var)
    io_constraint_dict[i] = var # add IntVar to the dictionary
#labelsp4=tk.Label(text='',bg='gray95')
#labelsp4.grid(row=16, column=1)
labelsp5=tk.Label(text='Sinter Plant',bg='yellow')
labelsp5.grid(row=18, column=1)
#sinter P1 constraints list
sinter_constraints_P1 =['Expected Production','Feeding Rate','Capacity Constraints',
                           'Fe(lower)','Fe(upper)','SiO2(lower)','SiO2(upper)','Al2O3(lower)','Al2O3(upper)',
                          'CaO(lower)','CaO(upper)','MgO(lower)','MgO(upper)','Basicity(lower)','Basicity(upper)']

sinter_constraint_dict_P1 = {} # dictionary to store all the IntVars
mb_sinter_P1=tk.Menubutton ( window, text=" Constraints P1",height=1,width=18, relief="raised",background='gray95'  )
mb_sinter_P1.menu  =  tk.Menu ( mb_sinter_P1, tearoff = 0 )
mb_sinter_P1["menu"] = mb_sinter_P1.menu
mb_sinter_P1.grid(column=1,row=19)#,sticky="nsew")#, padx=0, pady=0)


for i in sinter_constraints_P1:
    var = tk.IntVar()
    var.set(True)
    mb_sinter_P1.menu.add_checkbutton(label=i, onvalue=True, offvalue=0, variable=var)
    sinter_constraint_dict_P1[i] = var # add IntVar to the dictionary
sinter_constraint_dict_P1['Fe(lower)'].set(False)
sinter_constraint_dict_P1['Fe(upper)'].set(False)
sinter_constraint_dict_P1['SiO2(lower)'].set(False)
sinter_constraint_dict_P1['SiO2(upper)'].set(False)
sinter_constraint_dict_P1['CaO(lower)'].set(False)
sinter_constraint_dict_P1['CaO(upper)'].set(False)
sinter_constraint_dict_P1['MgO(lower)'].set(False)
sinter_constraint_dict_P1['MgO(upper)'].set(False)
    
#sinter P2 constraints list
sinter_constraints_P2 =['Expected Production','Feeding Rate','Capacity Constraints',
                           'Fe(lower)','Fe(upper)','SiO2(lower)','SiO2(upper)','Al2O3(lower)','Al2O3(upper)',
                          'CaO(lower)','CaO(upper)','MgO(lower)','MgO(upper)','Basicity(lower)','Basicity(upper)']

sinter_constraint_dict_P2 = {} # dictionary to store all the IntVars
mb_sinter_P2=tk.Menubutton ( window, text="Constraints P2",height=1,width=18, relief="raised",background='gray95')
mb_sinter_P2.menu  =  tk.Menu ( mb_sinter_P2, tearoff = 0 )
mb_sinter_P2["menu"] = mb_sinter_P2.menu
mb_sinter_P2.grid(column=1,row=20)#,sticky="nsew", padx=0, pady=0)

for i in sinter_constraints_P2:
    var = tk.IntVar()
    var.set(True)
    mb_sinter_P2.menu.add_checkbutton(label=i, onvalue=True, offvalue=0, variable=var)
    sinter_constraint_dict_P2[i] = var # add IntVar to the dictionary
sinter_constraint_dict_P2['Fe(lower)'].set(False)
sinter_constraint_dict_P2['Fe(upper)'].set(False)
sinter_constraint_dict_P2['SiO2(lower)'].set(False)
sinter_constraint_dict_P2['SiO2(upper)'].set(False)
sinter_constraint_dict_P2['CaO(lower)'].set(False)
sinter_constraint_dict_P2['CaO(upper)'].set(False)
sinter_constraint_dict_P2['MgO(lower)'].set(False)
sinter_constraint_dict_P2['MgO(upper)'].set(False)
    
#bf P1 constraints list
labelsp6=tk.Label(text='Blast Furnace',bg='yellow')
labelsp6.grid(row=21, column=1)

bf_constraints_P1 =['Hot Metal Production','Charge Capacity','Fe(lower)','Fe(upper)','Slag Basicity (B1)(lower)','Slag Basicity (B1)(upper)','Slag Amount(lower)',
                           'Slag Amount(upper)','Sulphur in HM(lower)','Sulphur in HM(upper)','Phosphorus(lower)','Phosphorus(upper)',
                             'CaO(lower)','CaO(upper)','MgO(lower)','MgO(upper)','SiO2(lower)','SiO2(upper)',
                          'Al2O3(lower)','Al2O3(upper)','Sinter Pellet Lump Ratio(lower)','Sinter Pellet Lump Ratio(upper)',
                       'Coke Inputs','Coke Plant(C1)', 'Pellet Production','Pellet Demand']

bf_constraint_dict_P1 = {} # dictionary to store all the IntVars
mb_bf_P1=tk.Menubutton ( window, text="Constraints P1",height=1,width=18, relief="raised",background='gray95' )
mb_bf_P1.menu  =  tk.Menu ( mb_bf_P1, tearoff = 0 )
mb_bf_P1["menu"] = mb_bf_P1.menu
mb_bf_P1.grid(column=1,row=22)#,sticky="nsew", padx=0, pady=0)

for i in bf_constraints_P1:
    var = tk.IntVar()
    var.set(True)
    mb_bf_P1.menu.add_checkbutton(label=i, onvalue=True, offvalue=0, variable=var)
    bf_constraint_dict_P1[i] = var # add IntVar to the dictionary
bf_constraint_dict_P1['Fe(upper)'].set(False)
bf_constraint_dict_P1['Sulphur in HM(lower)'].set(False)
bf_constraint_dict_P1['Sulphur in HM(upper)'].set(False)
bf_constraint_dict_P1['CaO(upper)'].set(False)
bf_constraint_dict_P1['SiO2(upper)'].set(False)
    

#bf P2 constraints list
bf_constraints_P2 =['Hot Metal Production','Charge Capacity','Fe(lower)','Fe(upper)','Slag Basicity (B1)(lower)','Slag Basicity (B1)(upper)','Slag Amount(lower)',
                           'Slag Amount(upper)','Sulphur in HM(lower)','Sulphur in HM(upper)','Phosphorus(lower)','Phosphorus(upper)',
                             'CaO(lower)','CaO(upper)','MgO(lower)','MgO(upper)','SiO2(lower)','SiO2(upper)',
                          'Al2O3(lower)','Al2O3(upper)','Sinter Pellet Lump Ratio(lower)','Sinter Pellet Lump Ratio(upper)',
                         'Coke Inputs','Coke Plant(C1)','Pellet Production','Pellet Demand']

bf_constraint_dict_P2 = {} # dictionary to store all the IntVars
mb_bf_P2=tk.Menubutton ( window, text="Constraints Raiagrh",height=1,width=18, relief="raised",background='gray95'  )
mb_bf_P2.menu  =  tk.Menu ( mb_bf_P2, tearoff = 0 )
mb_bf_P2["menu"] = mb_bf_P2.menu
mb_bf_P2.grid(column=1,row=23)#,sticky="nsew", padx=0, pady=0)

for i in bf_constraints_P2:
    var = tk.IntVar()
    var.set(True)
    mb_bf_P2.menu.add_checkbutton(label=i, onvalue=True, offvalue=0, variable=var)
    bf_constraint_dict_P2[i] = var # add IntVar to the dictionary
bf_constraint_dict_P2['Fe(upper)'].set(False)
bf_constraint_dict_P2['Sulphur in HM(lower)'].set(False)
bf_constraint_dict_P2['Sulphur in HM(upper)'].set(False)
bf_constraint_dict_P2['CaO(upper)'].set(False)
bf_constraint_dict_P2['SiO2(upper)'].set(False)
#label2=tk.Label(text='',bg='gray95')
#label2.grid(row=23, column=1)

#coke P1 constraints list
label2=tk.Label(text='Coke Plant',bg='yellow')
label2.grid(row=24, column=1)
coke_constraints_P1 =['VM(lower)','VM(upper)','CSN(lower)','CSN(upper)','MMR(lower)','MMR(upper)','Vitrinite(lower)',
                           'Vitrinite(upper)','Log(Fluidity)(lower)','Log(Fluidity)(upper)','V9-V14(lower)','V9-V14(upper)',
                             'Sulphur(lower)','Prime Hard Mix(lower)','Prime Hard Mix(upper)','Semi Hard Mix(lower)','Semi Hard Mix(upper)',
                          'Semi Soft Mix(lower)','Semi Soft Mix(upper)','PCI Mix(lower)','PCI Mix(upper)','Pet Coke Mix(lower)','Pet Coke Mix(upper)','Non Coking Coal Mix(lower)',
                         'Non Coking Coal Mix(upper)','Coke Ash','Coke Sulphur','CSR','Inventory']

coke_constraint_dict_P1 = {} # dictionary to store all the IntVars
mb_coke_P1=tk.Menubutton ( window, text="Coke Plant Constraints",height=1,width=1, relief="raised",background='gray95' )
mb_coke_P1.menu  =  tk.Menu ( mb_coke_P1, tearoff = 0 )
mb_coke_P1["menu"] = mb_coke_P1.menu
mb_coke_P1.grid(column=1,row=25,sticky="nsew", padx=0, pady=0)

for i in coke_constraints_P1:
    var = tk.IntVar()
    var.set(True)
    mb_coke_P1.menu.add_checkbutton(label=i, onvalue=True, offvalue=0, variable=var)
    coke_constraint_dict_P1[i] = var # add IntVar to the dictionary

#coke P2 constraints list
coke_constraints_P2 =['VM(lower)','VM(upper)','CSN(lower)','CSN(upper)','MMR(lower)','MMR(upper)','Vitrinite(lower)',
                           'Vitrinite(upper)','Log(Fluidity)(lower)','Log(Fluidity)(upper)','V9-V14(lower)','V9-V14(upper)',
                             'Sulphur(lower)','Prime Hard Mix(lower)','Prime Hard Mix(upper)','Semi Hard Mix(lower)','Semi Hard Mix(upper)',
                          'Semi Soft Mix(lower)','Semi Soft Mix(upper)','PCI Mix(lower)','PCI Mix(upper)','Pet Coke Mix(lower)','Pet Coke Mix(upper)','Non Coking Coal Mix(lower)',
                         'Non Coking Coal Mix(upper)','Coke Ash','Coke Sulphur','CSR','Inventory']

coke_constraint_dict_P2 = {} # dictionary to store all the IntVars
mb_coke_P2=tk.Menubutton ( window, text="Coke Plant Constraints",height=1,width=1, relief="raised",background='gray95' )
mb_coke_P2.menu  =  tk.Menu ( mb_coke_P2, tearoff = 0 )
mb_coke_P2["menu"] = mb_coke_P2.menu
mb_coke_P2.grid(column=1,row=26,sticky="nsew", padx=0, pady=0)

for i in coke_constraints_P2:
    var = tk.IntVar()
    var.set(True)
    mb_coke_P2.menu.add_checkbutton(label=i, onvalue=True, offvalue=0, variable=var)
    coke_constraint_dict_P2[i] = var # add IntVar to the dictionary
label2=tk.Label(text='',bg='brown')
label2.grid(row=27, column=1)

# getting input from interface
#class solver():
def model():
    no_of_days_in_month=30 
    no_of_days_input=float(enter_days_input.get("1.0",'end-1c'))#str(enter_days_input.get())
    hot_metal_input_P1=float(enter_hm_input_P1.get("1.0",'end-1c'))
    hot_metal_input_P2=float(enter_hm_input_P2.get("1.0",'end-1c'))
    dri_input_P1=float(enter_dri_input_P1.get("1.0",'end-1c'))
    dri_input_P2=float(enter_dri_input_P2.get("1.0",'end-1c'))
    #print(no_of_days_input*2)
    mines=pd.read_csv('Data Input/Head Office/Iron Ore/mines.csv')
    mines.loc[mines['Siding'].isna(),'Siding']='Road'
    mines['possibilities']=mines['Mine']+mines['Mine Owners']+mines['Plant']+mines['Siding']
    mines.head()


    # In[21]:


    exmine_cost=pd.read_csv('Data Input/Head Office/Iron Ore/Exmine cost.csv')
    exmine_cost.loc[exmine_cost['with_int_lumps'].isna(),'with_int_lumps']=0
    exmine_cost.loc[exmine_cost['with_int_fines'].isna(),'with_int_fines']=0
    exmine_cost.head()


    # In[22]:


    fines=list()
    lumps=list()
    for j in mines['Mine']:
        fines.append(exmine_cost.loc[(exmine_cost['Mines']==j),'with_int_fines'].item()) 
        lumps.append(exmine_cost.loc[(exmine_cost['Mines']==j),'with_int_lumps'].item())
    fines_list= pd.Series(fines)
    lump_list= pd.Series(lumps)
    mines['with_int_fines'] = fines_list.values
    mines['with_int_lumps']=  lump_list.values
    mines.head()


    # In[24]:


    #linear programming

    #variables


    # In[25]:

    #constraints input
    supply_constraint_table=pd.read_csv('Data Input/Head Office/Iron Ore/supply constraint.csv')
    supply_constraint_table['Fines']=supply_constraint_table['Fines'].str.replace(',','')
    supply_constraint_table['Fines']=supply_constraint_table['Fines'].str.replace('-','0').astype(float)
    supply_constraint_table['Lumps']=supply_constraint_table['Lumps'].str.replace(',','')
    supply_constraint_table['Lumps']=supply_constraint_table['Lumps'].str.replace('-','0').astype(float)
    supply_constraint_table.loc[supply_constraint_table['Lumps'].isna(),'Lumps']=0
    supply_constraint_table.head()


    # In[27]:


    ltc_constraint_table=pd.read_csv('Data Input/Head Office/Iron Ore/LTC Constraints.csv')
    ltc_constraint_table['Min']=ltc_constraint_table['Min'].str.replace(',','').astype(float)
    ltc_constraint_table['Max']=ltc_constraint_table['Max'].str.replace(',','').astype(float)
    ltc_constraint_table



    # In[28]:


    logistics_constraint_table=pd.read_csv('Data Input/Head Office/Iron Ore/Logistics Constraint.csv')
    logistics_constraint_table


    # In[29]:


    miner_constraint_table=pd.read_csv('Data Input/Head Office/Iron Ore/Miner Constraint.csv')
    miner_constraint_table


    # In[30]:


    demand_constraint_table=pd.read_csv('Data Input/Head Office/Iron Ore/demand constraint.csv')
    demand_constraint_table['Fines']=demand_constraint_table['Fines'].str.replace(',','').astype(float)
    demand_constraint_table['Lumps']=demand_constraint_table['Lumps'].str.replace(',','').astype(float)
    demand_constraint_table.loc[demand_constraint_table['Lumps'].isna(),'Lumps']=0
    demand_constraint_table


    # In[31]:


    siding_constraint_table=pd.read_csv('Data Input/Head Office/Iron Ore/siding handling constraints.csv')


    # In[3]:


    mines['Total Cost Fines']=mines['TF']+mines['with_int_fines']
    mines['Total Cost Lumps']=mines['TF']+mines['with_int_lumps']
    mines.head()

    quality_fines=pd.read_csv('Data Input/Plant/Iron Ore/Fines Mines Quality .csv')
    quality_fines['1-loi']=1-(quality_fines['LoI']/100)
    quality_fines['1-moi']=1-(quality_fines['Moisture']/100)
    #quality_fines['CaO']=0.5
    #quality_fines['MgO']=0.1


    for i in quality_fines['Miner']:
        mines.loc[mines['Mine']==i,'Fe_fines']=quality_fines.loc[quality_fines['Miner']==i,'Fe'].item()  
        mines.loc[mines['Mine']==i,'Moisture_fines']=quality_fines.loc[quality_fines['Miner']==i,'Moisture'].item()
        mines.loc[mines['Mine']==i,'SiO2_fines']=quality_fines.loc[quality_fines['Miner']==i,'SiO2'].item()
        mines.loc[mines['Mine']==i,'Al2O3_fines']=quality_fines.loc[quality_fines['Miner']==i,'Al2O3'].item()
        mines.loc[mines['Mine']==i,'1-loi_fines']=quality_fines.loc[quality_fines['Miner']==i,'1-loi'].item()
        mines.loc[mines['Mine']==i,'1-moi_fines']=quality_fines.loc[quality_fines['Miner']==i,'1-moi'].item()
        mines.loc[mines['Mine']==i,'CaO_fines']=quality_fines.loc[quality_fines['Miner']==i,'CaO'].item()
        mines.loc[mines['Mine']==i,'MgO_fines']=quality_fines.loc[quality_fines['Miner']==i,'MgO'].item()


    # In[4]:


    quality_lumps=pd.read_csv('Data Input/Plant/Iron Ore/Lumps Mines Quality.csv')
    quality_lumps['1-loi']=1-quality_lumps['LoI']/100
    quality_lumps['1-moi']=1-quality_lumps['Moisture']/100
    quality_lumps.head()

    for i in quality_lumps['Miner']:
        mines.loc[mines['Mine']==i,'Fe_lumps']=quality_lumps.loc[quality_lumps['Miner']==i,'Fe'].item()  
        mines.loc[mines['Mine']==i,'Moisture_lumps']=quality_lumps.loc[quality_lumps['Miner']==i,'Moisture'].item()
        mines.loc[mines['Mine']==i,'SiO2_lumps']=quality_lumps.loc[quality_lumps['Miner']==i,'SiO2'].item()
        mines.loc[mines['Mine']==i,'Al2O3_lumps']=quality_lumps.loc[quality_lumps['Miner']==i,'Al2O3'].item()
        mines.loc[mines['Mine']==i,'1-loi_lumps']=quality_lumps.loc[quality_lumps['Miner']==i,'1-loi'].item()
        mines.loc[mines['Mine']==i,'1-moi_lumps']=quality_lumps.loc[quality_lumps['Miner']==i,'1-moi'].item()
        mines.loc[mines['Mine']==i,'CaO_lumps']=quality_lumps.loc[quality_lumps['Miner']==i,'CaO'].item()
        mines.loc[mines['Mine']==i,'MgO_lumps']=quality_lumps.loc[quality_lumps['Miner']==i,'MgO'].item()


    # In[5]:


    #separating lumps
    mines_lumps=mines[mines.Plant!='No Value'] #can't copy directly
    mine=list()
    for row in mines.iloc[:,7]:
        mine.append(row)
    var_lumps=LpVariable.dicts("Lumps Quantity from",mine,0,None,LpContinuous)


    # In[6]:


    #disintegrating Barbil 
    mines_bf_P1=mines[mines.Plant=='Barbil']
    mines_bf_P2=mines[mines.Plant=='Barbil']
    mines_dri_P1=mines[mines.Plant=='Barbil']
    mines_dri_P2=mines[mines.Plant=='Barbil']
    mines_pellet_market=mines[mines.Plant=='Barbil']
    mines=mines[mines.Plant!='Barbil']


    # In[7]:


    mine=list()
    for row in mines.iloc[:,7]:
        mine.append(row)
    var_fines=LpVariable.dicts("Sinter Fines Quantity from",mine,0,None,LpContinuous)
    model=LpProblem('Contribution maximization',LpMaximize)

    var_fines_bf_P1=LpVariable.dicts("BF(P1) Fines Quantity from ",mines_bf_P1['possibilities'],0,None,LpContinuous)
    var_fines_bf_P2=LpVariable.dicts("BF(P2) Fines Quantity from ",mines_bf_P2['possibilities'],0,None,LpContinuous)
    var_fines_dri_P1=LpVariable.dicts("DRI(P1) Fines Quantity from ",mines_dri_P1['possibilities'],0,None,LpContinuous)
    var_fines_dri_P2=LpVariable.dicts("DRI(Raiagrh) Fines Quantity from ",mines_dri_P2['possibilities'],0,None,LpContinuous)
    var_fines_pellet_market=LpVariable.dicts("Pellet to Market Fines Quantity from ",mines_pellet_market['possibilities'],0,None,LpContinuous)


    # In[8]:


    #updating fines quality parameters(new values) sinter
    mines.loc[mines['1-moi_fines'].isna(),'1-moi_fines']=0
    mines.loc[mines['1-loi_fines'].isna(),'1-loi_fines']=0
    mines.loc[mines['Fe_fines'].isna(),'Fe_fines']=0
    mines.loc[mines['SiO2_fines'].isna(),'SiO2_fines']=0
    mines.loc[mines['Al2O3_fines'].isna(),'Al2O3_fines']=0
    mines.loc[mines['CaO_fines'].isna(),'CaO_fines']=0
    mines.loc[mines['MgO_fines'].isna(),'MgO_fines']=0


    residue_fines_list=list()
    for i in mines['possibilities']:
        x=mines.loc[mines['possibilities']==i,'1-loi_fines'].item()*mines.loc[mines['possibilities']==i,'1-moi_fines'].item()*var_fines[i]
        residue_fines_list.append(x)
    mines['Residue_fines']=residue_fines_list


    fe_new_fines_list=list()
    for i in mines['possibilities']:
        x=mines.loc[mines['possibilities']==i,'Fe_fines'].item()*mines.loc[mines['possibilities']==i,'1-moi_fines'].item()*var_fines[i]/100
        fe_new_fines_list.append(x)
    mines['Fe(new)_fines']=fe_new_fines_list

    SiO2_new_fines_list=list()
    for i in mines['possibilities']:
        x=mines.loc[mines['possibilities']==i,'SiO2_fines'].item()*mines.loc[mines['possibilities']==i,'1-moi_fines'].item()*var_fines[i]/100
        SiO2_new_fines_list.append(x)
    mines['SiO2(new)_fines']=SiO2_new_fines_list

    Al2O3_new_fines_list=list()
    for i in mines['possibilities']:
        x=mines.loc[mines['possibilities']==i,'Al2O3_fines'].item()*mines.loc[mines['possibilities']==i,'1-moi_fines'].item()*var_fines[i]/100
        Al2O3_new_fines_list.append(x)
    mines['Al2O3(new)_fines']=Al2O3_new_fines_list

    CaO_new_fines_list=list()
    for i in mines['possibilities']:
        x=mines.loc[mines['possibilities']==i,'CaO_fines'].item()*mines.loc[mines['possibilities']==i,'1-moi_fines'].item()*var_fines[i]/100
        CaO_new_fines_list.append(x)
    mines['CaO(new)_fines']=CaO_new_fines_list

    MgO_new_fines_list=list()
    for i in mines['possibilities']:
        x=mines.loc[mines['possibilities']==i,'MgO_fines'].item()*mines.loc[mines['possibilities']==i,'1-moi_fines'].item()*var_fines[i]/100
        MgO_new_fines_list.append(x)
    mines['MgO(new)_fines']=MgO_new_fines_list
        #  mines['Fe(new)_fines']=mines['SiO2_fines']*mines['1-moi_fines']*var_fines[i]    


    # In[9]:
    pellet_converison_df=pd.read_csv('Data Input/Head Office/Iron ore/Pellet Conversion Factors.csv')
    pellet_converison_df

    #updating fines quality parameters(new values) BF P1
    pellet_conversion_bf_P1=pellet_converison_df.iloc[0,3].item()
    mines_bf_P1.loc[mines_bf_P1['1-moi_fines'].isna(),'1-moi_fines']=0
    mines_bf_P1.loc[mines_bf_P1['1-loi_fines'].isna(),'1-loi_fines']=0
    mines_bf_P1.loc[mines_bf_P1['Fe_fines'].isna(),'Fe_fines']=0
    mines_bf_P1.loc[mines_bf_P1['SiO2_fines'].isna(),'SiO2_fines']=0
    mines_bf_P1.loc[mines_bf_P1['Al2O3_fines'].isna(),'Al2O3_fines']=0
    mines_bf_P1.loc[mines_bf_P1['CaO_fines'].isna(),'CaO_fines']=0
    mines_bf_P1.loc[mines_bf_P1['MgO_fines'].isna(),'MgO_fines']=0
    mines_bf_P1.loc[mines_bf_P1['Moisture_fines'].isna(),'Moisture_fines']=0

    feeding_dmt_list=list()
    for i in mines_bf_P1['possibilities']:
        x=mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Moisture_fines'].item()*var_fines_bf_P1[i]/100
        feeding_dmt_list.append(x)
    mines_bf_P1['Feeding DMT']=feeding_dmt_list

    residue_fines_list=list()
    for i in mines_bf_P1['possibilities']:
        x=mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'1-loi_fines'].item()*mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'1-moi_fines'].item()*var_fines_bf_P1[i]/pellet_conversion_bf_P1
        residue_fines_list.append(x)
    mines_bf_P1['Residue_fines']=residue_fines_list


    fe_new_fines_list=list()
    for i in mines_bf_P1['possibilities']:
        x=((mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Fe_fines'].item()*mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'1-moi_fines'].item()*var_fines_bf_P1[i]/100)/pellet_conversion_bf_P1)
        fe_new_fines_list.append(x)
    mines_bf_P1['Fe(new)_fines']=fe_new_fines_list

    Al2O3_new_fines_list=list()
    for i in mines_bf_P1['possibilities']:
        x=((mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Al2O3_fines'].item()*mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'1-moi_fines'].item()*var_fines_bf_P1[i]/100)/pellet_conversion_bf_P1)
        Al2O3_new_fines_list.append(x)
    mines_bf_P1['Al2O3(new)_fines']=Al2O3_new_fines_list

    SiO2_new_fines_list=list()
    for i in mines_bf_P1['possibilities']:
        x=((mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'SiO2_fines'].item()*mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'1-moi_fines'].item()*var_fines_bf_P1[i]/100)/pellet_conversion_bf_P1)
        SiO2_new_fines_list.append(x)
    mines_bf_P1['SiO2(new)_fines']=SiO2_new_fines_list

    CaO_new_fines_list=list()
    for i in mines_bf_P1['possibilities']:
        x=((mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'CaO_fines'].item()*mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'1-moi_fines'].item()*var_fines_bf_P1[i]/100)/pellet_conversion_bf_P1)
        CaO_new_fines_list.append(x)
    mines_bf_P1['CaO(new)_fines']=CaO_new_fines_list

    MgO_new_fines_list=list()
    for i in mines_bf_P1['possibilities']:
        x=((mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'MgO_fines'].item()*mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'1-moi_fines'].item()*var_fines_bf_P1[i]/100)/pellet_conversion_bf_P1)
        MgO_new_fines_list.append(x)
    mines_bf_P1['MgO(new)_fines']=MgO_new_fines_list

    #  mines_bf_P1['Fe(new)_fines']=mines_bf_P1['SiO2_fines']*mines_bf_P1['1-moi_fines']*var_fines_bf_P1[i]    
    # In[10]:


    #updating fines quality parameters(new values) BF P2
    pellet_conversion_bf_P2=pellet_converison_df.iloc[1,3].item()
    mines_bf_P2.loc[mines_bf_P2['1-moi_fines'].isna(),'1-moi_fines']=0
    mines_bf_P2.loc[mines_bf_P2['1-loi_fines'].isna(),'1-loi_fines']=0
    mines_bf_P2.loc[mines_bf_P2['Fe_fines'].isna(),'Fe_fines']=0
    mines_bf_P2.loc[mines_bf_P2['SiO2_fines'].isna(),'SiO2_fines']=0
    mines_bf_P2.loc[mines_bf_P2['Al2O3_fines'].isna(),'Al2O3_fines']=0
    mines_bf_P2.loc[mines_bf_P2['CaO_fines'].isna(),'CaO_fines']=0
    mines_bf_P2.loc[mines_bf_P2['MgO_fines'].isna(),'MgO_fines']=0
    mines_bf_P2.loc[mines_bf_P2['Moisture_fines'].isna(),'Moisture_fines']=0

    feeding_dmt_list=list()
    for i in mines_bf_P2['possibilities']:
        x=mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Moisture_fines'].item()*var_fines_bf_P2[i]/100
        feeding_dmt_list.append(x)
    mines_bf_P2['Feeding DMT']=feeding_dmt_list

    residue_fines_list=list()
    for i in mines_bf_P2['possibilities']:
        x=mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'1-loi_fines'].item()*mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'1-moi_fines'].item()*var_fines_bf_P2[i]/pellet_conversion_bf_P2
        residue_fines_list.append(x)
    mines_bf_P2['Residue_fines']=residue_fines_list


    fe_new_fines_list=list()
    for i in mines_bf_P2['possibilities']:
        x=((mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Fe_fines'].item()*mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'1-moi_fines'].item()*var_fines_bf_P2[i]/100)/pellet_conversion_bf_P2)
        fe_new_fines_list.append(x)
    mines_bf_P2['Fe(new)_fines']=fe_new_fines_list

    SiO2_new_fines_list=list()
    for i in mines_bf_P2['possibilities']:
        x=((mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'SiO2_fines'].item()*mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'1-moi_fines'].item()*var_fines_bf_P2[i]/100)/pellet_conversion_bf_P2)
        SiO2_new_fines_list.append(x)
    mines_bf_P2['SiO2(new)_fines']=SiO2_new_fines_list

    Al2O3_new_fines_list=list()
    for i in mines_bf_P2['possibilities']:
        x=((mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Al2O3_fines'].item()*mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'1-moi_fines'].item()*var_fines_bf_P2[i]/100)/pellet_conversion_bf_P2)
        Al2O3_new_fines_list.append(x)
    mines_bf_P2['Al2O3(new)_fines']=Al2O3_new_fines_list


    CaO_new_fines_list=list()
    for i in mines_bf_P2['possibilities']:
        x=((mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'CaO_fines'].item()*mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'1-moi_fines'].item()*var_fines_bf_P2[i]/100)/pellet_conversion_bf_P2)
        CaO_new_fines_list.append(x)
    mines_bf_P2['CaO(new)_fines']=CaO_new_fines_list

    MgO_new_fines_list=list()
    for i in mines_bf_P2['possibilities']:
        x=((mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'MgO_fines'].item()*mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'1-moi_fines'].item()*var_fines_bf_P2[i]/100)/pellet_conversion_bf_P2)
        MgO_new_fines_list.append(x)
    mines_bf_P2['MgO(new)_fines']=MgO_new_fines_list
    #  mines_bf_P2['Fe(new)_fines']=mines_bf_P2['SiO2_fines']*mines_bf_P2['1-moi_fines']*var_fines_bf_P2[i]    

    # In[11]:


    #updating fines quality parameters(new values) DRI P1
    pellet_conversion_dri_P1=pellet_converison_df.iloc[2,3].item()
    mines_dri_P1.loc[mines_dri_P1['1-moi_fines'].isna(),'1-moi_fines']=0
    mines_dri_P1.loc[mines_dri_P1['1-loi_fines'].isna(),'1-loi_fines']=0
    mines_dri_P1.loc[mines_dri_P1['Fe_fines'].isna(),'Fe_fines']=0
    mines_dri_P1.loc[mines_dri_P1['SiO2_fines'].isna(),'SiO2_fines']=0
    mines_dri_P1.loc[mines_dri_P1['Al2O3_fines'].isna(),'Al2O3_fines']=0
    mines_dri_P1.loc[mines_dri_P1['CaO_fines'].isna(),'CaO_fines']=0
    mines_dri_P1.loc[mines_dri_P1['MgO_fines'].isna(),'MgO_fines']=0
    mines_dri_P1.loc[mines_dri_P1['Moisture_fines'].isna(),'Moisture_fines']=0

    feeding_dmt_list=list()
    for i in mines_dri_P1['possibilities']:
        x=mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Moisture_fines'].item()*var_fines_dri_P1[i]/100
        feeding_dmt_list.append(x)
    mines_dri_P1['Feeding DMT']=feeding_dmt_list

    residue_fines_list=list()
    for i in mines_dri_P1['possibilities']:
        x=mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'1-loi_fines'].item()*mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'1-moi_fines'].item()*var_fines_dri_P1[i]/pellet_conversion_dri_P1
        residue_fines_list.append(x)
    mines_dri_P1['Residue_fines']=residue_fines_list


    fe_new_fines_list=list()
    for i in mines_dri_P1['possibilities']:
        x=((mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Fe_fines'].item()*mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'1-moi_fines'].item()*var_fines_dri_P1[i]/100)/pellet_conversion_dri_P1)
        fe_new_fines_list.append(x)
    mines_dri_P1['Fe(new)_fines']=fe_new_fines_list

    SiO2_new_fines_list=list()
    for i in mines_dri_P1['possibilities']:
        x=((mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'SiO2_fines'].item()*mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'1-moi_fines'].item()*var_fines_dri_P1[i]/100)/pellet_conversion_dri_P1)
        SiO2_new_fines_list.append(x)
    mines_dri_P1['SiO2(new)_fines']=SiO2_new_fines_list

    Al2O3_new_fines_list=list()
    for i in mines_dri_P1['possibilities']:
        x=((mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Al2O3_fines'].item()*mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'1-moi_fines'].item()*var_fines_dri_P1[i]/100)/pellet_conversion_dri_P1)
        Al2O3_new_fines_list.append(x)
    mines_dri_P1['Al2O3(new)_fines']=Al2O3_new_fines_list


    CaO_new_fines_list=list()
    for i in mines_dri_P1['possibilities']:
        x=((mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'CaO_fines'].item()*mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'1-moi_fines'].item()*var_fines_dri_P1[i]/100)/pellet_conversion_dri_P1)
        CaO_new_fines_list.append(x)
    mines_dri_P1['CaO(new)_fines']=CaO_new_fines_list

    MgO_new_fines_list=list()
    for i in mines_dri_P1['possibilities']:
        x=((mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'MgO_fines'].item()*mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'1-moi_fines'].item()*var_fines_dri_P1[i]/100)/pellet_conversion_dri_P1)
        MgO_new_fines_list.append(x)
    mines_dri_P1['MgO(new)_fines']=MgO_new_fines_list

    #  mines_dri_P1['Fe(new)_fines']=mines_dri_P1['SiO2_fines']*mines_dri_P1['1-moi_fines']*var_fines_dri_P1[i]    
    #  mines_dri_P1['Fe(new)_fines']=mines_dri_P1['SiO2_fines']*mines_dri_P1['1-moi_fines']*var_fines_dri_P1[i]    


    # In[12]:


    #updating fines quality parameters(new values) DRI P2
    pellet_conversion_dri_P2=pellet_converison_df.iloc[3,3].item()
    mines_dri_P2.loc[mines_dri_P2['1-moi_fines'].isna(),'1-moi_fines']=0
    mines_dri_P2.loc[mines_dri_P2['1-loi_fines'].isna(),'1-loi_fines']=0
    mines_dri_P2.loc[mines_dri_P2['Fe_fines'].isna(),'Fe_fines']=0
    mines_dri_P2.loc[mines_dri_P2['SiO2_fines'].isna(),'SiO2_fines']=0
    mines_dri_P2.loc[mines_dri_P2['Al2O3_fines'].isna(),'Al2O3_fines']=0
    mines_dri_P2.loc[mines_dri_P2['CaO_fines'].isna(),'CaO_fines']=0
    mines_dri_P2.loc[mines_dri_P2['MgO_fines'].isna(),'MgO_fines']=0
    mines_dri_P2.loc[mines_dri_P2['Moisture_fines'].isna(),'Moisture_fines']=0

    feeding_dmt_list=list()
    for i in mines_dri_P2['possibilities']:
        x=mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Moisture_fines'].item()*var_fines_dri_P2[i]/100
        feeding_dmt_list.append(x)
    mines_dri_P2['Feeding DMT']=feeding_dmt_list

    residue_fines_list=list()
    for i in mines_dri_P2['possibilities']:
        x=mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'1-loi_fines'].item()*mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'1-moi_fines'].item()*var_fines_dri_P2[i]/pellet_conversion_dri_P2
        residue_fines_list.append(x)
    mines_dri_P2['Residue_fines']=residue_fines_list


    fe_new_fines_list=list()
    for i in mines_dri_P2['possibilities']:
        x=((mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Fe_fines'].item()*mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'1-moi_fines'].item()*var_fines_dri_P2[i]/100)/pellet_conversion_dri_P2)
        fe_new_fines_list.append(x)
    mines_dri_P2['Fe(new)_fines']=fe_new_fines_list

    SiO2_new_fines_list=list()
    for i in mines_dri_P2['possibilities']:
        x=((mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'SiO2_fines'].item()*mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'1-moi_fines'].item()*var_fines_dri_P2[i]/100)/pellet_conversion_dri_P2)
        SiO2_new_fines_list.append(x)
    mines_dri_P2['SiO2(new)_fines']=SiO2_new_fines_list

    Al2O3_new_fines_list=list()
    for i in mines_dri_P2['possibilities']:
        x=((mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Al2O3_fines'].item()*mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'1-moi_fines'].item()*var_fines_dri_P2[i]/100)/pellet_conversion_dri_P2)
        Al2O3_new_fines_list.append(x)
    mines_dri_P2['Al2O3(new)_fines']=Al2O3_new_fines_list


    CaO_new_fines_list=list()
    for i in mines_dri_P2['possibilities']:
        x=((mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'CaO_fines'].item()*mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'1-moi_fines'].item()*var_fines_dri_P2[i]/100)/pellet_conversion_dri_P2)
        CaO_new_fines_list.append(x)
    mines_dri_P2['CaO(new)_fines']=CaO_new_fines_list

    MgO_new_fines_list=list()
    for i in mines_dri_P2['possibilities']:
        x=((mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'MgO_fines'].item()*mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'1-moi_fines'].item()*var_fines_dri_P2[i]/100)/pellet_conversion_dri_P2)
        MgO_new_fines_list.append(x)
    mines_dri_P2['MgO(new)_fines']=MgO_new_fines_list

    #  mines_dri_P2['Fe(new)_fines']=mines_dri_P2['SiO2_fines']*mines_dri_P2['1-moi_fines']*var_fines_dri_P2[i]    
    # In[13]:


    #updating lumps quality parameters(new values)
    mines_lumps.loc[mines_lumps['1-moi_lumps'].isna(),'1-moi_lumps']=0
    mines_lumps.loc[mines_lumps['1-loi_lumps'].isna(),'1-loi_lumps']=0
    mines_lumps.loc[mines_lumps['Fe_lumps'].isna(),'Fe_lumps']=0
    mines_lumps.loc[mines_lumps['SiO2_lumps'].isna(),'SiO2_lumps']=0
    mines_lumps.loc[mines_lumps['Al2O3_lumps'].isna(),'Al2O3_lumps']=0
    mines_lumps.loc[mines_lumps['CaO_lumps'].isna(),'CaO_lumps']=0
    mines_lumps.loc[mines_lumps['MgO_lumps'].isna(),'MgO_lumps']=0



    residue_lumps_list=list()
    for i in mines_lumps['possibilities']:
        x=mines_lumps.loc[mines_lumps['possibilities']==i,'1-loi_lumps'].item()*mines_lumps.loc[mines_lumps['possibilities']==i,'1-moi_lumps'].item()*var_lumps[i]
        residue_lumps_list.append(x)
    mines_lumps['Residue_lumps']=residue_lumps_list


    fe_new_lumps_list=list()
    for i in mines_lumps['possibilities']:
        x=mines_lumps.loc[mines_lumps['possibilities']==i,'Fe_lumps'].item()*mines_lumps.loc[mines_lumps['possibilities']==i,'1-moi_lumps'].item()*var_lumps[i]/100
        fe_new_lumps_list.append(x)
    mines_lumps['Fe(new)_lumps']=fe_new_lumps_list

    SiO2_new_lumps_list=list()
    for i in mines_lumps['possibilities']:
        x=mines_lumps.loc[mines_lumps['possibilities']==i,'SiO2_lumps'].item()*mines_lumps.loc[mines_lumps['possibilities']==i,'1-moi_lumps'].item()*var_lumps[i]/100
        SiO2_new_lumps_list.append(x)
    mines_lumps['SiO2(new)_lumps']=SiO2_new_lumps_list

    Al2O3_new_lumps_list=list()
    for i in mines_lumps['possibilities']:
        x=mines_lumps.loc[mines_lumps['possibilities']==i,'Al2O3_lumps'].item()*mines_lumps.loc[mines_lumps['possibilities']==i,'1-moi_lumps'].item()*var_lumps[i]/100
        Al2O3_new_lumps_list.append(x)
    mines_lumps['Al2O3(new)_lumps']=Al2O3_new_lumps_list

    CaO_new_lumps_list=list()
    for i in mines_lumps['possibilities']:
        x=mines_lumps.loc[mines_lumps['possibilities']==i,'CaO_lumps'].item()*mines_lumps.loc[mines_lumps['possibilities']==i,'1-moi_lumps'].item()*var_lumps[i]/100
        CaO_new_lumps_list.append(x)
    mines_lumps['CaO(new)_lumps']=CaO_new_lumps_list

    MgO_new_lumps_list=list()
    for i in mines_lumps['possibilities']:
        x=mines_lumps.loc[mines_lumps['possibilities']==i,'MgO_lumps'].item()*mines_lumps.loc[mines_lumps['possibilities']==i,'1-moi_lumps'].item()*var_lumps[i]/100
        MgO_new_lumps_list.append(x)
    mines_lumps['MgO(new)_lumps']=MgO_new_lumps_list



        #  mines_lumps['Fe(new)_lumps']=mines_lumps['SiO2_lumps']*mines_lumps['1-moi_lumps']*var_lumps[i]    


    # In[14]:


   #updating fines quality parameters(new values) Market
    pellet_conversion_market=pellet_converison_df.iloc[4,3].item()
    mines_pellet_market.loc[mines_pellet_market['1-moi_fines'].isna(),'1-moi_fines']=0
    mines_pellet_market.loc[mines_pellet_market['1-loi_fines'].isna(),'1-loi_fines']=0
    mines_pellet_market.loc[mines_pellet_market['Fe_fines'].isna(),'Fe_fines']=0
    mines_pellet_market.loc[mines_pellet_market['SiO2_fines'].isna(),'SiO2_fines']=0
    mines_pellet_market.loc[mines_pellet_market['Al2O3_fines'].isna(),'Al2O3_fines']=0
    mines_pellet_market.loc[mines_pellet_market['CaO_fines'].isna(),'CaO_fines']=0
    mines_pellet_market.loc[mines_pellet_market['MgO_fines'].isna(),'MgO_fines']=0
    mines_pellet_market.loc[mines_pellet_market['Moisture_fines'].isna(),'Moisture_fines']=0

    feeding_dmt_list=list()
    for i in mines_pellet_market['possibilities']:
        x=mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Moisture_fines'].item()*var_fines_pellet_market[i]/100
        feeding_dmt_list.append(x)
    mines_pellet_market['Feeding DMT']=feeding_dmt_list

    residue_fines_list=list()
    for i in mines_pellet_market['possibilities']:
        x=mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'1-loi_fines'].item()*mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'1-moi_fines'].item()*var_fines_pellet_market[i]/pellet_conversion_market
        residue_fines_list.append(x)
    mines_pellet_market['Residue_fines']=residue_fines_list


    fe_new_fines_list=list()
    for i in mines_pellet_market['possibilities']:
        x=((mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Fe_fines'].item()*mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'1-moi_fines'].item()*var_fines_pellet_market[i]/100)/pellet_conversion_market)
        fe_new_fines_list.append(x)
    mines_pellet_market['Fe(new)_fines']=fe_new_fines_list

    SiO2_new_fines_list=list()
    for i in mines_pellet_market['possibilities']:
        x=((mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'SiO2_fines'].item()*mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'1-moi_fines'].item()*var_fines_pellet_market[i]/100)/pellet_conversion_market)
        SiO2_new_fines_list.append(x)
    mines_pellet_market['SiO2(new)_fines']=SiO2_new_fines_list

    Al2O3_new_fines_list=list()
    for i in mines_pellet_market['possibilities']:
        x=((mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Al2O3_fines'].item()*mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'1-moi_fines'].item()*var_fines_pellet_market[i]/100)/pellet_conversion_market)
        Al2O3_new_fines_list.append(x)
    mines_pellet_market['Al2O3(new)_fines']=Al2O3_new_fines_list


    CaO_new_fines_list=list()
    for i in mines_pellet_market['possibilities']:
        x=((mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'CaO_fines'].item()*mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'1-moi_fines'].item()*var_fines_pellet_market[i]/100)/pellet_conversion_market)
        CaO_new_fines_list.append(x)
    mines_pellet_market['CaO(new)_fines']=CaO_new_fines_list

    MgO_new_fines_list=list()
    for i in mines_pellet_market['possibilities']:
        x=((mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'MgO_fines'].item()*mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'1-moi_fines'].item()*var_fines_pellet_market[i]/100)/pellet_conversion_market)
        MgO_new_fines_list.append(x)
    mines_pellet_market['MgO(new)_fines']=MgO_new_fines_list


    #  mines_pellet_market['Fe(new)_fines']=mines_pellet_market['SiO2_fines']*mines_pellet_market['1-moi_fines']*var_fines_pellet_market[i]    
    # In[15]:

    #Pellet DRI Integration
    #Barbil to bf
    pellet_freight=pd.read_csv('Data Input/Head Office/Iron ore/Pellet Freight.csv')

    mines_bf_P1['Serving Plant']='P1'
    mines_bf_P1['Freight 2']=pellet_freight.loc[pellet_freight['Serving Plant']=='P1','Freight2'].item()
    mines_bf_P1['Fixed']=pellet_freight.loc[pellet_freight['Serving Plant']=='P1','Fixed'].item()
    mines_bf_P1['Final Cost Fines']=mines_bf_P1['Freight 2']+mines_bf_P1['Fixed']+mines_bf_P1['Total Cost Fines']
    mines_bf_P2['Serving Plant']='P2'
    mines_bf_P2['Freight 2']=pellet_freight.loc[pellet_freight['Serving Plant']=='P2','Freight2'].item()
    mines_bf_P2['Fixed']=pellet_freight.loc[pellet_freight['Serving Plant']=='P2','Fixed'].item()
    mines_bf_P2['Final Cost Fines']=mines_bf_P2['Freight 2']+mines_bf_P2['Fixed']+mines_bf_P2['Total Cost Fines']
    #Barbil to DRI
    mines_dri_P1['Serving Plant']='P1'
    mines_dri_P1['Freight 2']=pellet_freight.loc[pellet_freight['Serving Plant']=='P1','Freight2'].item()
    mines_dri_P1['Fixed']=pellet_freight.loc[pellet_freight['Serving Plant']=='P1','Fixed'].item()
    mines_dri_P1['Final Cost Fines']=mines_dri_P1['Freight 2']+mines_dri_P1['Fixed']+mines_dri_P1['Total Cost Fines']
    mines_dri_P2['Serving Plant']='P2'
    mines_dri_P2['Freight 2']=pellet_freight.loc[pellet_freight['Serving Plant']=='P2','Freight2'].item()
    mines_dri_P2['Fixed']=pellet_freight.loc[pellet_freight['Serving Plant']=='P2','Fixed'].item()
    mines_dri_P2['Final Cost Fines']=mines_dri_P2['Freight 2']+mines_dri_P2['Fixed']+mines_dri_P2['Total Cost Fines']

    # In[ ]:





    # In[16]:


    zxc=list()
    #Constraints sinter

    #Supply Constraint

    #Supply Fines Constraint
    for j in supply_constraint_table['Mines']:
        a=list()
        b=list()
        for i in var_fines:
            if mines.loc[mines['possibilities']==i,'Mine'].item()==j:
                a.append(var_fines[i])
        for i in var_fines_bf_P1:
            if mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Mine'].item()==j:
                a.append(var_fines_bf_P1[i])
        for i in var_fines_bf_P2:
            if mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Mine'].item()==j:
                a.append(var_fines_bf_P2[i])
        for i in var_fines_dri_P1:
            if mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Mine'].item()==j:
                a.append(var_fines_dri_P1[i])
        for i in var_fines_dri_P2:
            if mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Mine'].item()==j:
                a.append(var_fines_dri_P2[i])
        for i in var_fines_pellet_market:
            if mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Mine'].item()==j:
                a.append(var_fines_pellet_market[i])
        b=lpSum(a)
        supply_constraint_fines=b<=supply_constraint_table.loc[supply_constraint_table['Mines']==j,'Fines']*(no_of_days_input/no_of_days_in_month)
        if io_constraint_dict['Supply Fines'].get()==1:
            model+=supply_constraint_fines


    #Supply Lumps Constraint
    for j in supply_constraint_table['Mines']:
        a=list()
        b=list()
        for i in var_lumps:
            if mines_lumps.loc[mines_lumps['possibilities']==i,'Mine'].item()==j:
                a.append(var_lumps[i])
        b=lpSum(a)
        supply_constraint_lumps=b<=supply_constraint_table.loc[supply_constraint_table['Mines']==j,'Lumps']*(no_of_days_input/no_of_days_in_month)
        if io_constraint_dict['Supply Lumps'].get()==1:
            model+=supply_constraint_lumps    


    # In[33]:


    #Demand Constraint

    p=list()
    #demand_constraint_fines
    #for j in demand_constraint_table['Plant']:
    a=list()
    b=list()
    for i in var_fines:
        if mines.loc[mines['possibilities']==i,'Plant'].item()=='Barbil':
            a.append(var_fines[i])
    b=lpSum(a)
    #model+=b>=demand_constraint_table.loc[demand_constraint_table['Plant']=='Barbil','Fines']

    #demand_constraint_lumps
    for j in demand_constraint_table['Plant']:
        a=list()
        b=list()
        for i in var_lumps:
            if mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()==j:
                a.append(var_lumps[i])
        b=lpSum(a)
    #    model+=b>=demand_constraint_table.loc[demand_constraint_table['Plant']==j,'Lumps']


    # In[34]:


    #Siding Constraints

    for j in siding_constraint_table['Siding']:
        a=list()
        b=list()
        for i in var_fines:
            if mines.loc[mines['possibilities']==i,'Siding'].item()==j:
                a.append(var_fines[i])
        for i in var_fines_bf_P1:
            if mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Siding'].item()==j:
                a.append(var_fines_bf_P1[i])
        for i in var_fines_bf_P2:
            if mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Siding'].item()==j:
                a.append(var_fines_bf_P2[i])
        for i in var_fines_dri_P1:
            if mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Siding'].item()==j:
                a.append(var_fines_dri_P1[i])
        for i in var_fines_dri_P2:
            if mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Siding'].item()==j:
                a.append(var_fines_dri_P2[i])
        for i in var_fines_pellet_market:
            if mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Siding'].item()==j:
                a.append(var_fines_pellet_market[i])
        for i in var_lumps:
            if mines_lumps.loc[mines_lumps['possibilities']==i,'Siding'].item()==j:
                a.append(var_lumps[i])
        b=lpSum(a)
        siding_constraint=b<=siding_constraint_table.loc[siding_constraint_table['Siding']==j,'Limit']*(no_of_days_input/no_of_days_in_month)
        if io_constraint_dict['Siding'].get()==1:
            model+=siding_constraint




    # In[35]:


    #Logistics Constraints
    #P1 & Raiagrh
    #raod
    for j in logistics_constraint_table.loc[:1,'Plant']:
        a=list()
        b=list()
        for i in var_fines:
            if mines.loc[mines['possibilities']==i,'Plant'].item()==j and (mines.loc[mines['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_fines[i])
        for i in var_lumps:
            if (mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()==j) and (mines_lumps.loc[mines_lumps['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_lumps[i])
        b=lpSum(a)
        logistics_constraint_road_P1_and_P2=b<=logistics_constraint_table.loc[logistics_constraint_table['Plant']==j,'Road']*(no_of_days_input/no_of_days_in_month)
        if io_constraint_dict['Logistics Road(P1 and P2)'].get()==1:
            model+=logistics_constraint_road_P1_and_P2


    #rail
    for j in logistics_constraint_table.loc[:1,'Plant']:
        a=list()
        b=list()
        for i in var_fines:
            if mines.loc[mines['possibilities']==i,'Plant'].item()==j and (mines.loc[mines['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_fines[i])
        for i in var_lumps:
            if (mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()==j) and (mines_lumps.loc[mines_lumps['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_lumps[i])
        b=lpSum(a)
        logistics_constraint_rail_P1_and_P2=b<=logistics_constraint_table.loc[logistics_constraint_table['Plant']==j,'Rail']*(no_of_days_input/no_of_days_in_month)
        if io_constraint_dict['Logistics Rail(P1 and P2)'].get()==1:
            model+=logistics_constraint_rail_P1_and_P2    

    #Barbil
    #raod
    for j in logistics_constraint_table.loc[2:,'Plant']:
        a=list()
        b=list()
        for i in var_fines_bf_P1:
            if mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Plant'].item()==j and (mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_fines_bf_P1[i])
        for i in var_fines_bf_P2:
            if mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Plant'].item()==j and (mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_fines_bf_P2[i])
        for i in var_fines_dri_P1:
            if mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Plant'].item()==j and (mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_fines_dri_P1[i])
        for i in var_fines_dri_P2:
            if mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Plant'].item()==j and (mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_fines_dri_P2[i])
        for i in var_fines_pellet_market:
            if mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Plant'].item()==j and (mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_fines_pellet_market[i]) 
        for i in var_lumps:
            if (mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()==j) and (mines_lumps.loc[mines_lumps['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_lumps[i])
        b=lpSum(a)
        logistics_constraint_road_barbil=b<=logistics_constraint_table.loc[logistics_constraint_table['Plant']==j,'Road']*(no_of_days_input/no_of_days_in_month)
        if io_constraint_dict['Logistics Road(Barbil)'].get()==1:
            model+=logistics_constraint_road_barbil


    #rail
    for j in logistics_constraint_table.loc[2:,'Plant']:
        a=list()
        b=list()
        for i in var_fines_bf_P1:
            if mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Plant'].item()==j and (mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_fines_bf_P1[i])
        for i in var_fines_bf_P2:
            if mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Plant'].item()==j and (mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_fines_bf_P2[i])
        for i in var_fines_dri_P1:
            if mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Plant'].item()==j and (mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_fines_dri_P1[i])
        for i in var_fines_dri_P2:
            if mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Plant'].item()==j and (mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_fines_dri_P2[i])
        for i in var_fines_pellet_market:
            if mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Plant'].item()==j and (mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_fines_pellet_market[i]) 
        for i in var_lumps:
            if (mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()==j) and (mines_lumps.loc[mines_lumps['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_lumps[i])
        b=lpSum(a)
        logistics_constraint_rail_barbil=b<=logistics_constraint_table.loc[logistics_constraint_table['Plant']==j,'Rail']*(no_of_days_input/no_of_days_in_month)
        if io_constraint_dict['Logistics Rail(Barbil)'].get()==1:
            model+=logistics_constraint_rail_barbil    


    #ltc constraints
    key=[0,0,1,1]
    ltc_constraint_table['key']=key

    for j in ltc_constraint_table['Mine Owners']:
        a=list()
        b=list()
        for i in var_fines:
            if mines.loc[mines['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_fines[i])
        for i in var_fines_bf_P1:
            if mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_fines_bf_P1[i])
        for i in var_fines_bf_P2:
            if mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_fines_bf_P2[i])
        for i in var_fines_dri_P1:
            if mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_fines_dri_P1[i])
        for i in var_fines_dri_P2:
            if mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_fines_dri_P2[i])
        for i in var_fines_pellet_market:
            if mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_fines_pellet_market[i])
        for i in var_lumps:
            if mines_lumps.loc[mines_lumps['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_lumps[i])
        b=lpSum(a)
        ltc_constraint_lower=b>=ltc_constraint_table.loc[ltc_constraint_table['Mine Owners']==j,'Min']*(no_of_days_input/no_of_days_in_month)
        if io_constraint_dict['LTC Lower'].get()==1:
            model+=ltc_constraint_lower

    for j in ltc_constraint_table['Mine Owners']:
        if ltc_constraint_table.loc[ltc_constraint_table['Mine Owners']==j,'key'].item()==1:
            a=list()
            b=list()
            for i in var_fines:
                if mines.loc[mines['possibilities']==i,'Mine Owners'].item()==j:
                    a.append(var_fines[i])
            for i in var_fines_bf_P1:
                if mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Mine Owners'].item()==j:
                    a.append(var_fines_bf_P1[i])
            for i in var_fines_bf_P2:
                if mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Mine Owners'].item()==j:
                    a.append(var_fines_bf_P2[i])
            for i in var_fines_dri_P1:
                if mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Mine Owners'].item()==j:
                    a.append(var_fines_dri_P1[i])
            for i in var_fines_dri_P2:
                if mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Mine Owners'].item()==j:
                    a.append(var_fines_dri_P2[i])
            for i in var_fines_pellet_market:
                if mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Mine Owners'].item()==j:
                    a.append(var_fines_pellet_market[i])
            for i in var_lumps:
                if mines_lumps.loc[mines_lumps['possibilities']==i,'Mine Owners'].item()==j:
                    a.append(var_lumps[i])
            b=lpSum(a)
            ltc_constraint_upper=b<=ltc_constraint_table.loc[ltc_constraint_table['Mine Owners']==j,'Max']*(no_of_days_input/no_of_days_in_month)
            if io_constraint_dict['LTC Upper'].get()==1:
                model+=ltc_constraint_upper



    # In[17]:


    materials=pd.read_csv('Data Input/Plant/Sinter/Material.csv')



    #no_of_days_planned=30
    no_of_days_planned=no_of_days_input
    no_of_days_in_month=30


    #
    #updating Iron ore purchased values
    #a=list()
    #a.append(residue_fines_P1)
    #materials.loc[materials['Material']=='Iron Ore  Purchased','Residue']=a
    #b=list()
    #b.append(MgO_fines_P1)
    #materials.loc[materials['Material']=='Iron Ore  Purchased','MgO(new)']=b
    #c=list()
    #c.append(CaO_fines_P1)
    #materials.loc[materials['Material']=='Iron Ore  Purchased','CaO(new)']=c
    #d=list()
    #d.append(Al2O3_fines_P1)
    #materials.loc[materials['Material']=='Iron Ore  Purchased','Al2O3(new)']=d
    #e=list()
    #e.append(SiO2_fines_P1)
    #materials.loc[materials['Material']=='Iron Ore  Purchased','SiO2(new)']=e
    #f=list()
    #f.append(fe_fines_P1)
    #materials.loc[materials['Material']=='Iron Ore  Purchased','Fe(new)']=f
    materials=materials.drop([0,2])
    materials=materials.drop([1])
    materials=materials.drop([16,17,18,19])

    material=list()
    for i in materials['Material']:
        material.append(i)

    #materials_variables_P1
    var_P1=LpVariable.dicts('Material quantity in Tons P1',material,0,None,LpContinuous)    

    #materials_variables_P2
    var_P2=LpVariable.dicts('Material quantity in Tons P2',material,0,None,LpContinuous)    

    # In[4]:


    #calculating feeding_dmt
    #feeding_dmt=list()
    #for i in var:
     #   x=((var[i]-(var[i]*materials.loc[materials['Material']==i,'Moi'].item()/100))/no_of_days_planned)
      #  feeding_dmt.append(x)
    #materials['Feeding DMT']=feeding_dmt
    #materials.head()

    #P1
    #updating chemistry values P1
    materials['1-loi']=1-materials['LOI']/100
    materials['1-moi']=1-materials['Moi']/100
    materials
    #residue_P1
    residue_list_P1=list()
    for i in materials['Material']:
        x=materials.loc[materials['Material']==i,'1-loi'].item()*materials.loc[materials['Material']==i,'1-moi'].item()*var_P1[i]
        residue_list_P1.append(x)
    materials['Residue_P1']=residue_list_P1


    fe_new_list_P1=list()
    for i in materials['Material']:
        x=materials.loc[materials['Material']==i,'Fe'].item()*materials.loc[materials['Material']==i,'1-moi'].item()*var_P1[i]/100
        fe_new_list_P1.append(x)
    materials['Fe(new)_P1']=fe_new_list_P1

    SiO2_new_list_P1=list()
    for i in materials['Material']:
        x=materials.loc[materials['Material']==i,'SiO2'].item()*materials.loc[materials['Material']==i,'1-moi'].item()*var_P1[i]/100
        SiO2_new_list_P1.append(x)
    materials['SiO2(new)_P1']=SiO2_new_list_P1

    Al2O3_new_list_P1=list()
    for i in materials['Material']:
        x=materials.loc[materials['Material']==i,'Al2O3'].item()*materials.loc[materials['Material']==i,'1-moi'].item()*var_P1[i]/100
        Al2O3_new_list_P1.append(x)
    materials['Al2O3(new)_P1']=Al2O3_new_list_P1

    CaO_new_list_P1=list()
    for i in materials['Material']:
        x=materials.loc[materials['Material']==i,'CaO'].item()*materials.loc[materials['Material']==i,'1-moi'].item()*var_P1[i]/100
        CaO_new_list_P1.append(x)
    materials['CaO(new)_P1']=CaO_new_list_P1

    MgO_new_list_P1=list()
    for i in materials['Material']:
        x=materials.loc[materials['Material']==i,'MgO'].item()*materials.loc[materials['Material']==i,'1-moi'].item()*var_P1[i]/100
        MgO_new_list_P1.append(x)
    materials['MgO(new)_P1']=MgO_new_list_P1
    materials

    #P2
    #updating chemistry values P2
    materials['1-loi']=1-materials['LOI']/100
    materials['1-moi']=1-materials['Moi']/100
    materials
    #residue
    residue_list_P2=list()
    for i in materials['Material']:
        x=materials.loc[materials['Material']==i,'1-loi'].item()*materials.loc[materials['Material']==i,'1-moi'].item()*var_P2[i]
        residue_list_P2.append(x)
    materials['Residue_P2']=residue_list_P2


    fe_new_list_P2=list()
    for i in materials['Material']:
        x=materials.loc[materials['Material']==i,'Fe'].item()*materials.loc[materials['Material']==i,'1-moi'].item()*var_P2[i]/100
        fe_new_list_P2.append(x)
    materials['Fe(new)_P2']=fe_new_list_P2

    SiO2_new_list_P2=list()
    for i in materials['Material']:
        x=materials.loc[materials['Material']==i,'SiO2'].item()*materials.loc[materials['Material']==i,'1-moi'].item()*var_P2[i]/100
        SiO2_new_list_P2.append(x)
    materials['SiO2(new)_P2']=SiO2_new_list_P2

    Al2O3_new_list_P2=list()
    for i in materials['Material']:
        x=materials.loc[materials['Material']==i,'Al2O3'].item()*materials.loc[materials['Material']==i,'1-moi'].item()*var_P2[i]/100
        Al2O3_new_list_P2.append(x)
    materials['Al2O3(new)_P2']=Al2O3_new_list_P2

    CaO_new_list_P2=list()
    for i in materials['Material']:
        x=materials.loc[materials['Material']==i,'CaO'].item()*materials.loc[materials['Material']==i,'1-moi'].item()*var_P2[i]/100
        CaO_new_list_P2.append(x)
    materials['CaO(new)_P2']=CaO_new_list_P2

    MgO_new_list_P2=list()
    for i in materials['Material']:
        x=materials.loc[materials['Material']==i,'MgO'].item()*materials.loc[materials['Material']==i,'1-moi'].item()*var_P2[i]/100
        MgO_new_list_P2.append(x)
    materials['MgO(new)_P2']=MgO_new_list_P2


    # In[18]:


    #For P1
    #other_P1 constraints & assumptions
    other_P1=pd.read_csv('Data Input/Plant/Blast Furnace/P1/P1 Constraints & Assumptions.csv')
    RF_Internal_perc_P1=other_P1.loc[other_P1['Const&Assump.']=='RF Internal(%)','Value'].item()
    RF_BF_perc_P1=other_P1.loc[other_P1['Const&Assump.']=='RF BF(%)','Value'].item()
    
    #expected_production_constraint
    Production_target=other_P1.loc[other_P1['Const&Assump.']=='Production Target/Day','Value'].item()*no_of_days_planned
    Historical_per_day=other_P1.loc[other_P1['Const&Assump.']=='Production Target/Day','Value'].item()
    if sinter_constraint_dict_P1['Expected Production'].get()==1:
        model+=lpSum(mines.loc[mines['Plant']=='P1','Residue_fines'])+lpSum(materials['Residue_P1']) <= Production_target*(1+RF_Internal_perc_P1/100)
    #feeding rate constraint
    P1_fines_variables_list=list()
    for i in var_fines:
        if mines.loc[mines['possibilities']==i,'Plant'].item()=='P1':
            P1_fines_variables_list.append(var_fines[i])
    if sinter_constraint_dict_P1['Feeding Rate'].get()==1:
        model+=(lpSum(P1_fines_variables_list)+lpSum(var_P1)-var_P1['RF Internal'])/(no_of_days_input*24)<=other_P1.loc[other_P1['Const&Assump.']=='Feeding Rate','Value'].item()


    other_P1.loc[other_P1['Const&Assump.']=='Coke Rate - Own','Value']=25*Production_target/1000
    other_P1.loc[other_P1['Const&Assump.']=='Coke Rate - Purchase','Value']=50*Production_target/1000
    other_P1


    #capacity input constraints
    capacity_input=pd.read_csv('Data Input/Plant/Sinter/Capacity Input.csv')
    capacity_input['Monthly']=capacity_input['Monthly'].str.replace('Model','0').astype(float)
    capacity_input['key']=[5,5,5,1,5,1,1,2,2,2,2,2,1,1,2,2,1,1,5,5,5,5,0,0]


    capacity_input.loc[capacity_input['Material']=='Coke Fines-own','Monthly']=other_P1.loc[other_P1['Const&Assump.']=='Coke Rate - Own','Value'].item()
    capacity_input.loc[capacity_input['Material']=='Coke Fines-purchase','Monthly']=other_P1.loc[other_P1['Const&Assump.']=='Coke Rate - Purchase','Value'].item()

    #smo&sinter relation
    #capacity_input.loc[capacity_input['Material']=='Tensa Total','Monthly']=164542.0
    #capacity_input.loc[capacity_input['Material']=='Tensa for P1','Monthly']=capacity_input.loc[capacity_input['Material']=='Tensa Total','Monthly'].item()*0.5
    #capacity_input.loc[capacity_input['Material']=='Iron Ore Tensa','Monthly']=capacity_input.loc[capacity_input['Material']=='Tensa Total','Monthly'].item()-capacity_input.loc[capacity_input['Material']=='Tensa for P1','Monthly'].item()
    #RF Internal Input
    expected_production_P1=(lpSum(mines.loc[mines['Plant']=='P1','Residue_fines'])+lpSum(materials['Residue_P1']))/(1+RF_Internal_perc_P1/100)
    rf_internal=list()
    rf_internal.append((expected_production_P1)*(RF_Internal_perc_P1/100))
    capacity_input.loc[capacity_input['Material']=='RF Internal','Monthly']=rf_internal
    capacity_input.loc[capacity_input['Material']=='RF Internal','Monthly'].item()
    #RF BF Input
    capacity_input.loc[capacity_input['Material']=='RF BF','Monthly']=Historical_per_day*RF_BF_perc_P1/100*no_of_days_planned
    capacity_input.head(7)


    capacity_input['Planned']=capacity_input['Monthly']*no_of_days_planned/no_of_days_in_month
    capacity_input['Daily']=capacity_input['Monthly']/no_of_days_in_month
    capacity_input.head()
    materials.head()


    # In[29]:



    capacity_input.head()    


    # In[38]:

    #capacity_constraints
    a=list()
    b=list()
    for i in var_P1:
        if capacity_input.loc[capacity_input['Material']==i,'key'].item()==1:
            x=var_P1[i]<=capacity_input.loc[capacity_input['Material']==i,'Planned']
            if sinter_constraint_dict_P1['Capacity Constraints'].get()==1:
                model+=x
        if capacity_input.loc[capacity_input['Material']==i,'key'].item()==2:
            x=var_P1[i]<=capacity_input.loc[capacity_input['Material']==i,'Planned']
            if sinter_constraint_dict_P1['Capacity Constraints'].get()==1:
                model+=x
        if capacity_input.loc[capacity_input['Material']==i,'key'].item()==0:
            x=var_P1[i]>=capacity_input.loc[capacity_input['Material']==i,'Planned']
            if sinter_constraint_dict_P1['Capacity Constraints'].get()==1:
                model+=x




    # In[19]:


    #For P2
    #other constraints & assumptions
    other=pd.read_csv('Data Input/Plant/Blast Furnace/P2/P2 Constraints & Assumptions.csv')
    RF_Internal_perc_P2=other.loc[other['Const&Assump.']=='RF Internal(%)','Value'].item()
    RF_BF_perc_P2=other.loc[other['Const&Assump.']=='RF BF(%)','Value'].item()
    
    #expected_production_constraint
    Production_target=other.loc[other['Const&Assump.']=='Production Target/Day','Value'].item()*no_of_days_planned
    Historical_per_day=other.loc[other['Const&Assump.']=='Production Target/Day','Value'].item()
    if sinter_constraint_dict_P2['Expected Production'].get()==1:
        model+=lpSum(mines.loc[mines['Plant']=='P2','Residue_fines'])+lpSum(materials['Residue_P2']) <= Production_target*(1+RF_Internal_perc_P2/100)
    #feeding rate constraint
    P2_fines_variables_list=list()
    for i in var_fines:
        if mines.loc[mines['possibilities']==i,'Plant'].item()=='P2':
            P2_fines_variables_list.append(var_fines[i])
    if sinter_constraint_dict_P2['Feeding Rate'].get()==1:
        model+=(lpSum(P2_fines_variables_list)+lpSum(var_P2)-var_P2['RF Internal'])/(no_of_days_input*24)<=other.loc[other['Const&Assump.']=='Feeding Rate','Value'].item()


    other.loc[other['Const&Assump.']=='Coke Rate - Own','Value']=25*Production_target/1000
    other.loc[other['Const&Assump.']=='Coke Rate - Purchase','Value']=50*Production_target/1000
    other


    #capacity input constraints
    capacity_input=pd.read_csv('Data Input/Plant/Sinter/Capacity Input.csv')
    capacity_input['Monthly']=capacity_input['Monthly'].str.replace('Model','0').astype(float)
    capacity_input['key']=[5,5,5,1,5,1,1,2,2,2,2,2,1,1,2,2,1,1,5,5,5,5,0,0]


    capacity_input.loc[capacity_input['Material']=='Coke Fines-own','Monthly']=other.loc[other['Const&Assump.']=='Coke Rate - Own','Value'].item()
    capacity_input.loc[capacity_input['Material']=='Coke Fines-purchase','Monthly']=other.loc[other['Const&Assump.']=='Coke Rate - Purchase','Value'].item()

    #smo&sinter relation
    #capacity_input.loc[capacity_input['Material']=='Tensa Total','Monthly']=164542.0
    #capacity_input.loc[capacity_input['Material']=='Tensa for P1','Monthly']=capacity_input.loc[capacity_input['Material']=='Tensa Total','Monthly'].item()*0.5
    #capacity_input.loc[capacity_input['Material']=='Iron Ore Tensa','Monthly']=capacity_input.loc[capacity_input['Material']=='Tensa Total','Monthly'].item()-capacity_input.loc[capacity_input['Material']=='Tensa for P1','Monthly'].item()
    #RF Internal Input
    expected_production_P2=(lpSum(mines.loc[mines['Plant']=='P2','Residue_fines'])+lpSum(materials['Residue_P2']))/(1+RF_Internal_perc_P2/100)
    rf_internal=list()
    rf_internal.append((expected_production_P2)*(RF_Internal_perc_P2/100))
    capacity_input.loc[capacity_input['Material']=='RF Internal','Monthly']=rf_internal
    capacity_input.loc[capacity_input['Material']=='RF Internal','Monthly'].item()
    #RF BF Input
    capacity_input.loc[capacity_input['Material']=='RF BF','Monthly']=Historical_per_day*RF_BF_perc_P2/100*no_of_days_planned
    capacity_input.head(7)


    capacity_input['Planned']=capacity_input['Monthly']*no_of_days_planned/no_of_days_in_month
    capacity_input['Daily']=capacity_input['Monthly']/no_of_days_in_month
    capacity_input.head()
    materials.head()


    # In[29]:



    capacity_input.head()    


    # In[38]:

    #capacity_constraints
    a=list()
    b=list()
    for i in var_P2:
        if capacity_input.loc[capacity_input['Material']==i,'key'].item()==1:
            x=var_P2[i]<=capacity_input.loc[capacity_input['Material']==i,'Planned']
            if sinter_constraint_dict_P2['Capacity Constraints'].get()==1:
                model+=x
        if capacity_input.loc[capacity_input['Material']==i,'key'].item()==2:
            x=var_P2[i]<=capacity_input.loc[capacity_input['Material']==i,'Planned']
            if sinter_constraint_dict_P2['Capacity Constraints'].get()==1:
                model+=x
        if capacity_input.loc[capacity_input['Material']==i,'key'].item()==0:
            x=var_P2[i]>=capacity_input.loc[capacity_input['Material']==i,'Planned']
            if sinter_constraint_dict_P2['Capacity Constraints'].get()==1:
                model+=x




    # In[20]:


    #sinter chemistry constraints_P1
    sinter_chem_constraints=pd.read_csv('Data Input/Plant/Sinter/Sinter Chemsitry Constraints.csv')
    total_residue_fines_P1=lpSum(mines.loc[mines['Plant']=='P1','Residue_fines'])+lpSum(materials['Residue_P1'])
    sinter_chem_constraints['min_res_P1']=sinter_chem_constraints['MIN']*total_residue_fines_P1
    sinter_chem_constraints['max_res_P1']=sinter_chem_constraints['MAX']*total_residue_fines_P1
    sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Basicity','min_res_P1']=[(lpSum(mines.loc[mines['Plant']=='P1','SiO2(new)_fines'])+lpSum(materials['SiO2(new)_P1']))*100*sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Basicity','MIN'].item()]
    sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Basicity','max_res_P1']=[(lpSum(mines.loc[mines['Plant']=='P1','SiO2(new)_fines'])+lpSum(materials['SiO2(new)_P1']))*100*sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Basicity','MAX'].item()]


    # In[21]:


    #sinter chemistry constraints_P2
    total_residue_fines_P2=lpSum(mines.loc[mines['Plant']=='P2','Residue_fines'])+lpSum(materials['Residue_P2'])
    sinter_chem_constraints['min_res_P2']=sinter_chem_constraints['MIN']*total_residue_fines_P2
    sinter_chem_constraints['max_res_P2']=sinter_chem_constraints['MAX']*total_residue_fines_P2
    sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Basicity','min_res_P2']=[(lpSum(mines.loc[mines['Plant']=='P2','SiO2(new)_fines'])+lpSum(materials['SiO2(new)_P2']))*100*sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Basicity','MIN'].item()]
    sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Basicity','max_res_P2']=[(lpSum(mines.loc[mines['Plant']=='P2','SiO2(new)_fines'])+lpSum(materials['SiO2(new)_P2']))*100*sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Basicity','MAX'].item()]


    # In[22]:


    p=list()
    #P1
    #Fe
    Fe_lower_P1=(lpSum(mines.loc[mines['Plant']=='P1','Fe(new)_fines'])+lpSum(materials['Fe(new)_P1']))*100>=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Fe','min_res_P1'].item()
    Fe_upper_P1=(lpSum(mines.loc[mines['Plant']=='P1','Fe(new)_fines'])+lpSum(materials['Fe(new)_P1']))*100<=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Fe','max_res_P1'].item()
    if sinter_constraint_dict_P1['Fe(lower)'].get()==1:
        model+=Fe_lower_P1
    if sinter_constraint_dict_P1['Fe(upper)'].get()==1:
        model+=Fe_upper_P1

    #SiO2
    SiO2_upper_P1=(lpSum(mines.loc[mines['Plant']=='P1','SiO2(new)_fines'])+lpSum(materials['SiO2(new)_P1']))*100>=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='SiO2','min_res_P1'].item()
    SiO2_lower_P1=(lpSum(mines.loc[mines['Plant']=='P1','SiO2(new)_fines'])+lpSum(materials['SiO2(new)_P1']))*100<=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='SiO2','max_res_P1'].item()
    if sinter_constraint_dict_P1['SiO2(lower)'].get()==1:
        model+=SiO2_lower_P1
    if sinter_constraint_dict_P1['SiO2(upper)'].get()==1:
        model+=SiO2_upper_P1


    #Al2O3
    Al2O3_lower_P1=(lpSum(mines.loc[mines['Plant']=='P1','Al2O3(new)_fines'])+lpSum(materials['Al2O3(new)_P1']))*100>=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Al2O3','min_res_P1'].item()
    Al2O3_upper_P1=(lpSum(mines.loc[mines['Plant']=='P1','Al2O3(new)_fines'])+lpSum(materials['Al2O3(new)_P1']))*100<=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Al2O3','max_res_P1'].item()
    if sinter_constraint_dict_P1['Al2O3(lower)'].get()==1:
        model+=Al2O3_lower_P1
    if sinter_constraint_dict_P1['Al2O3(upper)'].get()==1:
        model+=Al2O3_upper_P1

    #CaO
    CaO_lower_P1=(lpSum(mines.loc[mines['Plant']=='P1','CaO(new)_fines'])+lpSum(materials['CaO(new)_P1']))*100>=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='CaO','min_res_P1'].item()
    CaO_upper_P1=(lpSum(mines.loc[mines['Plant']=='P1','CaO(new)_fines'])+lpSum(materials['CaO(new)_P1']))*100<=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='CaO','max_res_P1'].item()
    if sinter_constraint_dict_P1['CaO(lower)'].get()==1:
        model+=CaO_lower_P1
    if sinter_constraint_dict_P1['CaO(upper)'].get()==1:
        model+=CaO_upper_P1

    #MgO
    MgO_lower_P1=(lpSum(mines.loc[mines['Plant']=='P1','MgO(new)_fines'])+lpSum(materials['MgO(new)_P1']))*100>=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='MgO','min_res_P1'].item()
    MgO_upper_P1=(lpSum(mines.loc[mines['Plant']=='P1','MgO(new)_fines'])+lpSum(materials['MgO(new)_P1']))*100<=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='MgO','max_res_P1'].item()
    if sinter_constraint_dict_P1['MgO(lower)'].get()==1:
        model+=MgO_lower_P1
    if sinter_constraint_dict_P1['MgO(upper)'].get()==1:
        model+=MgO_upper_P1

    #Basicity
    Basicity_lower_P1=(lpSum(mines.loc[mines['Plant']=='P1','CaO(new)_fines'])+lpSum(materials['CaO(new)_P1']))*100>=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Basicity','min_res_P1'].item()
    Basicity_upper_P1=(lpSum(mines.loc[mines['Plant']=='P1','CaO(new)_fines'])+lpSum(materials['CaO(new)_P1']))*100<=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Basicity','max_res_P1'].item()
    if sinter_constraint_dict_P1['Basicity(lower)'].get()==1:
        model+=Basicity_lower_P1
    if sinter_constraint_dict_P1['Basicity(upper)'].get()==1:
        model+=Basicity_upper_P1

    # In[23]:


    #P2
    #Fe
    Fe_lower_P2=(lpSum(mines.loc[mines['Plant']=='P2','Fe(new)_fines'])+lpSum(materials['Fe(new)_P2']))*100>=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Fe','min_res_P2'].item()
    Fe_upper_P2=(lpSum(mines.loc[mines['Plant']=='P2','Fe(new)_fines'])+lpSum(materials['Fe(new)_P2']))*100<=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Fe','max_res_P2'].item()
    if sinter_constraint_dict_P2['Fe(lower)'].get()==1:
        model+=Fe_lower_P2
    if sinter_constraint_dict_P2['Fe(upper)'].get()==1:
        model+=Fe_upper_P2
    #SiO2
    SiO2_upper_P2=(lpSum(mines.loc[mines['Plant']=='P2','SiO2(new)_fines'])+lpSum(materials['SiO2(new)_P2']))*100>=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='SiO2','min_res_P2'].item()
    SiO2_lower_P2=(lpSum(mines.loc[mines['Plant']=='P2','SiO2(new)_fines'])+lpSum(materials['SiO2(new)_P2']))*100<=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='SiO2','max_res_P2'].item()
    if sinter_constraint_dict_P2['SiO2(lower)'].get()==1:
        model+=SiO2_lower_P2
    if sinter_constraint_dict_P2['SiO2(upper)'].get()==1:
        model+=SiO2_upper_P2
    
    #Al2O3
    Al2O3_lower_P2=(lpSum(mines.loc[mines['Plant']=='P2','Al2O3(new)_fines'])+lpSum(materials['Al2O3(new)_P2']))*100>=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Al2O3','min_res_P2'].item()
    Al2O3_upper_P2=(lpSum(mines.loc[mines['Plant']=='P2','Al2O3(new)_fines'])+lpSum(materials['Al2O3(new)_P2']))*100<=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Al2O3','max_res_P2'].item()
    if sinter_constraint_dict_P2['Al2O3(lower)'].get()==1:
        model+=Al2O3_lower_P2
    if sinter_constraint_dict_P2['Al2O3(upper)'].get()==1:
        model+=Al2O3_upper_P2
    


    #CaO
    CaO_lower_P2=(lpSum(mines.loc[mines['Plant']=='P2','CaO(new)_fines'])+lpSum(materials['CaO(new)_P2']))*100>=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='CaO','min_res_P2'].item()
    CaO_upper_P2=(lpSum(mines.loc[mines['Plant']=='P2','CaO(new)_fines'])+lpSum(materials['CaO(new)_P2']))*100<=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='CaO','max_res_P2'].item()
    if sinter_constraint_dict_P2['CaO(lower)'].get()==1:
        model+=CaO_lower_P2
    if sinter_constraint_dict_P2['CaO(upper)'].get()==1:
        model+=CaO_upper_P2
    

    #MgO
    MgO_lower_P2=(lpSum(mines.loc[mines['Plant']=='P2','MgO(new)_fines'])+lpSum(materials['MgO(new)_P2']))*100>=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='MgO','min_res_P2'].item()
    MgO_upper_P2=(lpSum(mines.loc[mines['Plant']=='P2','MgO(new)_fines'])+lpSum(materials['MgO(new)_P2']))*100<=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='MgO','max_res_P2'].item()
    if sinter_constraint_dict_P2['MgO(lower)'].get()==1:
        model+=MgO_lower_P2
    if sinter_constraint_dict_P2['MgO(upper)'].get()==1:
        model+=MgO_upper_P2
    


    #Basicity
    Basicity_lower_P2=(lpSum(mines.loc[mines['Plant']=='P2','CaO(new)_fines'])+lpSum(materials['CaO(new)_P2']))*100>=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Basicity','min_res_P2'].item()
    Basicity_upper_P2=(lpSum(mines.loc[mines['Plant']=='P2','CaO(new)_fines'])+lpSum(materials['CaO(new)_P2']))*100<=sinter_chem_constraints.loc[sinter_chem_constraints['Parameters']=='Basicity','max_res_P2'].item()
    if sinter_constraint_dict_P2['Basicity(lower)'].get()==1:
        model+=Basicity_lower_P2
    if sinter_constraint_dict_P2['Basicity(upper)'].get()==1:
        model+=Basicity_upper_P2
    


    # In[24]:


    #P1 Sinter Objective
    #objective_sinter
    sinter_obj_P1=list()
    for i in var_P1:
        sinter_obj_P1.append(var_P1[i]*materials.loc[materials['Material']==i,'Cost/MT'].item())
    lpSum(sinter_obj_P1) 
    #objective_smo_fines_P1
    smo_obj_fines=list()
    for i in var_fines:
        if mines.loc[mines['possibilities']==i,'Plant'].item()=='P1':
            smo_obj_fines.append(var_fines[i]*mines.loc[mines['possibilities']==i,'Total Cost Fines'].item())
    lpSum(smo_obj_fines)   

    model_P1=lpSum(sinter_obj_P1+smo_obj_fines)

    #updating objective

    #P1

    #Cost
    cost_P1=model_P1

    #residue
    #residue_P1_output=total_residue_fines_P2

    #Return_fines_Credit_P1
    Return_Fines_Credit_P1=expected_production_P1*(RF_Internal_perc_P1/100)*materials.loc[materials['Material']=='RF Internal','Cost/MT'].item()


    #Return_fines_bf_Credit_P1
    Return_Fines_BF_Credit_P1=materials.loc[materials['Material']=='RF BF','Cost/MT'].item()*(RF_BF_perc_P1/100)*var_P1['RF BF']

    Total_Cost_P1=cost_P1-Return_Fines_Credit_P1-Return_Fines_BF_Credit_P1
    Total_Cost_P1

    #Riagarh Sinter Objective
    #objective_sinter
    sinter_obj_P2=list()
    for i in var_P2:
        sinter_obj_P2.append(var_P2[i]*materials.loc[materials['Material']==i,'Cost/MT'].item())
    lpSum(sinter_obj_P2) 
    #objective_smo_fines_P2
    smo_obj_fines=list()
    for i in var_fines:
        if mines.loc[mines['possibilities']==i,'Plant'].item()=='P2':
            smo_obj_fines.append(var_fines[i]*mines.loc[mines['possibilities']==i,'Total Cost Fines'].item())
    lpSum(smo_obj_fines)   
    model_P2=lpSum(sinter_obj_P2+smo_obj_fines)

    #objective_smo_lumps
    smo_obj_lumps=list()
    for i in var_lumps:
        smo_obj_lumps.append(var_lumps[i]*mines_lumps.loc[mines_lumps['possibilities']==i,'Total Cost Lumps'].item())
    lpSum(smo_obj_lumps)   



    #P2


    #Cost
    cost_P2=model_P2

    #residue
    #residue_P2_output=total_residue_fines_P2

    #Return_fines_Credit_P2
    Return_Fines_Credit_P2=expected_production_P2*(RF_Internal_perc_P2/100)*materials.loc[materials['Material']=='RF Internal','Cost/MT'].item()


    #Return_fines_bf_Credit_P2
    Return_Fines_BF_Credit_P2=materials.loc[materials['Material']=='RF BF','Cost/MT'].item()*(RF_BF_perc_P2/100)*var_P2['RF BF']

    #Total_Cost_P2
    Total_Cost_P2=cost_P2-Return_Fines_Credit_P2-Return_Fines_BF_Credit_P2


    #p+=Total_Cost_P1+Total_Cost_P2+lpSum(smo_obj_lumps)


    # In[25]:


    #bmo starts here
    bmo_materials=pd.read_csv('Data Input/Plant/Blast Furnace/Common BMO Materials/bmo_materials.csv')
    bmo_materials.loc[bmo_materials['Ash'].isna(),'Ash']=0
    bmo_materials=bmo_materials.drop([0,4])


    # In[26]:


    
    coals=pd.read_csv('Data Input/Head Office/Coke Plant/P2/Coals.csv')
    coals=coals.iloc[:,:7]
    coals.rename(columns={'Coals ': 'coals',
                         'Type':'type', 
                          'Min':'min_perc',
                          'Max': 'max_perc',
                          'Inventory (MT)':'inventory',
                          'Input Price (INR)':'input_price',
                          "Price based on Platt's (INR)":"platt_price"} , inplace=True)
    #coals=coals.iloc[:6,:]
    coals["min_perc"]=coals["min_perc"].str.replace('%','').astype('float')
    coals["max_perc"]=coals["max_perc"].str.replace('%','').astype('float')
    coals["inventory"]=coals["inventory"].str.replace(',','').astype('float')
    coals['input_price']=coals['input_price'].astype('float')

    #costing
    #coal_cost_P2=pd.read_csv('Data Input/Coke/Coking Coal- Monthly Costs.csv')
    #coal_cost_P1=pd.read_csv('Data Input/Coke/Coking Coal Monthly Costs P1.csv')
    #coal_cost_P2

    #for i in coal_cost_P2['Coking Coal']:
        #coals.loc[coals['type']==i,'input_price']=coal_cost_P2.loc[coal_cost_P2['Coking Coal']==i,'December'].item()
    #coals.loc[coals['input_price'].isna(),'input_price']=0   



    bf=pd.read_csv('Data Input/Plant/Coke Plant/P2/blast_furnace.csv')
    bf1=bf.iloc[:8,:4]
    bf1.rename(columns={'Parameters ': 'parameters',
                         'Blend':'blend', 
                          'Lower Limit':'lower_limit',
                          'Upper Limit': 'upper_limit'} , inplace=True)


    # In[3]:


    #bf2=bf.iloc[:8,4:9]
    #bf2.rename(columns={ 'Blend.1':'blend', 
     #                     'Lower Limit.1':'lower_limit',
      #                    'Upper Limit.1': 'upper_limit',} , inplace=True)


    # In[4]:


    blend_type=pd.read_csv('Data Input/Plant/Coke Plant/P2/blend_type.csv')
    blend_type.rename(columns={'Type of Blend ': 'blend_type',
                         'Blend %':'blend_perc', 
                          'Lower Limit':'lower_limit_perc',
                          'Upper Limit': 'upper_limit_perc'},inplace=True)
                          #'Blend %.1':'blend_perc',
                          #'Lower Limit.1':'lower_limit_perc',
                          #"Upper Limit.1":"upper_limit_perc"} , inplace=True)


    # In[5]:


    blend_bf1=blend_type.iloc[:,:4]
    #blend_bf2=blend_type.iloc[:,4:9]
    blend_bf1['upper_limit_perc']=blend_bf1['upper_limit_perc'].str.replace('%','').astype('float')
    blend_bf1['lower_limit_perc']=blend_bf1['lower_limit_perc'].str.replace('%','').astype('float')
    #blend_bf2['upper_limit_perc']=blend_bf2['upper_limit_perc'].str.replace('%','').astype('float')
    #blend_bf2['lower_limit_perc']=blend_bf2['lower_limit_perc'].str.replace('%','').astype('float')


    # In[6]:


    blend_control_parameters=pd.read_csv('Data Input/Plant/Coke Plant/P2/blend_control_parameters.csv')


    # In[7]:


    coal_data=pd.read_csv('Data Input/Plant/Coke Plant/P2/coal_data.csv')
    coal_data=coal_data.iloc[:45,:]
    coal_data.rename(columns={'Category': 'category',
                          'VM':'vm',
                          'CSN': 'csn',
                          'MMR':'mmr',
                          'Log of Fluidity':'log_fluidity',
                          "Vitrinite":"vitrinite",
                             "V9- V14":"v9_v14",
                             "Sulphur":"sulphur",
                             "CSR":"csr",
                             "TM":"tm"} , inplace=True)
    coal_data['MBI']=coal_data['MBI'].astype('float')
    for i in coal_data.columns:
        coal_data.loc[coal_data[i].isna(),i]=0

    # In[8]:


    coke_yield=pd.read_csv('Data Input/Head Office/Coke Plant/P2/coke_yield.csv')

    coke_yield['Yield of coke %']=coke_yield['Yield of coke %'].str.replace('%','').astype(float)
    Credit_For_Nut_Coke_10_25_mm=(coke_yield.loc[(coke_yield["Coke "]=='Nut Coke'),'Yield of coke %']*coke_yield.loc[(coke_yield["Coke "]=='Nut Coke'),'Selling Price per MT'])/100
    Credit_For_Nut_Coke_10_25_mm
    Credit_For_Coke_Fines_0_10_mm=(coke_yield.loc[(coke_yield["Coke "]=='Fines'),'Yield of coke %']*coke_yield.loc[(coke_yield["Coke "]=='Fines'),'Selling Price per MT'])/100
    Credit_For_Coke_Fines_0_10_mm


    # In[9]:


    thumb_rules=pd.read_csv("Data Input/Plant/Coke Plant/P2/Thumb_rule.csv")
    thumb_rules['BF1']=thumb_rules.loc[:,'BF1'].str.replace('%','').astype(float)
    burning_loss=thumb_rules.loc[(thumb_rules["Thumb Rules"]=='Burning Loss'),'BF1']
    end_vm=thumb_rules.loc[(thumb_rules["Thumb Rules"]=='End VM'),'BF1']
    end_moisture=thumb_rules.loc[(thumb_rules["Thumb Rules"]=='End Moisture'),'BF1']
    burning_loss=burning_loss.item()
    end_vm=end_vm.item()
    end_moisture=end_moisture.item()
    inc_fact_coke_ash=thumb_rules.loc[(thumb_rules["Thumb Rules"]=='Incremental Factor for Coke Ash'),'BF1'].item()
    inc_fact_coke_ash
    sulphur_retention=thumb_rules.loc[(thumb_rules["Thumb Rules"]=='Sulphur Retention'),'BF1'].item()
    sulphur_retention=sulphur_retention/100
    sulphur_retention


    # In[10]:


    coke_parameters=pd.read_csv('Data Input/Plant/Coke Plant/P2/coke_ash_sulphur.csv')
    coke_parameters.rename(columns={'Upper Limit':'upper_limit'},inplace=True)
                                    #'Upper Limit.1':'upper_limit1'},inplace=True)
    coke_parameters_bf1=coke_parameters.iloc[:,:3]
    #coke_parameters_bf2=coke_parameters.iloc[:,3:]
    coke_parameters_bf1


    # In[11]:


    #linear programming
    

    #variables
    coal=list()
    for row in coals.iloc[:,0]:
        coal.append(row)
    var_coke_P2=LpVariable.dicts("P2 Coal quantity in tonns",coal,0,None,LpContinuous)



    # In[12]:

    #objective
    cost_objective_coke_P2_list=list()
    for i in var_coke_P2:
        n=coals.loc[(coals['coals']==i),'input_price']
        mod=n*var_coke_P2[i]
        cost_objective_coke_P2_list.append(mod)
        cost_objective_coke_P2=lpSum(cost_objective_coke_P2_list)



    # In[13]:
    no_of_days=30
    coke_required_per_day_tonnes=(blend_control_parameters.iloc[0,1].item()*blend_control_parameters.iloc[1,1].item()+blend_control_parameters.iloc[0,4].item()*blend_control_parameters.iloc[1,4].item())
    coke_required_per_month_tonnes=coke_required_per_day_tonnes*no_of_days_input
    #coke_kgs constraint
    #model+=lpSum(var_coke_P2[i]*(((100-(coal_data.loc[(coal_data['category']==i),'tm'])-((100-(coal_data.loc[(coal_data['category']==i),'tm']))*((coal_data.loc[(coal_data['category']==i),'vm'])/100))-burning_loss+end_vm+end_moisture)/100).item()) for i in var_coke_P2)>=coke_required_per_month_tonnes


    # In[14]:


    #constraints

    #coal type percentage
    for j in var_coke_P2:
        model+= var_coke_P2[j]<=((coals.loc[(coals['coals']==j),'max_perc'])/100)*lpSum(var_coke_P2[i] for i in var_coke_P2)
    for j in var_coke_P2:
        model+= var_coke_P2[j]>=((coals.loc[(coals['coals']==j),'min_perc'])/100)*lpSum(var_coke_P2[i] for i in var_coke_P2)

    #blast furnace parameters
    #VM
    if coke_constraint_dict_P2['VM(upper)'].get()==1:
        model+= lpSum((coal_data.loc[(coal_data['category']==i),'vm'])*var_coke_P2[i] for i in var_coke_P2)<=bf1.loc[(bf1['parameters']=='VM'),'upper_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)
    if coke_constraint_dict_P2['VM(lower)'].get()==1:
        model+= lpSum((coal_data.loc[(coal_data['category']==i),'vm'])*var_coke_P2[i] for i in var_coke_P2)>=bf1.loc[(bf1['parameters']=='VM'),'lower_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)

    #CSN
    if coke_constraint_dict_P2['CSN(upper)'].get()==1:
        model+= lpSum((coal_data.loc[(coal_data['category']==i),'csn'])*var_coke_P2[i] for i in var_coke_P2)<=bf1.loc[(bf1['parameters']=='CSN'),'upper_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)
    if coke_constraint_dict_P2['CSN(lower)'].get()==1:
        model+= lpSum((coal_data.loc[(coal_data['category']==i),'csn'])*var_coke_P2[i] for i in var_coke_P2)>=bf1.loc[(bf1['parameters']=='CSN'),'lower_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)

    #MMR
    if coke_constraint_dict_P2['MMR(upper)'].get()==1:
        model+= lpSum((coal_data.loc[(coal_data['category']==i),'mmr'])*var_coke_P2[i] for i in var_coke_P2)<=bf1.loc[(bf1['parameters']=='MMR'),'upper_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)
    if coke_constraint_dict_P2['MMR(lower)'].get()==1:
        model+= lpSum((coal_data.loc[(coal_data['category']==i),'mmr'])*var_coke_P2[i] for i in var_coke_P2)>=bf1.loc[(bf1['parameters']=='MMR'),'lower_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)

    #vitrinite
    if coke_constraint_dict_P2['Vitrinite(upper)'].get()==1:
        model+= lpSum((coal_data.loc[(coal_data['category']==i),'vitrinite'])*var_coke_P2[i] for i in var_coke_P2)<=bf1.loc[(bf1['parameters']=='Vitrinite'),'upper_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)
    if coke_constraint_dict_P2['Vitrinite(lower)'].get()==1:
        model+= lpSum((coal_data.loc[(coal_data['category']==i),'vitrinite'])*var_coke_P2[i] for i in var_coke_P2)>=bf1.loc[(bf1['parameters']=='Vitrinite'),'lower_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)

    #log(fluidity) 
    if coke_constraint_dict_P2['Log(Fluidity)(lower)'].get()==1:
        model+= lpSum((log10(coal_data.loc[(coal_data['category']==i),'Fluidity'])*100)*var_coke_P2[i] for i in var_coke_P2)>=bf1.loc[(bf1['parameters']=='Log(Fluidity)'),'lower_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)
    if coke_constraint_dict_P2['Log(Fluidity)(upper)'].get()==1:
        model+= lpSum((log10(coal_data.loc[(coal_data['category']==i),'Fluidity'])*100)*var_coke_P2[i] for i in var_coke_P2)<=bf1.loc[(bf1['parameters']=='Log(Fluidity)'),'upper_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)


    #V9-V14
    if coke_constraint_dict_P2['V9-V14(upper)'].get()==1:
        model+= lpSum((coal_data.loc[(coal_data['category']==i),'v9_v14'])*var_coke_P2[i] for i in var_coke_P2)<=bf1.loc[(bf1['parameters']=='V9-V14'),'upper_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)
    if coke_constraint_dict_P2['V9-V14(lower)'].get()==1:
        model+= lpSum((coal_data.loc[(coal_data['category']==i),'v9_v14'])*var_coke_P2[i] for i in var_coke_P2)>=bf1.loc[(bf1['parameters']=='V9-V14'),'lower_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)

    #Sulphur:No upper limit
    if coke_constraint_dict_P2['Sulphur(lower)'].get()==1:
        model+= lpSum((coal_data.loc[(coal_data['category']==i),'sulphur'])*var_coke_P2[i] for i in var_coke_P2)>=bf1.loc[(bf1['parameters']=='Sulphur'),'lower_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)

    #blend_type_constraints
    #lower limit
    if coke_constraint_dict_P2['Prime Hard Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P2[i] for i in coals.loc[(coals['type']=='Prime Hard'),'coals'])>=((blend_bf1.loc[(blend_bf1['blend_type']=='Prime Hard'),'lower_limit_perc'])/100)*lpSum(var_coke_P2[i] for j in var_coke_P2)
    if coke_constraint_dict_P2['Semi Hard Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P2[i] for i in coals.loc[(coals['type']=='Semi Hard'),'coals'])>=((blend_bf1.loc[(blend_bf1['blend_type']=='Semi Hard'),'lower_limit_perc'])/100)*lpSum(var_coke_P2[i] for j in var_coke_P2)
    if coke_constraint_dict_P2['Semi Soft Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P2[i] for i in coals.loc[(coals['type']=='Semi Soft'),'coals'])>=((blend_bf1.loc[(blend_bf1['blend_type']=='Semi Soft'),'lower_limit_perc'])/100)*lpSum(var_coke_P2[i] for j in var_coke_P2)
    if coke_constraint_dict_P2['PCI Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P2[i] for i in coals.loc[(coals['type']=='PCI'),'coals'])>=((blend_bf1.loc[(blend_bf1['blend_type']=='PCI'),'lower_limit_perc'])/100)*lpSum(var_coke_P2[i] for j in var_coke_P2)
    if coke_constraint_dict_P2['Pet Coke Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P2[i] for i in coals.loc[(coals['type']=='Pet Coke'),'coals'])>=((blend_bf1.loc[(blend_bf1['blend_type']=='Pet Coke'),'lower_limit_perc'])/100)*lpSum(var_coke_P2[i] for j in var_coke_P2)
    if coke_constraint_dict_P2['Non Coking Coal Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P2[i] for i in coals.loc[(coals['type']=='Non Coking Coal'),'coals'])>=((blend_bf1.loc[(blend_bf1['blend_type']=='Non Coking Coal'),'lower_limit_perc'])/100)*lpSum(var_coke_P2[i] for j in var_coke_P2)

    #upper limit
    if coke_constraint_dict_P2['Prime Hard Mix(upper)'].get()==1:
        model+=lpSum(var_coke_P2[i] for i in coals.loc[(coals['type']=='Prime Hard'),'coals'])-lpSum((blend_bf1.loc[(blend_bf1['blend_type']=='Prime Hard'),'upper_limit_perc']/100)*lpSum(var_coke_P2[k] for k in var_coke_P2))<=0
    if coke_constraint_dict_P2['Semi Hard Mix(upper)'].get()==1:
        model+=lpSum(var_coke_P2[i] for i in coals.loc[(coals['type']=='Semi Hard'),'coals'])-lpSum((blend_bf1.loc[(blend_bf1['blend_type']=='Semi Hard'),'upper_limit_perc']/100)*lpSum(var_coke_P2[k] for k in var_coke_P2))<=0
    if coke_constraint_dict_P2['Semi Soft Mix(upper)'].get()==1:
        model+=lpSum(var_coke_P2[i] for i in coals.loc[(coals['type']=='Semi Soft'),'coals'])-lpSum((blend_bf1.loc[(blend_bf1['blend_type']=='Semi Soft'),'upper_limit_perc']/100)*lpSum(var_coke_P2[k] for k in var_coke_P2))<=0
    if coke_constraint_dict_P2['PCI Mix(upper)'].get()==1:
        model+=lpSum(var_coke_P2[i] for i in coals.loc[(coals['type']=='PCI'),'coals'])-lpSum((blend_bf1.loc[(blend_bf1['blend_type']=='PCI'),'upper_limit_perc']/100)*lpSum(var_coke_P2[k] for k in var_coke_P2))<=0
    if coke_constraint_dict_P2['Pet Coke Mix(upper)'].get()==1:
        model+=lpSum(var_coke_P2[i] for i in coals.loc[(coals['type']=='Pet Coke'),'coals'])-lpSum((blend_bf1.loc[(blend_bf1['blend_type']=='Pet Coke'),'upper_limit_perc']/100)*lpSum(var_coke_P2[k] for k in var_coke_P2))<=0
    if coke_constraint_dict_P2['Non Coking Coal Mix(upper)'].get()==1:
        model+=lpSum(var_coke_P2[i] for i in coals.loc[(coals['type']=='Non Coking Coal'),'coals'])-lpSum((blend_bf1.loc[(blend_bf1['blend_type']=='Non Coking Coal'),'upper_limit_perc']/100)*lpSum(var_coke_P2[k] for k in var_coke_P2))<=0


    # In[15]:


    #coke ash constraint
    if coke_constraint_dict_P2['Coke Ash'].get()==1:
        model+= (lpSum((coal_data.loc[(coal_data['category']==i),'Ash'])*var_coke_P2[i] for i in var_coke_P2))*inc_fact_coke_ash<=coke_parameters_bf1.loc[(coke_parameters_bf1['Parameters']=='Coke Ash'),'upper_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)


    # In[16]:


    #coke sulphur constraint
    if coke_constraint_dict_P2['Coke Sulphur'].get()==1:
        model+= (lpSum((coal_data.loc[(coal_data['category']==i),'sulphur'])*var_coke_P2[i] for i in var_coke_P2))*sulphur_retention<=coke_parameters_bf1.loc[(coke_parameters_bf1['Parameters']=='Coke Sulphur'),'upper_limit']*lpSum(var_coke_P2[j]for j in var_coke_P2)


    # In[17]:



    vm_db=lpSum((coal_data.loc[(coal_data['category']==i),'VM (db)'])*var_coke_P2[i] for i in var_coke_P2)
    basic=lpSum((coal_data.loc[(coal_data['category']==i),'Basic'])*var_coke_P2[i] for i in var_coke_P2)
    acidic=lpSum((coal_data.loc[(coal_data['category']==i),'Acidic'])*var_coke_P2[i] for i in var_coke_P2)
    mbi=lpSum(coal_data.loc[(coal_data['category']==i),'MBI'])



    #csr constraint
    mbi=lpSum((coal_data.loc[(coal_data['category']==i),'MBI'])*var_coke_P2[i] for i in var_coke_P2)
    if coke_constraint_dict_P2['CSR'].get()==1:
        model+=68.5*lpSum(var_coke_P2[i]for i in var_coke_P2)+0.512*mbi+0.02308*lpSum((log10(coal_data.loc[(coal_data['category']==i),'Fluidity'])*100)*var_coke_P2[i] for i in var_coke_P2)-1.775*lpSum((coal_data.loc[(coal_data['category']==i),'mmr'])*var_coke_P2[i] for i in var_coke_P2)-0.1035*lpSum((coal_data.loc[(coal_data['category']==i),'vitrinite'])*var_coke_P2[i] for i in var_coke_P2)-(lpSum(var_coke_P2[i]for i in var_coke_P2)*bf1.loc[(bf1['parameters']=='CSR '),'lower_limit'].item())>=0


    # inventory constraint
    p=list()
    for i in var_coke_P2:
        if coke_constraint_dict_P2['Inventory'].get()==1:
            model+=var_coke_P2[i]*(blend_control_parameters.iloc[0,1].item()*blend_control_parameters.iloc[1,1].item()+blend_control_parameters.iloc[0,4].item()*blend_control_parameters.iloc[1,4].item())<=coals.loc[coals['coals']==i,'inventory'].item()*lpSum(var_coke_P2[i] for i in var_coke_P2)


    # In[27]:


    final_cost=cost_objective_coke_P2


    # In[28]:




    # In[21]:


    cost_calculation=pd.read_csv('Data Input/Head Office/Coke Plant/P2/Cost Calculation.csv')
    cost_calculation.rename(columns={'Unnamed: 1':'bf1'},inplace=True)
                                    #'Unnamed: 2':'bf2',
                                    #'Unnamed: 3':'total'},inplace=True)
    cost_calculation['bf1']=cost_calculation['bf1'].str.replace(',','')
    cost_calculation['bf1']=cost_calculation['bf1'].str.replace('-','0').astype(float)
    cost_calculation


    # In[22]:


    #other_operation_cost
    Maintenance_Cost_and_Consumables=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Maintenance Cost & Consumables'),'bf1'].item()
    power=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Power '),'bf1'].item()
    RMHS_Allocation=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='RMHS Allocation'),'bf1'].item()
    Refractories=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Refractories'),'bf1'].item()
    Manpower_Supply=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Manpower Supply'),'bf1'].item()
    Job_Contract_Outsourcing=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Job Contract & Outsourcing'),'bf1'].item()
    Manpower_Tonnage=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Manpower Tonnage'),'bf1'].item()
    Manpower_on_Roll=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Manpower on Roll'),'bf1'].item()
    other_operation_cost=Maintenance_Cost_and_Consumables+power+RMHS_Allocation+Refractories+Manpower_Supply+Job_Contract_Outsourcing+Manpower_Tonnage+Manpower_on_Roll

    #Coke - Gross Basis
    final_cost_mt=final_cost/coke_required_per_month_tonnes
    Credit_For_Nut_Coke_10_25_mm=((coke_yield.loc[(coke_yield["Coke "]=='Nut Coke'),'Yield of coke %']*coke_yield.loc[(coke_yield["Coke "]=='Nut Coke'),'Selling Price per MT'])/100).item()
    Credit_For_Coke_Fines_0_10_mm=((coke_yield.loc[(coke_yield["Coke "]=='Fines'),'Yield of coke %']*coke_yield.loc[(coke_yield["Coke "]=='Fines'),'Selling Price per MT'])/100).item()
    Coke_Gross_Basis=final_cost_mt-Credit_For_Nut_Coke_10_25_mm-Credit_For_Coke_Fines_0_10_mm
    Coke_Gross_Basis_req=final_cost-(Credit_For_Nut_Coke_10_25_mm+Credit_For_Coke_Fines_0_10_mm)*coke_required_per_month_tonnes

    #utilities
    BF_Gas_NM3_t_=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='BF Gas (NM3/t)'),'bf1'].item()
    Utilites=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Utilites '),'bf1'].item()
    utilities=BF_Gas_NM3_t_+Utilites
    utilities

    #Credits for Flue gas GCAL/tn
    Credits_for_Flue_gas_GCAL_tn=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Credits for Flue gas GCAL/tn'),'bf1'].item()
    Credits_for_Flue_gas_GCAL_tn

    #Total_Operational_Coke_Cost
    Total_Operational_Coke_Cost=Coke_Gross_Basis+utilities+other_operation_cost-Credits_for_Flue_gas_GCAL_tn
    Total_Operational_Coke_Cost_req=Total_Operational_Coke_Cost*coke_required_per_month_tonnes
    #Indirect OH 
    Indirect_Personnel=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Indirect Personnel'),'bf1'].item()
    Indirect_Stores=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Indirect Stores'),'bf1'].item()
    Indirect_Power=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Indirect Power'),'bf1'].item()
    Outsourcing_II=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Outsourcing - II'),'bf1'].item()
    Administrative_Overhead=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Administrative Overhead'),'bf1'].item()
    Corporate_Expenses=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Corporate Expenses'),'bf1'].item()
    indirect_oh=Indirect_Personnel+Indirect_Stores+Indirect_Power+Outsourcing_II+Administrative_Overhead+Corporate_Expenses
    indirect_oh_req=indirect_oh*coke_required_per_month_tonnes
    #Total Variable Cost 
    total_variable_cost=Total_Operational_Coke_Cost+indirect_oh
    total_variable_cost_req=Total_Operational_Coke_Cost_req+indirect_oh_req
    #Total Cost
    finance_cost=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Finance Cost'),'bf1'].item()
    depreciation_cost=cost_calculation.loc[(cost_calculation['Cost Calculation (INR)']=='Depreciation'),'bf1'].item()
    fixed_cost=finance_cost+depreciation_cost
    total_cost=total_variable_cost+fixed_cost
    total_cost_req=total_variable_cost_req+(fixed_cost*coke_required_per_month_tonnes)
    #final model
    #Total Cost of Usable Coke at Blast Furnace
    Total_Cost_of_Usable_Coke_at_Blast_Furnace=total_cost/(coke_yield.loc[(coke_yield["Coke "]=='BF Coke '),'Yield of coke %'].item()/100) 
    Total_Cost_of_Usable_Coke_at_Blast_Furnace_req=total_cost_req/(coke_yield.loc[(coke_yield["Coke "]=='BF Coke '),'Yield of coke %'].item()/100)



    # In[23]:


    #Cost of Coke for MTHM 
    Cost_of_Coke_for_MTHM=Total_Cost_of_Usable_Coke_at_Blast_Furnace*(thumb_rules.loc[(thumb_rules['Thumb Rules']=='Consumption Rate'),'BF1'].item()/1000)
    Total_Cost_of_Usable_Coke_at_Blast_Furnace_req




    # In[29]:



    
    coals_P1=pd.read_csv('Data Input/Head Office/Coke Plant/P1/Coals_P1.csv')
    coals_P1=coals_P1.iloc[:,:7]
    coals_P1.rename(columns={'Coals ': 'coals_P1',
                         'Type':'type', 
                          'Min':'min_perc',
                          'Max': 'max_perc',
                          'Inventory (MT)':'inventory',
                          'Input Price (INR)':'input_price',
                          "Price based on Platt's (INR)":"platt_price"} , inplace=True)
    #coals_P1=coals_P1.iloc[:6,:]
    coals_P1["min_perc"]=coals_P1["min_perc"].str.replace('%','').astype('float')
    coals_P1["max_perc"]=coals_P1["max_perc"].str.replace('%','').astype('float')
    coals_P1["inventory"]=coals_P1["inventory"].str.replace(',','').astype('float')
    coals_P1['input_price']=coals_P1['input_price'].astype('float')

    #costing
    #coal_cost_P1=pd.read_csv('Data Input/Coke/Coking Coal Monthly Costs P1.csv')

    #for i in coal_cost_P1['Coking Coal']:
        #coals_P1.loc[coals_P1['type']==i,'input_price']=coal_cost_P1.loc[coal_cost_P1['Coking Coal']==i,'December'].item()
    #coals_P1.loc[coals_P1['input_price'].isna(),'input_price']=0 
    # In[2]:


    bf_P1=pd.read_csv('Data Input/Plant/Coke Plant/P1/blast_furnace_P1.csv')
    bf1_P1=bf_P1.iloc[:8,:4]
    bf1_P1.rename(columns={'Parameters ': 'parameters',
                         'Blend':'blend', 
                          'Lower Limit':'lower_limit',
                          'Upper Limit': 'upper_limit',} , inplace=True)


    # In[4]:


    blend_type_P1=pd.read_csv('Data Input/Plant/Coke Plant/P1/blend_type_P1.csv')
    blend_type_P1.rename(columns={'Type of Blend ': 'blend_type_P1',
                         'Blend %':'blend_perc', 
                          'Lower Limit':'lower_limit_perc',
                          'Upper Limit': 'upper_limit_perc'},inplace=True)
                          #'Blend %.1':'blend_perc',
                          #'Lower Limit.1':'lower_limit_perc',
                          #"Upper Limit.1":"upper_limit_perc"} , inplace=True)


    # In[5]:


    blend_bf1_P1=blend_type_P1.iloc[:,:4]
    blend_bf1_P1['upper_limit_perc']=blend_bf1_P1['upper_limit_perc'].str.replace('%','').astype('float')
    blend_bf1_P1['lower_limit_perc']=blend_bf1_P1['lower_limit_perc'].str.replace('%','').astype('float')



    # In[6]:


    blend_control_parameters_P1=pd.read_csv('Data Input/Plant/Coke Plant/P1/blend_control_parameters_P1.csv')


    # In[7]:


    coal_data_P1=pd.read_csv('Data Input/Plant/Coke Plant/P1/coal_data_P1.csv')
    coal_data_P1=coal_data_P1.iloc[:47,:]
    coal_data_P1.rename(columns={'Category': 'category',
                          'VM':'vm',
                          'CSN': 'csn',
                          'MMR':'mmr',
                          'Log of Fluidity':'log_fluidity',
                          "Vitrinite":"vitrinite",
                             "V9- V14":"v9_v14",
                             "Sulphur":"sulphur",
                             "CSR":"csr",
                             "TM":"tm"} , inplace=True)
    coal_data_P1['MBI']=coal_data_P1['MBI'].astype('float')
    for i in coal_data_P1.columns:
        coal_data_P1.loc[coal_data_P1[i].isna(),i]=0


    # In[8]:


    coke_yield_P1=pd.read_csv('Data Input/Head Office/Coke Plant/P1/coke_yield_P1.csv')

    coke_yield_P1['Yield of coke %']=coke_yield_P1['Yield of coke %'].str.replace('%','').astype(float)
    Credit_For_Nut_Coke_10_25_mm_P1=(coke_yield_P1.loc[(coke_yield_P1["Coke "]=='Nut Coke'),'Yield of coke %']*coke_yield_P1.loc[(coke_yield_P1["Coke "]=='Nut Coke'),'Selling Price per MT'])/100
    Credit_For_Coke_Fines_0_10_mm_P1=(coke_yield_P1.loc[(coke_yield_P1["Coke "]=='Fines'),'Yield of coke %']*coke_yield_P1.loc[(coke_yield_P1["Coke "]=='Fines'),'Selling Price per MT'])/100

    # In[9]:


    thumb_rules_P1=pd.read_csv("Data Input/Plant/Coke Plant/P1/Thumb_rule_P1.csv")
    thumb_rules_P1['BF1']=thumb_rules_P1.loc[:,'BF1'].str.replace('%','').astype(float)
    burning_loss_P1=thumb_rules_P1.loc[(thumb_rules_P1["Thumb Rules"]=='Burning Loss'),'BF1']
    end_vm_P1_P1=thumb_rules_P1.loc[(thumb_rules_P1["Thumb Rules"]=='End VM'),'BF1']
    end_moisture_P1=thumb_rules_P1.loc[(thumb_rules_P1["Thumb Rules"]=='End Moisture'),'BF1']
    burning_loss_P1=burning_loss_P1.item()
    end_vm_P1_P1=end_vm_P1_P1.item()
    end_moisture_P1=end_moisture_P1.item()
    inc_fact_coke_ash_P1=thumb_rules_P1.loc[(thumb_rules_P1["Thumb Rules"]=='Incremental Factor for Coke Ash'),'BF1'].item()
    sulphur_retention_P1=thumb_rules_P1.loc[(thumb_rules_P1["Thumb Rules"]=='Sulphur Retention'),'BF1'].item()
    sulphur_retention_P1=sulphur_retention_P1/100

    # In[10]:


    coke_parameters_P1=pd.read_csv('Data Input/Plant/Coke Plant/P1/coke_ash_sulphur_P1.csv')
    coke_parameters_P1.rename(columns={'Upper Limit':'upper_limit'},inplace=True)
                                    #'Upper Limit.1':'upper_limit1'},inplace=True)
    coke_parameters_bf1_P1=coke_parameters_P1.iloc[:,:3]


    # In[11]:


    #linear programming
    
    #model=LpProblem('Cost minimising blending problem',LpMinimize)
    #variables
    coal_P1=list()
    for row in coals_P1.iloc[:,0]:
        coal_P1.append(row)
    var_coke_P1=LpVariable.dicts("P1 Coal quantity in tonns",coal_P1,0,None,LpContinuous)



    # In[12]:

    #objective
    cost_objective_coke_P1_list=list()
    for i in var_coke_P1:
        n=coals_P1.loc[(coals_P1['coals_P1']==i),'input_price']
        mod=n*var_coke_P1[i]
        cost_objective_coke_P1_list.append(mod)
        cost_objective_coke_P1=lpSum(cost_objective_coke_P1_list)



    # In[13]:


    #coke_kgs constraint
    no_of_days=30
    coke_required_per_day_tonnes_P1=(blend_control_parameters_P1.iloc[0,1].item()*blend_control_parameters_P1.iloc[1,1].item())
    coke_required_per_month_tonnes_P1=coke_required_per_day_tonnes_P1*no_of_days_input
    #coke_kgs constraint
    #model+=lpSum(var_coke_P1[i]*(((100-(coal_data_P1.loc[(coal_data_P1['category']==i),'tm'])-((100-(coal_data_P1.loc[(coal_data_P1['category']==i),'tm']))*((coal_data_P1.loc[(coal_data_P1['category']==i),'vm'])/100))-burning_loss_P1+end_vm_P1_P1+end_moisture_P1)/100).item()) for i in var_coke_P1)>=coke_required_per_month_tonnes_P1


    # In[14]:


    #constraints

    #coal type percentage
    for j in var_coke_P1:
        model+= var_coke_P1[j]<=((coals_P1.loc[(coals_P1['coals_P1']==j),'max_perc'])/100)*lpSum(var_coke_P1[i] for i in var_coke_P1)
    for j in var_coke_P1:
        model+= var_coke_P1[j]>=((coals_P1.loc[(coals_P1['coals_P1']==j),'min_perc'])/100)*lpSum(var_coke_P1[i] for i in var_coke_P1)

    #blast furnace parameters
    #VM
    if coke_constraint_dict_P1['VM(upper)'].get()==1:
        model+= lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'vm'])*var_coke_P1[i] for i in var_coke_P1)<=bf1_P1.loc[(bf1_P1['parameters']=='VM'),'upper_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)
    if coke_constraint_dict_P1['VM(lower)'].get()==1:
        model+= lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'vm'])*var_coke_P1[i] for i in var_coke_P1)>=bf1_P1.loc[(bf1_P1['parameters']=='VM'),'lower_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)

    #CSN
    if coke_constraint_dict_P1['CSN(upper)'].get()==1:
        model+= lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'csn'])*var_coke_P1[i] for i in var_coke_P1)<=bf1_P1.loc[(bf1_P1['parameters']=='CSN'),'upper_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)
    if coke_constraint_dict_P1['CSN(lower)'].get()==1:
        model+= lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'csn'])*var_coke_P1[i] for i in var_coke_P1)>=bf1_P1.loc[(bf1_P1['parameters']=='CSN'),'lower_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)

    #MMR
    if coke_constraint_dict_P1['MMR(upper)'].get()==1:
        model+= lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'mmr'])*var_coke_P1[i] for i in var_coke_P1)<=bf1_P1.loc[(bf1_P1['parameters']=='MMR'),'upper_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)
    if coke_constraint_dict_P1['MMR(lower)'].get()==1:
        model+= lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'mmr'])*var_coke_P1[i] for i in var_coke_P1)>=bf1_P1.loc[(bf1_P1['parameters']=='MMR'),'lower_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)

    #vitrinite
    if coke_constraint_dict_P1['Vitrinite(upper)'].get()==1:
        model+= lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'vitrinite'])*var_coke_P1[i] for i in var_coke_P1)<=bf1_P1.loc[(bf1_P1['parameters']=='Vitrinite'),'upper_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)
    if coke_constraint_dict_P1['Vitrinite(lower)'].get()==1:
        model+= lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'vitrinite'])*var_coke_P1[i] for i in var_coke_P1)>=bf1_P1.loc[(bf1_P1['parameters']=='Vitrinite'),'lower_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)

    #log(fluidity) 
    if coke_constraint_dict_P1['Log(Fluidity)(lower)'].get()==1:
        model+= lpSum((log10(coal_data_P1.loc[(coal_data_P1['category']==i),'Fluidity'])*100)*var_coke_P1[i] for i in var_coke_P1)>=bf1_P1.loc[(bf1_P1['parameters']=='Log(Fluidity)'),'lower_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)
    if coke_constraint_dict_P1['Log(Fluidity)(upper)'].get()==1:
        model+= lpSum((log10(coal_data_P1.loc[(coal_data_P1['category']==i),'Fluidity'])*100)*var_coke_P1[i] for i in var_coke_P1)<=bf1_P1.loc[(bf1_P1['parameters']=='Log(Fluidity)'),'upper_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)


    #V9-V14
    if coke_constraint_dict_P1['V9-V14(upper)'].get()==1:
        model+= lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'v9_v14'])*var_coke_P1[i] for i in var_coke_P1)<=bf1_P1.loc[(bf1_P1['parameters']=='V9-V14'),'upper_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)
    if coke_constraint_dict_P1['V9-V14(lower)'].get()==1:
        model+= lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'v9_v14'])*var_coke_P1[i] for i in var_coke_P1)>=bf1_P1.loc[(bf1_P1['parameters']=='V9-V14'),'lower_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)

    #Sulphur:No upper limit
    if coke_constraint_dict_P1['Sulphur(lower)'].get()==1:
        model+= lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'sulphur'])*var_coke_P1[i] for i in var_coke_P1)>=bf1_P1.loc[(bf1_P1['parameters']=='Sulphur'),'lower_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)

    #blend_type_P1_constraints
    #lower limit
    if coke_constraint_dict_P1['Prime Hard Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P1[i] for i in coals_P1.loc[(coals_P1['type']=='Prime Hard'),'coals_P1'])>=((blend_bf1_P1.loc[(blend_bf1_P1['blend_type_P1']=='Prime Hard'),'lower_limit_perc'])/100)*lpSum(var_coke_P1[j] for j in var_coke_P1)
    if coke_constraint_dict_P1['Semi Hard Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P1[i] for i in coals_P1.loc[(coals_P1['type']=='Semi Hard'),'coals_P1'])>=((blend_bf1_P1.loc[(blend_bf1_P1['blend_type_P1']=='Semi Hard'),'lower_limit_perc'])/100)*lpSum(var_coke_P1[j] for j in var_coke_P1)
    if coke_constraint_dict_P1['Semi Soft Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P1[i] for i in coals_P1.loc[(coals_P1['type']=='Semi Soft'),'coals_P1'])>=((blend_bf1_P1.loc[(blend_bf1_P1['blend_type_P1']=='Semi Soft'),'lower_limit_perc'])/100)*lpSum(var_coke_P1[j] for j in var_coke_P1)
    if coke_constraint_dict_P1['PCI Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P1[i] for i in coals_P1.loc[(coals_P1['type']=='PCI'),'coals_P1'])>=((blend_bf1_P1.loc[(blend_bf1_P1['blend_type_P1']=='PCI'),'lower_limit_perc'])/100)*lpSum(var_coke_P1[j] for j in var_coke_P1)
    if coke_constraint_dict_P1['Pet Coke Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P1[i] for i in coals_P1.loc[(coals_P1['type']=='Pet Coke'),'coals_P1'])>=((blend_bf1_P1.loc[(blend_bf1_P1['blend_type_P1']=='Pet Coke'),'lower_limit_perc'])/100)*lpSum(var_coke_P1[j] for j in var_coke_P1)
    if coke_constraint_dict_P1['Non Coking Coal Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P1[i] for i in coals_P1.loc[(coals_P1['type']=='Non Coking Coal'),'coals_P1'])>=((blend_bf1_P1.loc[(blend_bf1_P1['blend_type_P1']=='Non Coking Coal'),'lower_limit_perc'])/100)*lpSum(var_coke_P1[j] for j in var_coke_P1)

    #upper limit
    if coke_constraint_dict_P1['Prime Hard Mix(upper)'].get()==1:
        model+=lpSum(var_coke_P1[i] for i in coals_P1.loc[(coals_P1['type']=='Prime Hard'),'coals_P1'])-lpSum((blend_bf1_P1.loc[(blend_bf1_P1['blend_type_P1']=='Prime Hard'),'upper_limit_perc']/100)*lpSum(var_coke_P1[k] for k in var_coke_P1))<=0
    if coke_constraint_dict_P1['Semi Hard Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P1[i] for i in coals_P1.loc[(coals_P1['type']=='Semi Hard'),'coals_P1'])-lpSum((blend_bf1_P1.loc[(blend_bf1_P1['blend_type_P1']=='Semi Hard'),'upper_limit_perc']/100)*lpSum(var_coke_P1[k] for k in var_coke_P1))<=0
    if coke_constraint_dict_P1['Semi Soft Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P1[i] for i in coals_P1.loc[(coals_P1['type']=='Semi Soft'),'coals_P1'])-lpSum((blend_bf1_P1.loc[(blend_bf1_P1['blend_type_P1']=='Semi Soft'),'upper_limit_perc']/100)*lpSum(var_coke_P1[k] for k in var_coke_P1))<=0
    if coke_constraint_dict_P1['PCI Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P1[i] for i in coals_P1.loc[(coals_P1['type']=='PCI'),'coals_P1'])-lpSum((blend_bf1_P1.loc[(blend_bf1_P1['blend_type_P1']=='PCI'),'upper_limit_perc']/100)*lpSum(var_coke_P1[k] for k in var_coke_P1))<=0
    if coke_constraint_dict_P1['Pet Coke Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P1[i] for i in coals_P1.loc[(coals_P1['type']=='Pet Coke'),'coals_P1'])-lpSum((blend_bf1_P1.loc[(blend_bf1_P1['blend_type_P1']=='Pet Coke'),'upper_limit_perc']/100)*lpSum(var_coke_P1[k] for k in var_coke_P1))<=0
    if coke_constraint_dict_P1['Non Coking Coal Mix(lower)'].get()==1:
        model+=lpSum(var_coke_P1[i] for i in coals_P1.loc[(coals_P1['type']=='Non Coking Coal'),'coals_P1'])-lpSum((blend_bf1_P1.loc[(blend_bf1_P1['blend_type_P1']=='Non Coking Coal'),'upper_limit_perc']/100)*lpSum(var_coke_P1[k] for k in var_coke_P1))<=0


    # In[15]:


    #coke ash constraint
    if coke_constraint_dict_P1['Coke Ash'].get()==1:
        model+= (lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'Ash'])*var_coke_P1[i] for i in var_coke_P1))*inc_fact_coke_ash_P1<=coke_parameters_bf1_P1.loc[(coke_parameters_bf1_P1['Parameters']=='Coke Ash'),'upper_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)


    # In[16]:


    #coke sulphur constraint
    if coke_constraint_dict_P1['Coke Sulphur'].get()==1:
        model+= (lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'sulphur'])*var_coke_P1[i] for i in var_coke_P1))*sulphur_retention_P1<=coke_parameters_bf1_P1.loc[(coke_parameters_bf1_P1['Parameters']=='Coke Sulphur'),'upper_limit']*lpSum(var_coke_P1[j]for j in var_coke_P1)


    # In[17]:



    vm_db_P1=lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'VM (db)'])*var_coke_P1[i] for i in var_coke_P1)
    basic_P1=lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'Basic'])*var_coke_P1[i] for i in var_coke_P1)
    acidic_P1=lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'Acidic'])*var_coke_P1[i] for i in var_coke_P1)
    mbi_P1=lpSum(coal_data_P1.loc[(coal_data_P1['category']==i),'MBI'])



    #csr constraint
    mbi_P1=lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'MBI'])*var_coke_P1[i] for i in var_coke_P1)
    if coke_constraint_dict_P1['CSR'].get()==1:
        model+=68.5*lpSum(var_coke_P1[i]for i in var_coke_P1)+0.512*mbi_P1+0.02308*lpSum((log10(coal_data_P1.loc[(coal_data_P1['category']==i),'Fluidity'])*100)*var_coke_P1[i] for i in var_coke_P1)-1.775*lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'mmr'])*var_coke_P1[i] for i in var_coke_P1)-0.075*lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'vitrinite'])*var_coke_P1[i] for i in var_coke_P1)-(lpSum(var_coke_P1[i]for i in var_coke_P1)*bf1_P1.loc[(bf1_P1['parameters']=='CSR '),'lower_limit'].item())>=0


    # inventory constraint
    p=list()
    for i in var_coke_P1:
        if coke_constraint_dict_P1['Inventory'].get()==1:
            model+=var_coke_P1[i]*(blend_control_parameters_P1.iloc[0,1].item()*blend_control_parameters_P1.iloc[1,1].item())<=coals_P1.loc[coals_P1['coals_P1']==i,'inventory'].item()*lpSum(var_coke_P1[i] for i in var_coke_P1)


    # In[30]:


    final_cost_P1=cost_objective_coke_P1


    # In[31]:


    cost_calculation_P1=pd.read_csv('Data Input/Head Office/Coke Plant/P1/Cost Calculation_P1.csv')
    cost_calculation_P1.rename(columns={'Unnamed: 1':'bf1'},inplace=True)
                                    #'Unnamed: 2':'bf2',
                                    #'Unnamed: 3':'total'},inplace=True)
    cost_calculation_P1['bf1']=cost_calculation_P1['bf1'].str.replace(',','')
    cost_calculation_P1['bf1']=cost_calculation_P1['bf1'].str.replace('-','0').astype(float)

    # In[22]:
    CO_Gas_NM3_t_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='CO Gas (NM3/t)'),'bf1'].item()
    CO_Gas_flaring_NM3_t_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='CO Gas flaring (NM3/t)'),'bf1'].item()
    MP_Steam_NM3_t_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='MP Steam (NM3/t)'),'bf1'].item()
    Nitrogen_NM3_t_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Nitrogen (NM3/t)'),'bf1'].item()
    Water_Treatment_Exp_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Water Treatment Exp'),'bf1'].item()
    Credits_for_Coke_Oven_Gas_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Credits for Coke Oven Gas'),'bf1'].item()
    Credits_for_Tar_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Credits for Tar'),'bf1'].item()
    Credits_for_Cruze_Benzol_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Credits for Cruze Benzol'),'bf1'].item()
    Direct_Admin_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Direct Admin'),'bf1'].item()
    Direct_Manpower_Allied_Co_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Direct Manpower/Allied Co'),'bf1'].item()
    Direct_Salary_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Direct Salary'),'bf1'].item()
    Corporate_Allocation_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Corporate Allocation'),'bf1'].item()
    Stock_Yard_Marketing_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Stock Yard & Marketing'),'bf1'].item()
    Indirect_OH_t_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Indirect OH'),'bf1'].item()
    CGP_FC_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='CGP FC'),'bf1'].item()
    Water_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Water'),'bf1'].item()

    #other_operation_cost
    Maintenance_Cost_and_Consumables_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Maintenance Cost & Consumables'),'bf1'].item()
    power_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Power '),'bf1'].item()
    RMHS_Allocation_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='RMHS Allocation'),'bf1'].item()
    #Refractories_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Refractories'),'bf1'].item()
    #Manpower_Supply_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Manpower Supply'),'bf1'].item()
    Job_Contract_Outsourcing_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Job Contract & Outsourcing'),'bf1'].item()
    #Manpower_Tonnage_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Manpower Tonnage'),'bf1'].item()
    #Manpower_on_Roll_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Manpower on Roll'),'bf1'].item()
    other_operation_cost_P1=Maintenance_Cost_and_Consumables_P1+power_P1+RMHS_Allocation_P1+Job_Contract_Outsourcing_P1+Water_Treatment_Exp_P1

    #Coke - Gross Basis
    final_cost_mt_P1=final_cost_P1/coke_required_per_month_tonnes_P1
    Credit_For_Nut_Coke_10_25_mm_P1=((coke_yield_P1.loc[(coke_yield_P1["Coke "]=='Nut Coke'),'Yield of coke %']*coke_yield_P1.loc[(coke_yield_P1["Coke "]=='Nut Coke'),'Selling Price per MT'])/100).item()
    Credit_For_Coke_Fines_0_10_mm_P1=((coke_yield_P1.loc[(coke_yield_P1["Coke "]=='Fines'),'Yield of coke %']*coke_yield_P1.loc[(coke_yield_P1["Coke "]=='Fines'),'Selling Price per MT'])/100).item()
    Coke_Gross_Basis_P1=final_cost_mt_P1-Credit_For_Nut_Coke_10_25_mm_P1-Credit_For_Coke_Fines_0_10_mm_P1
    Coke_Gross_Basis_req_P1=final_cost_P1-(Credit_For_Nut_Coke_10_25_mm_P1+Credit_For_Coke_Fines_0_10_mm_P1)*coke_required_per_month_tonnes_P1

    #utilities
    BF_Gas_NM3_t_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='BF Gas (NM3/t)'),'bf1'].item()
    #Utilites_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Utilites '),'bf1'].item()
    utilities_P1=BF_Gas_NM3_t_P1+CO_Gas_NM3_t_P1+CO_Gas_flaring_NM3_t_P1+Nitrogen_NM3_t_P1+MP_Steam_NM3_t_P1
    utilities_P1

    #Credits for Flue gas GCAL/tn
    #Credits_for_Flue_gas_GCAL_tn_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Credits for Flue gas GCAL/tn'),'bf1'].item()
    #Credits_for_Flue_gas_GCAL_tn_P1
    Total_Credit_P1=Credits_for_Coke_Oven_Gas_P1+Credits_for_Tar_P1+Credits_for_Cruze_Benzol_P1

    #Total_Operational_Coke_Cost
    Total_Operational_Coke_Cost_P1=Coke_Gross_Basis_P1+utilities_P1+other_operation_cost_P1-Total_Credit_P1
    Total_Operational_Coke_Cost_req_P1=Total_Operational_Coke_Cost_P1*coke_required_per_month_tonnes_P1
    #Indirect OH 
    indirect_oh_P1=Direct_Admin_P1 +Direct_Manpower_Allied_Co_P1+Direct_Salary_P1+Corporate_Allocation_P1+Stock_Yard_Marketing_P1+Indirect_OH_t_P1+CGP_FC_P1+Water_P1
    indirect_oh_req_P1=indirect_oh_P1*coke_required_per_month_tonnes_P1
    #Total Variable Cost 
    total_variable_cost_P1=Total_Operational_Coke_Cost_P1+indirect_oh_P1
    total_variable_cost_req_P1=Total_Operational_Coke_Cost_req_P1+indirect_oh_req_P1
    #Total Cost
    finance_cost_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Finance Cost'),'bf1'].item()
    depreciation_cost_P1=cost_calculation_P1.loc[(cost_calculation_P1['Cost Calculation (INR)']=='Depreciation'),'bf1'].item()
    fixed_cost_P1=finance_cost_P1+depreciation_cost_P1
    total_cost_P1=total_variable_cost_P1+fixed_cost_P1
    total_cost_req_P1=total_variable_cost_req_P1+(fixed_cost_P1*coke_required_per_month_tonnes_P1)
    #final model
    #Total Cost of Usable Coke at Blast Furnace
    Total_Cost_of_Usable_Coke_at_Blast_Furnace_P1=total_cost_P1/(coke_yield_P1.loc[(coke_yield_P1["Coke "]=='BF Coke '),'Yield of coke %'].item()/100) 
    Total_Cost_of_Usable_Coke_at_Blast_Furnace_req_P1=total_cost_req_P1/(coke_yield_P1.loc[(coke_yield_P1["Coke "]=='BF Coke '),'Yield of coke %'].item()/100)



    # In[23]:


    #Cost of Coke for MTHM 
    Cost_of_Coke_for_MTHM_P1=Total_Cost_of_Usable_Coke_at_Blast_Furnace_P1*(thumb_rules_P1.loc[(thumb_rules_P1['Thumb Rules']=='Consumption Rate'),'BF1'].item()/1000)




    # In[32]:


    #change constants to values from bmo materials table
    coals['Ash']=[coal_data.loc[coal_data['category']==i,'Ash'].item()*thumb_rules.loc[(thumb_rules['Thumb Rules']=='Incremental Factor for Coke Ash'),'BF1'].item() for i in var_coke_P2]
    coals['Moi']=3.82
    coals['Fe']=0.67
    coals['SiO2']=0.52*coals['Ash']
    coals['Al2O3']=0.3*coals['Ash']
    coals['MgO']=0.07
    coals['CaO']=0.03
    coals['S']=[coal_data.loc[coal_data['category']==i,'sulphur'].item()*sulphur_retention for i in var_coke_P2]
    coals['P']=0.05


    # In[33]:


    #change constants to values from bmo materials table
    coals_P1['Ash']=[coal_data_P1.loc[coal_data_P1['category']==i,'Ash'].item()*thumb_rules_P1.loc[(thumb_rules['Thumb Rules']=='Incremental Factor for Coke Ash'),'BF1'].item() for i in var_coke_P1]
    coals_P1['Moi']=3.82
    coals_P1['Fe']=0.67
    coals_P1['SiO2']=0.52*coals_P1['Ash']
    coals_P1['Al2O3']=0.3*coals_P1['Ash']
    coals_P1['MgO']=0.07
    coals_P1['CaO']=0.03
    coals_P1['S']=[coal_data_P1.loc[coal_data_P1['category']==i,'sulphur'].item()*sulphur_retention_P1 for i in var_coke_P1]
    coals_P1['P']=0.05
    coals_P1


    # In[34]:


    #updating coals P2 chemistry values
    #Feeding DMT
    coke_feeding_dmt_list_P2=list()
    for i in coals['coals']:
        x=coals.loc[coals['coals']==i,'Moi'].item()*var_coke_P2[i]/100
        coke_feeding_dmt_list_P2.append(x)
    coals['feeding_dmt_P2']=coke_feeding_dmt_list_P2

    #Fe_new
    coke_fe_list_P2=list()
    for i in coals['coals']:
        x=((var_coke_P2[i])-coals.loc[coals['coals']==i,'feeding_dmt_P2'].item())*coals.loc[coals['coals']==i,'Fe'].item()/100
        coke_fe_list_P2.append(x)
    coals['Fe(new)_P2']=coke_fe_list_P2

    #SiO2_new
    coke_SiO2_list_P2=list()
    for i in coals['coals']:
        x=((var_coke_P2[i])-coals.loc[coals['coals']==i,'feeding_dmt_P2'].item())*coals.loc[coals['coals']==i,'SiO2'].item()/100
        coke_SiO2_list_P2.append(x)
    coals['SiO2(new)_P2']=coke_SiO2_list_P2
    coke_SiO2_list_P2

    #Al2O3_new
    coke_Al2O3_list_P2=list()
    for i in coals['coals']:
        x=((var_coke_P2[i])-coals.loc[coals['coals']==i,'feeding_dmt_P2'].item())*coals.loc[coals['coals']==i,'Al2O3'].item()/100
        coke_Al2O3_list_P2.append(x)
    coals['Al2O3(new)_P2']=coke_Al2O3_list_P2

    #CaO_new
    coke_CaO_list_P2=list()
    for i in coals['coals']:
        x=((var_coke_P2[i])-coals.loc[coals['coals']==i,'feeding_dmt_P2'].item())*coals.loc[coals['coals']==i,'CaO'].item()/100
        coke_CaO_list_P2.append(x)
    coals['CaO(new)_P2']=coke_CaO_list_P2

    #MgO_new
    coke_MgO_list_P2=list()
    for i in coals['coals']:
        x=((var_coke_P2[i])-coals.loc[coals['coals']==i,'feeding_dmt_P2'].item())*coals.loc[coals['coals']==i,'MgO'].item()/100
        coke_MgO_list_P2.append(x)
    coals['MgO(new)_P2']=coke_MgO_list_P2

    #P_new
    coke_P_list_P2=list()
    for i in coals['coals']:
        x=((var_coke_P2[i])-coals.loc[coals['coals']==i,'feeding_dmt_P2'].item())*coals.loc[coals['coals']==i,'P'].item()/100
        coke_P_list_P2.append(x)
    coals['P(new)_P2']=coke_P_list_P2


    #S_new
    coke_S_list_P2=list()
    for i in coals['coals']:
        x=((var_coke_P2[i])-coals.loc[coals['coals']==i,'feeding_dmt_P2'].item())*coals.loc[coals['coals']==i,'S'].item()/100
        coke_S_list_P2.append(x)
    coals['S(new)_P2']=coke_S_list_P2


    # In[35]:


    #updating coals_P1 P1 chemistry values
    #Feeding DMT
    coke_feeding_dmt_list_P1=list()
    for i in coals_P1['coals_P1']:
        x=coals_P1.loc[coals_P1['coals_P1']==i,'Moi'].item()*var_coke_P1[i]/100
        coke_feeding_dmt_list_P1.append(x)
    coals_P1['feeding_dmt_P1']=coke_feeding_dmt_list_P1

    #Fe_new
    coke_fe_list_P1=list()
    for i in coals_P1['coals_P1']:
        x=((var_coke_P1[i])-coals_P1.loc[coals_P1['coals_P1']==i,'feeding_dmt_P1'].item())*coals_P1.loc[coals_P1['coals_P1']==i,'Fe'].item()/100
        coke_fe_list_P1.append(x)
    coals_P1['Fe(new)_P1']=coke_fe_list_P1

    #SiO2_new
    coke_SiO2_list_P1=list()
    for i in coals_P1['coals_P1']:
        x=((var_coke_P1[i])-coals_P1.loc[coals_P1['coals_P1']==i,'feeding_dmt_P1'].item())*coals_P1.loc[coals_P1['coals_P1']==i,'SiO2'].item()/100
        coke_SiO2_list_P1.append(x)
    coals_P1['SiO2(new)_P1']=coke_SiO2_list_P1
    coke_SiO2_list_P1

    #Al2O3_new
    coke_Al2O3_list_P1=list()
    for i in coals_P1['coals_P1']:
        x=((var_coke_P1[i])-coals_P1.loc[coals_P1['coals_P1']==i,'feeding_dmt_P1'].item())*coals_P1.loc[coals_P1['coals_P1']==i,'Al2O3'].item()/100
        coke_Al2O3_list_P1.append(x)
    coals_P1['Al2O3(new)_P1']=coke_Al2O3_list_P1

    #CaO_new
    coke_CaO_list_P1=list()
    for i in coals_P1['coals_P1']:
        x=((var_coke_P1[i])-coals_P1.loc[coals_P1['coals_P1']==i,'feeding_dmt_P1'].item())*coals_P1.loc[coals_P1['coals_P1']==i,'CaO'].item()/100
        coke_CaO_list_P1.append(x)
    coals_P1['CaO(new)_P1']=coke_CaO_list_P1

    #MgO_new
    coke_MgO_list_P1=list()
    for i in coals_P1['coals_P1']:
        x=((var_coke_P1[i])-coals_P1.loc[coals_P1['coals_P1']==i,'feeding_dmt_P1'].item())*coals_P1.loc[coals_P1['coals_P1']==i,'MgO'].item()/100
        coke_MgO_list_P1.append(x)
    coals_P1['MgO(new)_P1']=coke_MgO_list_P1

    #P_new
    coke_P_list_P1=list()
    for i in coals_P1['coals_P1']:
        x=((var_coke_P1[i])-coals_P1.loc[coals_P1['coals_P1']==i,'feeding_dmt_P1'].item())*coals_P1.loc[coals_P1['coals_P1']==i,'P'].item()/100
        coke_P_list_P1.append(x)
    coals_P1['P(new)_P1']=coke_P_list_P1


    #S_new
    coke_S_list_P1=list()
    for i in coals_P1['coals_P1']:
        x=((var_coke_P1[i])-coals_P1.loc[coals_P1['coals_P1']==i,'feeding_dmt_P1'].item())*coals_P1.loc[coals_P1['coals_P1']==i,'S'].item()/100
        coke_S_list_P1.append(x)
    coals_P1['S(new)_P1']=coke_S_list_P1


    # In[36]:


    #model=Lpmodellem('bmo optimisation',LpMinimize)
    var_bmo_P1=LpVariable.dicts('BMO Material quantity in Tons P1',bmo_materials['Material'],0,None,LpContinuous) 
    var_bmo_P2=LpVariable.dicts('BMO Material quantity in Tons P2',bmo_materials['Material'],0,None,LpContinuous) 


    # In[37]:


    #values cell

    pellet_production=650000*(no_of_days_input/no_of_days_in_month)
    DRI=150000
    values_P1=pd.read_csv('Data Input/Head Office/Blast Furnace/P1/Values P1.csv')
    values_P2=pd.read_csv('Data Input/Head Office/Blast Furnace/P2/Values P2.csv')


    # In[38]:


    #updating P1 chemistry values
    #Feeding DMT
    bmo_feeding_dmt_list_P1=list()
    for i in bmo_materials['Material']:
        x=bmo_materials.loc[bmo_materials['Material']==i,'Moi'].item()*var_bmo_P1[i]/100
        bmo_feeding_dmt_list_P1.append(x)
    bmo_materials['feeding_dmt_P1']=bmo_feeding_dmt_list_P1

    #Fe_new
    bmo_fe_list_P1=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P1[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P1'].item())*bmo_materials.loc[bmo_materials['Material']==i,'Fe'].item()/100
        bmo_fe_list_P1.append(x)
    bmo_materials['Fe(new)_P1']=bmo_fe_list_P1

    #SiO2_new
    bmo_SiO2_list_P1=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P1[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P1'].item())*bmo_materials.loc[bmo_materials['Material']==i,'SiO2'].item()/100
        bmo_SiO2_list_P1.append(x)
    bmo_materials['SiO2(new)_P1']=bmo_SiO2_list_P1
    bmo_SiO2_list_P1

    #Al2O3_new
    bmo_Al2O3_list_P1=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P1[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P1'].item())*bmo_materials.loc[bmo_materials['Material']==i,'Al2O3'].item()/100
        bmo_Al2O3_list_P1.append(x)
    bmo_materials['Al2O3(new)_P1']=bmo_Al2O3_list_P1

    #CaO_new
    bmo_CaO_list_P1=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P1[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P1'].item())*bmo_materials.loc[bmo_materials['Material']==i,'CaO'].item()/100
        bmo_CaO_list_P1.append(x)
    bmo_materials['CaO(new)_P1']=bmo_CaO_list_P1

    #MgO_new
    bmo_MgO_list_P1=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P1[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P1'].item())*bmo_materials.loc[bmo_materials['Material']==i,'MgO'].item()/100
        bmo_MgO_list_P1.append(x)
    bmo_materials['MgO(new)_P1']=bmo_MgO_list_P1

    #P_new
    bmo_P_list_P1=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P1[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P1'].item())*bmo_materials.loc[bmo_materials['Material']==i,'P'].item()/100
        bmo_P_list_P1.append(x)
    bmo_materials['P(new)_P1']=bmo_P_list_P1


    #S_new
    bmo_S_list_P1=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P1[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P1'].item())*bmo_materials.loc[bmo_materials['Material']==i,'S'].item()/100
        bmo_S_list_P1.append(x)
    bmo_materials['S(new)_P1']=bmo_S_list_P1


    # In[39]:


    #updating P2 chemistry values
    #Feeding DMT
    bmo_feeding_dmt_list_P2=list()
    for i in bmo_materials['Material']:
        x=bmo_materials.loc[bmo_materials['Material']==i,'Moi'].item()*var_bmo_P2[i]/100
        bmo_feeding_dmt_list_P2.append(x)
    bmo_materials['feeding_dmt_P2']=bmo_feeding_dmt_list_P2

    #Fe_new
    bmo_fe_list_P2=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P2[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P2'].item())*bmo_materials.loc[bmo_materials['Material']==i,'Fe'].item()/100
        bmo_fe_list_P2.append(x)
    bmo_materials['Fe(new)_P2']=bmo_fe_list_P2

    #SiO2_new
    bmo_SiO2_list_P2=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P2[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P2'].item())*bmo_materials.loc[bmo_materials['Material']==i,'SiO2'].item()/100
        bmo_SiO2_list_P2.append(x)
    bmo_materials['SiO2(new)_P2']=bmo_SiO2_list_P2
    bmo_SiO2_list_P2

    #Al2O3_new
    bmo_Al2O3_list_P2=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P2[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P2'].item())*bmo_materials.loc[bmo_materials['Material']==i,'Al2O3'].item()/100
        bmo_Al2O3_list_P2.append(x)
    bmo_materials['Al2O3(new)_P2']=bmo_Al2O3_list_P2

    #CaO_new
    bmo_CaO_list_P2=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P2[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P2'].item())*bmo_materials.loc[bmo_materials['Material']==i,'CaO'].item()/100
        bmo_CaO_list_P2.append(x)
    bmo_materials['CaO(new)_P2']=bmo_CaO_list_P2

    #MgO_new
    bmo_MgO_list_P2=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P2[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P2'].item())*bmo_materials.loc[bmo_materials['Material']==i,'MgO'].item()/100
        bmo_MgO_list_P2.append(x)
    bmo_materials['MgO(new)_P2']=bmo_MgO_list_P2

    #P_new
    bmo_P_list_P2=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P2[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P2'].item())*bmo_materials.loc[bmo_materials['Material']==i,'P'].item()/100
        bmo_P_list_P2.append(x)
    bmo_materials['P(new)_P2']=bmo_P_list_P2


    #S_new
    bmo_S_list_P2=list()
    for i in bmo_materials['Material']:
        x=(var_bmo_P2[i]-bmo_materials.loc[bmo_materials['Material']==i,'feeding_dmt_P2'].item())*bmo_materials.loc[bmo_materials['Material']==i,'S'].item()/100
        bmo_S_list_P2.append(x)
    bmo_materials['S(new)_P2']=bmo_S_list_P2


    # In[40]:


    #P1 Chemistry Totals (Integration part)
    Fe_total_P1=lpSum([bmo_materials['Fe(new)_P1']])+lpSum(mines.loc[mines['Plant']=='P1','Fe(new)_fines'])+lpSum(mines_bf_P1['Fe(new)_fines'])+lpSum(materials['Fe(new)_P1'])+lpSum(mines_lumps.loc[mines_lumps['Plant']=='P1','Fe(new)_lumps'])+lpSum(coals_P1['Fe(new)_P1'])
    SiO2_total_P1=lpSum([bmo_materials['SiO2(new)_P1']])+lpSum(mines.loc[mines['Plant']=='P1','SiO2(new)_fines'])+lpSum(mines_bf_P1['SiO2(new)_fines'])+lpSum(materials['SiO2(new)_P1'])+lpSum(mines_lumps.loc[mines_lumps['Plant']=='P1','SiO2(new)_lumps']) + lpSum(coals_P1['SiO2(new)_P1'])
    Al2O3_total_P1=lpSum([bmo_materials['Al2O3(new)_P1']])+lpSum(mines.loc[mines['Plant']=='P1','Al2O3(new)_fines'])+lpSum(mines_bf_P1['Al2O3(new)_fines'])+lpSum(materials['Al2O3(new)_P1'])+lpSum(mines_lumps.loc[mines_lumps['Plant']=='P1','Al2O3(new)_lumps']) + lpSum(coals_P1['Al2O3(new)_P1'])
    CaO_total_P1=lpSum([bmo_materials['CaO(new)_P1']])+lpSum(mines.loc[mines['Plant']=='P1','CaO(new)_fines'])+lpSum(mines_bf_P1['CaO(new)_fines'])+lpSum(materials['CaO(new)_P1'])+lpSum(mines_lumps.loc[mines_lumps['Plant']=='P1','CaO(new)_lumps'])+lpSum(coals_P1['CaO(new)_P1'])
    MgO_total_P1=lpSum([bmo_materials['MgO(new)_P1']])+lpSum(mines.loc[mines['Plant']=='P1','MgO(new)_fines'])+lpSum(mines_bf_P1['MgO(new)_fines'])+lpSum(materials['MgO(new)_P1'])+lpSum(mines_lumps.loc[mines_lumps['Plant']=='P1','MgO(new)_lumps'])+lpSum(coals_P1['MgO(new)_P1'])
    P_total_P1=lpSum([bmo_materials['P(new)_P1']])+ lpSum(coals_P1['P(new)_P1'])
    S_total_P1=lpSum([bmo_materials['S(new)_P1']])+lpSum(coals_P1['S(new)_P1'])

    #Si_total(using Fe)
    Si_total_P1=(Fe_total_P1*100/94.5)*values_P1.loc[values_P1['Parameter']=='Si %','Value'].item()/100

    #Kg/Thm
    Kg_by_Thm_P1=Fe_total_P1*100/94.5


    # In[41]:


    #P2 Chemistry Totals (Integration part)
    Fe_total_P2=lpSum([bmo_materials['Fe(new)_P2']])+lpSum(mines.loc[mines['Plant']=='P2','Fe(new)_fines'])+lpSum(mines_bf_P2['Fe(new)_fines'])+lpSum(materials['Fe(new)_P2'])+lpSum(mines_lumps.loc[mines_lumps['Plant']=='P2','Fe(new)_lumps']) +lpSum(coals['Fe(new)_P2'])
    SiO2_total_P2=lpSum([bmo_materials['SiO2(new)_P2']])+lpSum(mines.loc[mines['Plant']=='P2','SiO2(new)_fines'])+lpSum(mines_bf_P2['SiO2(new)_fines'])+lpSum(materials['SiO2(new)_P2'])+lpSum(mines_lumps.loc[mines_lumps['Plant']=='P2','SiO2(new)_lumps']) + lpSum(coals['SiO2(new)_P2'])
    Al2O3_total_P2=lpSum([bmo_materials['Al2O3(new)_P2']])+lpSum(mines.loc[mines['Plant']=='P2','Al2O3(new)_fines'])+lpSum(mines_bf_P2['Al2O3(new)_fines'])+lpSum(materials['Al2O3(new)_P2'])+lpSum(mines_lumps.loc[mines_lumps['Plant']=='P2','Al2O3(new)_lumps']) + lpSum(coals['Al2O3(new)_P2'])
    CaO_total_P2=lpSum([bmo_materials['CaO(new)_P2']])+lpSum(mines.loc[mines['Plant']=='P2','CaO(new)_fines'])+lpSum(mines_bf_P2['CaO(new)_fines'])+lpSum(materials['CaO(new)_P2'])+lpSum(mines_lumps.loc[mines_lumps['Plant']=='P2','CaO(new)_lumps'])+lpSum(coals['CaO(new)_P2'])
    MgO_total_P2=lpSum([bmo_materials['MgO(new)_P2']])+lpSum(mines.loc[mines['Plant']=='P2','MgO(new)_fines'])+lpSum(mines_bf_P2['MgO(new)_fines'])+lpSum(materials['MgO(new)_P2'])+lpSum(mines_lumps.loc[mines_lumps['Plant']=='P2','MgO(new)_lumps'])+lpSum(coals['MgO(new)_P2'])
    P_total_P2=lpSum([bmo_materials['P(new)_P2']]) + lpSum(coals['P(new)_P2'])
    S_total_P2=lpSum([bmo_materials['S(new)_P2']])+lpSum(coals['S(new)_P2'])

    #Si_total(using Fe)
    Si_total_P2=(Fe_total_P2*99/95)*values_P2.loc[values_P2['Parameter']=='Si %','Value'].item()/100

    #Kg/Thm
    Kg_by_Thm_P2=Fe_total_P2*99/95


    # In[42]:


    #bmo chemistry constraints input
    bmo_chemistry_table_P1=pd.read_csv('Data Input/Plant/Blast Furnace/P1/BMO Chemistry Constraints_P1.csv')
    bmo_chemistry_table_P2=pd.read_csv('Data Input/Plant/Blast Furnace/P2/BMO Chemistry Constraints_P2.csv')
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Min'].isna(),'Min']=0
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Max'].isna(),'Max']=0
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Min'].isna(),'Min']=0
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Max'].isna(),'Max']=0


    # In[45]:


    #P1
    #bmo_chemistry_constraints
    #updating min(total)_P1
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','min(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Min'].item()*Kg_by_Thm_P1]
    #bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Basicity (B1)','min(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Basicity (B1)','Min'].item()*P_total_P1]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Amount','min(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Amount','Min'].item()*Kg_by_Thm_P1]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Phosphorus','min(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Phosphorus','Min'].item()*Kg_by_Thm_P1]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='MgO','min(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='MgO','Min'].item()*(CaO_total_P1+MgO_total_P1+(SiO2_total_P1-(Si_total_P1/0.466))+Al2O3_total_P1)/(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Min'].item()/1000)]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Al2O3','min(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Al2O3','Min'].item()*(CaO_total_P1+MgO_total_P1+(SiO2_total_P1-(Si_total_P1/0.466))+Al2O3_total_P1)/(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Min'].item()/1000)]

    #updating max(total)_P1
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Amount','max(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Amount','Max'].item()*Kg_by_Thm_P1]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Sulphur in HM','max(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Sulphur in HM','Max'].item()*Kg_by_Thm_P1/100000]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Phosphorus','max(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Phosphorus','Max'].item()*Kg_by_Thm_P1]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='MgO','max(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='MgO','Max'].item()*(CaO_total_P1+MgO_total_P1+(SiO2_total_P1-(Si_total_P1/0.466))+Al2O3_total_P1)/(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Min'].item()/1000)]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Al2O3','max(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Al2O3','Max'].item()*(CaO_total_P1+MgO_total_P1+(SiO2_total_P1-(Si_total_P1/0.466))+Al2O3_total_P1)/(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Min'].item()/1000)]

    #P2
    #bmo_chemistry_constraints
    #updating min(total)_P2
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','min(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Min'].item()*Kg_by_Thm_P2]
    #bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Basicity (B1)','min(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Basicity (B1)','Min'].item()*P_total_P2]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Amount','min(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Amount','Min'].item()*Kg_by_Thm_P2]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Phosphorus','min(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Phosphorus','Min'].item()*Kg_by_Thm_P2]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='MgO','min(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='MgO','Min'].item()*(CaO_total_P2+MgO_total_P2+(SiO2_total_P2-(Si_total_P2/0.466))+Al2O3_total_P2)/(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Min'].item()/1000)]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Al2O3','min(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Al2O3','Min'].item()*(CaO_total_P2+MgO_total_P2+(SiO2_total_P2-(Si_total_P2/0.466))+Al2O3_total_P2)/(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Min'].item()/1000)]

    #updating max(total)_P2
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Amount','max(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Amount','Max'].item()*Kg_by_Thm_P2]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Sulphur in HM','max(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Sulphur in HM','Max'].item()*Kg_by_Thm_P2/100000]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Phosphorus','max(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Phosphorus','Max'].item()*Kg_by_Thm_P2]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='MgO','max(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='MgO','Max'].item()*(CaO_total_P2+MgO_total_P2+(SiO2_total_P2-(Si_total_P2/0.466))+Al2O3_total_P2)/(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Min'].item()/1000)]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Al2O3','max(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Al2O3','Max'].item()*(CaO_total_P2+MgO_total_P2+(SiO2_total_P2-(Si_total_P2/0.466))+Al2O3_total_P2)/(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Min'].item()/1000)]


    # In[46]:


    #actual_values(constraints LHS) 
    #P1
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Actual Values_P1']=[Fe_total_P1*1000]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Amount','Actual Values_P1']=[((CaO_total_P1+MgO_total_P1+(SiO2_total_P1-(Si_total_P1/0.466))+Al2O3_total_P1)/(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Min'].item()/1000))*1000]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Sulphur in HM','Actual Values_P1']=[(S_total_P1/1000)-0.217]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Phosphorus','Actual Values_P1']=[P_total_P1*100]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='CaO','Actual Values_P1']=[(CaO_total_P1-(56/32*0.7*(CaO_total_P1+MgO_total_P1+(SiO2_total_P1-(Si_total_P1/0.466))+Al2O3_total_P1)/(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Min'].item()/1000)/100))*100]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='MgO','Actual Values_P1']=[MgO_total_P1*100]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='SiO2','Actual Values_P1']=[(SiO2_total_P1-(Si_total_P1/0.466))*100]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Al2O3','Actual Values_P1']=[Al2O3_total_P1*100]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Basicity (B1)','Actual Values_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='CaO','Actual Values_P1'].item()]

    #P2
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Actual Values_P2']=[Fe_total_P2*1000]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Amount','Actual Values_P2']=[((CaO_total_P2+MgO_total_P2+(SiO2_total_P2-(Si_total_P2/0.466))+Al2O3_total_P2)/(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Min'].item()/1000))*1000]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Sulphur in HM','Actual Values_P2']=[(S_total_P2/1000)-0.217]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Phosphorus','Actual Values_P2']=[P_total_P2*100]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='CaO','Actual Values_P2']=[(CaO_total_P2-(56/32*0.7*(CaO_total_P2+MgO_total_P2+(SiO2_total_P2-(Si_total_P2/0.466))+Al2O3_total_P2)/(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Min'].item()/1000)/100))*100]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='MgO','Actual Values_P2']=[MgO_total_P2*100]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='SiO2','Actual Values_P2']=[(SiO2_total_P2-(Si_total_P2/0.466))*100]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Al2O3','Actual Values_P2']=[Al2O3_total_P2*100]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Basicity (B1)','Actual Values_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='CaO','Actual Values_P2'].item()]


    # In[47]:


    #remaining related min/max(total) P1
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Basicity (B1)','min(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Basicity (B1)','Min'].item()*bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='SiO2','Actual Values_P1'].item()]
    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Basicity (B1)','max(total)_P1']=[bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Basicity (B1)','Max'].item()*bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='SiO2','Actual Values_P1'].item()]


    # In[48]:


    #remaining related min/max(total) P2
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Basicity (B1)','min(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Basicity (B1)','Min'].item()*bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='SiO2','Actual Values_P2'].item()]
    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Basicity (B1)','max(total)_P2']=[bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Basicity (B1)','Max'].item()*bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='SiO2','Actual Values_P2'].item()]


    # In[49]:


    #Coke inputs
    coke_input_P1=pd.read_csv('Data Input/Plant/Coke Plant/P1/Coke Input P1.csv')
    coke_input_P2=pd.read_csv('Data Input/Plant/Coke Plant/P2/Coke Input P2.csv')

    #sinter pellet lump inputs
    spl_input_P1=pd.read_csv('Data Input/Plant/Blast Furnace/P1/Sinter Pellet Lump P1.csv')
    spl_input_P2=pd.read_csv('Data Input/Plant/Blast Furnace/P2/Sinter Pellet Lump P2.csv')


    # In[50]:


    bmo_chemistry_table_P1['key_min']=[bf_constraint_dict_P1['Fe(lower)'].get(),bf_constraint_dict_P1['Slag Basicity (B1)(lower)'].get(),
                                          bf_constraint_dict_P1['Slag Amount(lower)'].get(),bf_constraint_dict_P1['Sulphur in HM(lower)'].get(),
                                          bf_constraint_dict_P1['Phosphorus(lower)'].get(),bf_constraint_dict_P1['CaO(lower)'].get(),
                                          bf_constraint_dict_P1['MgO(lower)'].get(),bf_constraint_dict_P1['SiO2(lower)'].get(),
                                          bf_constraint_dict_P1['Al2O3(lower)'].get()]
    bmo_chemistry_table_P1['key_max']=[bf_constraint_dict_P1['Fe(upper)'].get(),bf_constraint_dict_P1['Slag Basicity (B1)(upper)'].get(),
                                          bf_constraint_dict_P1['Slag Amount(upper)'].get(),bf_constraint_dict_P1['Sulphur in HM(upper)'].get(),
                                          bf_constraint_dict_P1['Phosphorus(upper)'].get(),bf_constraint_dict_P1['CaO(upper)'].get(),
                                          bf_constraint_dict_P1['MgO(upper)'].get(),bf_constraint_dict_P1['SiO2(upper)'].get(),
                                          bf_constraint_dict_P1['Al2O3(upper)'].get()]

    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['min(total)_P1'].isna(),'min(total)_P1']=0

    bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['max(total)_P1'].isna(),'max(total)_P1']=0

    bmo_chemistry_table_P2['key_min']=[bf_constraint_dict_P2['Fe(lower)'].get(),bf_constraint_dict_P2['Slag Basicity (B1)(lower)'].get(),
                                          bf_constraint_dict_P2['Slag Amount(lower)'].get(),bf_constraint_dict_P2['Sulphur in HM(lower)'].get(),
                                          bf_constraint_dict_P2['Phosphorus(lower)'].get(),bf_constraint_dict_P2['CaO(lower)'].get(),
                                          bf_constraint_dict_P2['MgO(lower)'].get(),bf_constraint_dict_P2['SiO2(lower)'].get(),
                                          bf_constraint_dict_P2['Al2O3(lower)'].get()]
     
    bmo_chemistry_table_P2['key_max']=[bf_constraint_dict_P2['Fe(upper)'].get(),bf_constraint_dict_P2['Slag Basicity (B1)(upper)'].get(),
                                          bf_constraint_dict_P2['Slag Amount(upper)'].get(),bf_constraint_dict_P2['Sulphur in HM(upper)'].get(),
                                          bf_constraint_dict_P2['Phosphorus(upper)'].get(),bf_constraint_dict_P2['CaO(upper)'].get(),
                                          bf_constraint_dict_P2['MgO(upper)'].get(),bf_constraint_dict_P2['SiO2(upper)'].get(),
                                          bf_constraint_dict_P2['Al2O3(upper)'].get()]

    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['min(total)_P2'].isna(),'min(total)_P2']=0

    bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['max(total)_P2'].isna(),'max(total)_P2']=0


    # In[51]:


    zxc=list()
    #bmo_constraints table
    #P1
    for i in bmo_chemistry_table_P1['Parameters']:
        if bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']==i,'key_min'].item()==1:
            a=bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']==i,'Actual Values_P1'].item()>=bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']==i,'min(total)_P1'].item()
            model+=a
    for i in bmo_chemistry_table_P1['Parameters']:
        if bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']==i,'key_max'].item()==1:
            a=bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']==i,'Actual Values_P1'].item()<=bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']==i,'max(total)_P1'].item()        
            model+=a

    #P2
    for i in bmo_chemistry_table_P2['Parameters']:
        if bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']==i,'key_min'].item()==1:
            a=bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']==i,'Actual Values_P2'].item()>=bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']==i,'min(total)_P2'].item()
            model+=a
    for i in bmo_chemistry_table_P2['Parameters']:
        if bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']==i,'key_max'].item()==1:
            a=bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']==i,'Actual Values_P2'].item()<=bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']==i,'max(total)_P2'].item()        
            model+=a                


    # In[52]:


    #total_production constraint
    #P1
    #values_P1.loc[values_P1['Parameter']=='Total Production','Value'].item()=hot_metal_input_P1
    #values_P2.loc[values_P2['Parameter']=='Total Production','Value'].item()=hot_metal_input_P2
    if bf_constraint_dict_P1['Hot Metal Production'].get()==1:
        model+=lpSum(Kg_by_Thm_P1)==hot_metal_input_P1
    #P2
    if bf_constraint_dict_P2['Hot Metal Production'].get()==1:
        model+=lpSum(Kg_by_Thm_P2)==hot_metal_input_P2


    # In[53]:


    lpSum(var_lumps)#sinter pellet lump constraints
    #P1
    #to be integrated
    expected_production_sinter_P1=[expected_production_P1]

    #
    total_lumps_P1=list()
    for i in var_lumps:
        if mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()=='P1':
            total_lumps_P1.append(var_lumps[i])
    lpSum(total_lumps_P1)  

    total_lumps_P1=[lpSum(total_lumps_P1)]
    spl_input_P1.loc[spl_input_P1['Parameter']=='Sinter','Actual']=expected_production_sinter_P1
    spl_input_P1.loc[spl_input_P1['Parameter']=='Pellet','Actual']=[(lpSum(var_fines_bf_P1))/pellet_conversion_bf_P1]
    spl_input_P1.loc[spl_input_P1['Parameter']=='Lump','Actual']=total_lumps_P1
    total_ibm_P1=lpSum(spl_input_P1['Actual'])
    #updating min 
    for i in spl_input_P1['Parameter']:
        spl_input_P1.loc[spl_input_P1['Parameter']==i,'min(wtd)']=[total_ibm_P1*spl_input_P1.loc[spl_input_P1['Parameter']==i,'Min'].item()]
    #updating max
    for i in spl_input_P1['Parameter']:
        spl_input_P1.loc[spl_input_P1['Parameter']==i,'max(wtd)']=[total_ibm_P1*spl_input_P1.loc[spl_input_P1['Parameter']==i,'Max'].item()]

    for i in spl_input_P1['Parameter']:
        a=spl_input_P1.loc[spl_input_P1['Parameter']==i,'Actual'].item()<=spl_input_P1.loc[spl_input_P1['Parameter']==i,'max(wtd)'].item()        
        if bf_constraint_dict_P1['Sinter Pellet Lump Ratio(upper)'].get()==1:
            model+=a  

    for i in spl_input_P1['Parameter']:
        a=spl_input_P1.loc[spl_input_P1['Parameter']==i,'Actual'].item()>=spl_input_P1.loc[spl_input_P1['Parameter']==i,'min(wtd)'].item()        
        if bf_constraint_dict_P1['Sinter Pellet Lump Ratio(lower)'].get()==1:
            model+=a      


    # In[54]:


    lpSum(var_lumps)#sinter pellet lump constraints
    #P2
    #to be integrated
    expected_production_sinter_P2=[expected_production_P2]

    #
    total_lumps_P2=list()
    for i in var_lumps:
        if mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()=='P2':
            total_lumps_P2.append(var_lumps[i])
    lpSum(total_lumps_P2)  

    total_lumps_P2=[lpSum(total_lumps_P2)]
    spl_input_P2.loc[spl_input_P2['Parameter']=='Sinter','Actual']=expected_production_sinter_P2
    spl_input_P2.loc[spl_input_P2['Parameter']=='Pellet','Actual']=[(lpSum(var_fines_bf_P2))/pellet_conversion_bf_P2]
    spl_input_P2.loc[spl_input_P2['Parameter']=='Lump','Actual']=total_lumps_P2
    total_ibm_P2=lpSum(spl_input_P2['Actual'])
    #updating min 
    for i in spl_input_P2['Parameter']:
        spl_input_P2.loc[spl_input_P2['Parameter']==i,'min(wtd)']=[total_ibm_P2*spl_input_P2.loc[spl_input_P2['Parameter']==i,'Min'].item()]
    #updating max
    for i in spl_input_P2['Parameter']:
        spl_input_P2.loc[spl_input_P2['Parameter']==i,'max(wtd)']=[total_ibm_P2*spl_input_P2.loc[spl_input_P2['Parameter']==i,'Max'].item()]

    for i in spl_input_P2['Parameter']:
        a=spl_input_P2.loc[spl_input_P2['Parameter']==i,'Actual'].item()<=spl_input_P2.loc[spl_input_P2['Parameter']==i,'max(wtd)'].item()        
        if bf_constraint_dict_P2['Sinter Pellet Lump Ratio(upper)'].get()==1:
            model+=a  

    for i in spl_input_P2['Parameter']:
        a=spl_input_P2.loc[spl_input_P2['Parameter']==i,'Actual'].item()>=spl_input_P2.loc[spl_input_P2['Parameter']==i,'min(wtd)'].item()        
        if bf_constraint_dict_P2['Sinter Pellet Lump Ratio(lower)'].get()==1:
            model+=a      


    # In[55]:


    #qw=list()
    #coke constraints
    #P1
    for i in coke_input_P1['Parameter']:
        if coke_input_P1.loc[coke_input_P1['Parameter']==i,'Parameter'].item()!='Coke (C1)':
            coke_input_P1.loc[coke_input_P1['Parameter']==i,'Actual']=[Kg_by_Thm_P1*coke_input_P1.loc[coke_input_P1['Parameter']==i,'Limit'].item()/1000]

    for i in coke_input_P1['Parameter']:
        if coke_input_P1.loc[coke_input_P1['Parameter']==i,'Parameter'].item()!='Coke (C1)':
            if bf_constraint_dict_P1['Coke Inputs'].get()==1:
                model+=var_bmo_P1[i]>=coke_input_P1.loc[coke_input_P1['Parameter']==i,'Actual'].item()



    # In[56]:


    #P2
    for i in coke_input_P2['Parameter']:
        if coke_input_P2.loc[coke_input_P2['Parameter']==i,'Parameter'].item()!='Coke (C1)':
            coke_input_P2.loc[coke_input_P2['Parameter']==i,'Actual']=[Kg_by_Thm_P2*coke_input_P2.loc[coke_input_P2['Parameter']==i,'Limit'].item()/1000]

    for i in coke_input_P2['Parameter']:
        if coke_input_P2.loc[coke_input_P2['Parameter']==i,'Parameter'].item()!='Coke (C1)':
            if bf_constraint_dict_P2['Coke Inputs'].get()==1:
                model+=var_bmo_P2[i]>=coke_input_P2.loc[coke_input_P2['Parameter']==i,'Actual'].item()



    # In[57]:


    if bf_constraint_dict_P1['Coke Plant(C1)'].get()==1:
        model+=lpSum(var_coke_P1[i]*(((100-(coal_data_P1.loc[(coal_data_P1['category']==i),'tm'])-((100-(coal_data_P1.loc[(coal_data_P1['category']==i),'tm']))*((coal_data_P1.loc[(coal_data_P1['category']==i),'vm'])/100))-burning_loss_P1+end_vm_P1_P1+end_moisture_P1)/100).item()) for i in var_coke_P1)>=[Kg_by_Thm_P1*coke_input_P1.loc[coke_input_P1['Parameter']=='Coke (C1)','Limit'].item()/1000]


    # In[58]:


    if bf_constraint_dict_P2['Coke Plant(C1)'].get()==1:
        model+=lpSum(var_coke_P2[i]*(((100-(coal_data.loc[(coal_data['category']==i),'tm'])-((100-(coal_data.loc[(coal_data['category']==i),'tm']))*((coal_data.loc[(coal_data['category']==i),'vm'])/100))-burning_loss+end_vm+end_moisture)/100).item()) for i in var_coke_P2)>=[Kg_by_Thm_P2*coke_input_P2.loc[coke_input_P2['Parameter']=='Coke (C1)','Limit'].item()/1000]


    # In[59]:


    #Pellet production
    pellet_production=650000*(no_of_days_input/no_of_days_in_month)
    pellet_used_in_bf=spl_input_P1.loc[spl_input_P1['Parameter']=='Pellet','Actual'].item()+spl_input_P2.loc[spl_input_P2['Parameter']=='Pellet','Actual'].item()
    model+=pellet_used_in_bf>=0
    #dri
    dri_input_df=pd.read_csv('Data Input/Plant/DRI Plant/DRI Inputs.csv')
    sponge_iron_P1=dri_input_df.iloc[0,2].item()
    sponge_iron_P2=dri_input_df.iloc[3,2].item()
    pellet_consumption_per_ton_P1=dri_input_df.iloc[1,2].item()
    pellet_consumption_per_ton_P2=dri_input_df.iloc[4,2].item()
    total_pellet_requirement_P1=sponge_iron_P1*pellet_consumption_per_ton_P1
    total_pellet_requirement_P2=sponge_iron_P2*pellet_consumption_per_ton_P2
    DRI=lpSum(var_fines_dri_P2)+lpSum(var_fines_dri_P1)
    model+=DRI>=0
    market_sale=pellet_production-(pellet_used_in_bf+DRI)
    model+=market_sale>=0
    total_pellet=pellet_used_in_bf+DRI+market_sale
    #pellet_contribution_P1=market_sale*values_P1.loc[values_P1['Parameter']=='Realization from pellet Sale','Value'].item()
    pellet_contribution=market_sale*values_P2.loc[values_P2['Parameter']=='Realization from pellet Sale','Value'].item()
    if bf_constraint_dict_P1['Pellet Production'].get()==1 & bf_constraint_dict_P1['Pellet Production'].get()==1:
        model+=total_pellet<=pellet_production


    # In[60]:


    #charge capacity_constraint
    #P1
    charge_capacity_P1=(spl_input_P1.loc[spl_input_P1['Parameter']=='Sinter','Actual'].item()+spl_input_P1.loc[spl_input_P1['Parameter']=='Pellet','Actual'].item()+lpSum(var_coke_P1)+spl_input_P1.loc[spl_input_P1['Parameter']=='Lump','Actual'].item()+lpSum(var_bmo_P1))/(values_P1.loc[values_P1['Parameter']=='Charge rate','Value'].item()*no_of_days_input*24)
    if bf_constraint_dict_P1['Charge Capacity'].get()==1:
        model+=charge_capacity_P1<=values_P1.loc[values_P1['Parameter']=='Charge Capacity','Value'].item()

    #P2
    charge_capacity_P2=(spl_input_P2.loc[spl_input_P2['Parameter']=='Sinter','Actual'].item()+spl_input_P2.loc[spl_input_P2['Parameter']=='Pellet','Actual'].item()+lpSum(var_coke_P2)+spl_input_P2.loc[spl_input_P2['Parameter']=='Lump','Actual'].item()+lpSum(var_bmo_P2))/(values_P2.loc[values_P2['Parameter']=='Charge rate','Value'].item()*no_of_days_input*24)
    if bf_constraint_dict_P2['Charge Capacity'].get()==1:
        model+=charge_capacity_P2<=values_P2.loc[values_P2['Parameter']=='Charge Capacity','Value'].item()


    # In[61]:


    #cost barbil
    barbil_cost=lpSum(var_fines_bf_P1[i]*mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Final Cost Fines'].item()/pellet_conversion_bf_P1 for i in var_fines_bf_P1)


    # In[62]:


    #Cost objective P1
    #updating lumps cost
    smo_obj_lumps_P1=list()
    for i in var_lumps:
        if mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()=='P1':
            smo_obj_lumps_P1.append(var_lumps[i]*mines_lumps.loc[mines_lumps['possibilities']==i,'Total Cost Lumps'].item())
    lpSum(smo_obj_lumps_P1)

    cost_objective_P1_bf=lpSum(var_fines_bf_P1[i]*mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Final Cost Fines'].item()/pellet_conversion_bf_P1 for i in var_fines_bf_P1)
    cost_objective_P1_dri=lpSum(var_fines_dri_P1[i]*mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Final Cost Fines'].item()/pellet_conversion_dri_P1 for i in var_fines_dri_P1)
    bmo_materials_cost_obj_P1=list()
    for i in var_bmo_P1:
        bmo_materials_cost_obj_P1.append(var_bmo_P1[i]*bmo_materials.loc[bmo_materials['Material']==i,'Cost/MT'].item())
    bmo_total_cost_P1=lpSum(bmo_materials_cost_obj_P1)
    int_total_cost_P1=cost_objective_P1_bf +bmo_total_cost_P1+Total_Cost_of_Usable_Coke_at_Blast_Furnace_req_P1+Total_Cost_P1+lpSum(smo_obj_lumps_P1)
    int_total_realization_P1=Kg_by_Thm_P1*values_P1.loc[values_P1['Parameter']=='Realization per ton of HM','Value'].item()
    int_total_contribution_P1=int_total_realization_P1-int_total_cost_P1


    # In[63]:


    #Cost objective P2
    #updating lumps list
    smo_obj_lumps_P2=list()
    for i in var_lumps:
        if mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()=='P2':
            smo_obj_lumps_P2.append(var_lumps[i]*mines_lumps.loc[mines_lumps['possibilities']==i,'Total Cost Lumps'].item())
    lpSum(smo_obj_lumps_P2)

    cost_objective_P2_bf=lpSum(var_fines_bf_P2[i]*mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Final Cost Fines'].item()/pellet_conversion_bf_P2 for i in var_fines_bf_P2)
    cost_objective_P2_dri=lpSum(var_fines_dri_P2[i]*mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Final Cost Fines'].item()/pellet_conversion_dri_P2 for i in var_fines_dri_P2)

    bmo_materials_cost_obj_P2=list()
    for i in var_bmo_P2:
        bmo_materials_cost_obj_P2.append(var_bmo_P2[i]*bmo_materials.loc[bmo_materials['Material']==i,'Cost/MT'].item())
    bmo_total_cost_P2=lpSum(bmo_materials_cost_obj_P2)
    int_total_cost_P2=cost_objective_P2_bf+bmo_total_cost_P2+Total_Cost_of_Usable_Coke_at_Blast_Furnace_req+Total_Cost_P2+lpSum(smo_obj_lumps_P2)
    int_total_realization_P2=Kg_by_Thm_P2*values_P2.loc[values_P2['Parameter']=='Realization per ton of HM','Value'].item()
    int_total_contribution_P2=int_total_realization_P2-int_total_cost_P2


    # In[64]:


    #demand_constraint_fines
    b=lpSum(var_fines_bf_P1)+lpSum(var_fines_dri_P1)+lpSum(var_fines_bf_P2)+lpSum(var_fines_dri_P2)+lpSum(var_fines_pellet_market)
    if bf_constraint_dict_P1['Pellet Demand'].get()==1 & bf_constraint_dict_P1['Pellet Demand'].get()==1:
        model+=b>=pellet_production*((pellet_conversion_bf_P1+pellet_conversion_bf_P2+pellet_conversion_dri_P1+pellet_conversion_dri_P2+pellet_conversion_market))/5


    # In[65]:


    #pellet consumtion constraints
    P2_target=dri_input_P2
    P1_target=dri_input_P1
    Pellet_Sp_Cons_Ton_P2_target=dri_input_df.iloc[4,2].item()
    Pellet_Sp_Cons_Ton_P1_target=dri_input_df.iloc[1,2].item()
    Pellet_Sp_Cons_Ton_P1_model=dri_input_df.iloc[1,2].item()
    Pellet_Sp_Cons_Ton_P2_model=dri_input_df.iloc[4,2].item()
    Other_Cost_per_ton_P1=dri_input_df.iloc[2,2].item()
    Other_Cost_per_ton_P2=dri_input_df.iloc[5,2].item()
    model+=lpSum(var_fines_dri_P1)<=P1_target*Pellet_Sp_Cons_Ton_P1_target
    model+=lpSum(var_fines_dri_P2)<=P2_target*Pellet_Sp_Cons_Ton_P2_target
    model+=(lpSum(var_fines_dri_P1)/Pellet_Sp_Cons_Ton_P1_model)==P1_target
    model+=(lpSum(var_fines_dri_P2)/Pellet_Sp_Cons_Ton_P2_model)==P2_target


    # In[66]:


    #cost objective Barbil
    cost_objective_barbil_list=list()
    for i in var_lumps:
        if mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()=='Barbil':
            cost_objective_barbil_list.append(var_lumps[i]*mines_lumps.loc[mines_lumps['possibilities']==i,'Total Cost Lumps'].item())
    for i in var_fines_pellet_market:
        cost_objective_barbil_list.append(var_fines_pellet_market[i]*mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Total Cost Fines'].item()/pellet_conversion_market)    
    lpSum(cost_objective_barbil_list)
    cost_objective_barbil=lpSum(cost_objective_barbil_list)


    # In[67]:


    #cost calculation barbil
    #barbil to P1
    total_cost_P1_dri_list=list()
    for i in var_fines_dri_P1:
        total_cost_P1_dri_list.append(var_fines_dri_P1[i]*mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Final Cost Fines'].item()/pellet_conversion_dri_P1)
    total_cost_P1_dri=lpSum(total_cost_P1_dri_list)+(lpSum(var_fines_dri_P1)/Pellet_Sp_Cons_Ton_P1_model*Other_Cost_per_ton_P1)  
    total_realisation_P1_dri=((lpSum(var_fines_dri_P1)/Pellet_Sp_Cons_Ton_P1_model))*values_P1.loc[values_P1['Parameter']=='Realization per ton of HM','Value'].item()
    total_contribution_P1_dri=total_realisation_P1_dri-total_cost_P1_dri

    #barbil to P2
    total_cost_P2_dri_list=list()
    for i in var_fines_dri_P2:
        total_cost_P2_dri_list.append(var_fines_dri_P2[i]*mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Final Cost Fines'].item()/pellet_conversion_dri_P2)
    total_cost_P2_dri=lpSum(total_cost_P2_dri_list)+(lpSum(var_fines_dri_P2)/Pellet_Sp_Cons_Ton_P2_model*Other_Cost_per_ton_P2)  
    total_realisation_P2_dri=((lpSum(var_fines_dri_P2)/Pellet_Sp_Cons_Ton_P2_model))*values_P2.loc[values_P2['Parameter']=='Realization per ton of HM','Value'].item()
    total_contribution_P2_dri=total_realisation_P2_dri-total_cost_P2_dri


    # In[68]:


    model+=int_total_contribution_P2+int_total_contribution_P1+pellet_contribution-lpSum(cost_objective_barbil_list)+total_contribution_P1_dri+total_contribution_P2_dri
    #int_total_contribution_P2+int_total_contribution_P1+pellet_contribution-lpSum(cost_objective_barbil_list)+total_contribution_P1_dri+total_contribution_P2_dri
    #pellet_contribution+int_total_contribution_P2+int_total_contribution_P1
    #+int_total_contribution_P2+int_total_contribution_P1+Total_Cost_of_Usable_Coke_at_Blast_Furnace
    #int_total_contribution_P2+int_total_contribution_P1
    #pellet_contribution
    #+int_total_contribution_P2+int_total_contribution_P1
    #+int_total_contribution_P2
    #+pellet_contribution_P2


    # In[69]:


    model


    # In[70]:


    model.solve()


    # In[71]:


    model.writeLP('Pulp Data/Integrated(sinter+smo+bmo+coke+pellet).lp')
    print("Status:",LpStatus[model.status])
    for v in model.variables():
        print(v.name,"",v.varValue)
    final_cost=value(model.objective)
    print(" Total contributon is ",final_cost)
    messagebox.showinfo("Status",LpStatus[model.status])

    def output(self):
        op_list=list()
        for v, coefficient in lpSum(self).items():
            op_list.append(coefficient*v.varValue)
        op_total=lpSum(op_list).value()
        return op_total
    #P1
    sinter_P1=output(spl_input_P1.loc[spl_input_P1['Parameter']=='Sinter','Actual'])
    sinter_P1_max_wtd=output(spl_input_P1.loc[spl_input_P1['Parameter']=='Sinter','max(wtd)'])
    sinter_P1_min_wtd=output(spl_input_P1.loc[spl_input_P1['Parameter']=='Sinter','min(wtd)'])
    pellet_P1=output(spl_input_P1.loc[spl_input_P1['Parameter']=='Pellet','Actual'])
    pellet_P1_max_wtd=output(spl_input_P1.loc[spl_input_P1['Parameter']=='Pellet','max(wtd)'])
    pellet_P1_min_wtd=output(spl_input_P1.loc[spl_input_P1['Parameter']=='Pellet','min(wtd)'])
    lump_P1=output(spl_input_P1.loc[spl_input_P1['Parameter']=='Lump','Actual'])
    lump_P1_max_wtd=output(spl_input_P1.loc[spl_input_P1['Parameter']=='Lump','max(wtd)'])
    hot_metal_prodction_P1=output(Kg_by_Thm_P1)
    charge_capacity_P1_op=output(charge_capacity_P1)
    dri_pellet_P1_model=output(lpSum(var_fines_dri_P1)/Pellet_Sp_Cons_Ton_P1_model)

    #P2
    sinter_P2=output(spl_input_P2.loc[spl_input_P2['Parameter']=='Sinter','Actual'])
    sinter_P2_max_wtd=output(spl_input_P2.loc[spl_input_P2['Parameter']=='Sinter','max(wtd)'])
    sinter_P2_min_wtd=output(spl_input_P2.loc[spl_input_P2['Parameter']=='Sinter','min(wtd)'])
    pellet_P2=output(spl_input_P2.loc[spl_input_P2['Parameter']=='Pellet','Actual'])
    pellet_P2_max_wtd=output(spl_input_P2.loc[spl_input_P2['Parameter']=='Pellet','max(wtd)'])
    pellet_P2_min_wtd=output(spl_input_P2.loc[spl_input_P2['Parameter']=='Pellet','min(wtd)'])
    lump_P2=output(spl_input_P2.loc[spl_input_P2['Parameter']=='Lump','Actual'])
    lump_P2_max_wtd=output(spl_input_P2.loc[spl_input_P2['Parameter']=='Lump','max(wtd)'])

    hot_metal_prodction_P2=output(Kg_by_Thm_P2)
    charge_capacity_P2_op=output(charge_capacity_P2)
    dri_pellet_P2_model=output(lpSum(var_fines_dri_P2)/Pellet_Sp_Cons_Ton_P2_model)


    # In[81]:


    scenario_P1_dict={'Parameter':[],'Blast Furnace(HM)':[],'Coke Oven':[],'Sinter':[],'DRI':[]}
    scenario_P1_df=pd.DataFrame(data=scenario_P1_dict)
    scenario_P1_df['Parameter']=['Excpected Production','Total Realisation','Total Contribution']
    scenario_P1_df['Blast Furnace(HM)']=[Kg_by_Thm_P1.value(),int_total_realization_P1.value(),int_total_contribution_P1.value()]
    scenario_P1_df['Coke Oven']=[output(lpSum(var_coke_P1[i]*(((100-(coal_data_P1.loc[(coal_data_P1['category']==i),'tm'])-((100-(coal_data_P1.loc[(coal_data_P1['category']==i),'tm']))*((coal_data_P1.loc[(coal_data_P1['category']==i),'vm'])/100))-burning_loss_P1+end_vm_P1_P1+end_moisture_P1)/100).item()) for i in var_coke_P1)),
                                    output(lpSum(var_coke_P1[i]*(((100-(coal_data_P1.loc[(coal_data_P1['category']==i),'tm'])-((100-(coal_data_P1.loc[(coal_data_P1['category']==i),'tm']))*((coal_data_P1.loc[(coal_data_P1['category']==i),'vm'])/100))-burning_loss_P1+end_vm_P1_P1+end_moisture_P1)/100).item()) for i in var_coke_P1))*values_P2.loc[values_P2['Parameter']=='Realization per ton of HM','Value'].item(),
                                    output(lpSum(var_coke_P1[i]*(((100-(coal_data_P1.loc[(coal_data_P1['category']==i),'tm'])-((100-(coal_data_P1.loc[(coal_data_P1['category']==i),'tm']))*((coal_data_P1.loc[(coal_data_P1['category']==i),'vm'])/100))-burning_loss_P1+end_vm_P1_P1+end_moisture_P1)/100).item()) for i in var_coke_P1))*values_P2.loc[values_P2['Parameter']=='Realization per ton of HM','Value'].item()-Total_Cost_of_Usable_Coke_at_Blast_Furnace_req_P1.value()]
    scenario_P1_df['DRI']=[output(lpSum(var_fines_dri_P1)),total_realisation_P1_dri.value(),total_contribution_P1_dri.value()]

    scenario_P1_df['Sinter']=[spl_input_P1.loc[spl_input_P1['Parameter']=='Sinter','Actual'].item().value(),
                                      spl_input_P1.loc[spl_input_P1['Parameter']=='Sinter','Actual'].item().value()*values_P2.loc[values_P2['Parameter']=='Realization per ton of HM','Value'].item(),
                                      spl_input_P1.loc[spl_input_P1['Parameter']=='Sinter','Actual'].item().value()*values_P2.loc[values_P2['Parameter']=='Realization per ton of HM','Value'].item()-lpSum(sinter_obj_P1).value()]
    scenario_P1_df['Blast Furnace(HM)']=scenario_P1_df['Blast Furnace(HM)'].apply(lambda x: '{:.1f}'.format(x))
    scenario_P1_df['Coke Oven']=scenario_P1_df['Coke Oven'].apply(lambda x: '{:.1f}'.format(x))
    scenario_P1_df['Sinter']=scenario_P1_df['Sinter'].apply(lambda x: '{:.1f}'.format(x))
    scenario_P1_df['DRI']=scenario_P1_df['DRI'].apply(lambda x: '{:.1f}'.format(x))
    scenario_P1_df


    # In[82]:


    scenario_P2_dict={'Parameter':[],'Blast Furnace(HM)':[],'Coke Oven':[],'Sinter':[],'DRI':[]}
    scenario_P2_df=pd.DataFrame(data=scenario_P2_dict)
    scenario_P2_df['Parameter']=['Excpected Production','Total Realisation','Total Contribution']
    scenario_P2_df['Blast Furnace(HM)']=[Kg_by_Thm_P2.value(),int_total_realization_P2.value(),int_total_contribution_P2.value()]
    scenario_P2_df['Coke Oven']=[output(lpSum(var_coke_P2[i]*(((100-(coal_data.loc[(coal_data['category']==i),'tm'])-((100-(coal_data.loc[(coal_data['category']==i),'tm']))*((coal_data.loc[(coal_data['category']==i),'vm'])/100))-burning_loss+end_vm+end_moisture)/100).item()) for i in var_coke_P2)),
                                    output(lpSum(var_coke_P2[i]*(((100-(coal_data.loc[(coal_data['category']==i),'tm'])-((100-(coal_data.loc[(coal_data['category']==i),'tm']))*((coal_data.loc[(coal_data['category']==i),'vm'])/100))-burning_loss+end_vm+end_moisture)/100).item()) for i in var_coke_P2))*values_P2.loc[values_P2['Parameter']=='Realization per ton of HM','Value'].item(),
                                    output(lpSum(var_coke_P2[i]*(((100-(coal_data.loc[(coal_data['category']==i),'tm'])-((100-(coal_data.loc[(coal_data['category']==i),'tm']))*((coal_data.loc[(coal_data['category']==i),'vm'])/100))-burning_loss+end_vm+end_moisture)/100).item()) for i in var_coke_P2))*values_P2.loc[values_P2['Parameter']=='Realization per ton of HM','Value'].item()-Total_Cost_of_Usable_Coke_at_Blast_Furnace_req.value()]
    scenario_P2_df['DRI']=[output(lpSum(var_fines_dri_P2)),total_realisation_P2_dri.value(),total_contribution_P2_dri.value()]

    scenario_P2_df['Sinter']=[spl_input_P2.loc[spl_input_P2['Parameter']=='Sinter','Actual'].item().value(),
                                      spl_input_P2.loc[spl_input_P2['Parameter']=='Sinter','Actual'].item().value()*values_P2.loc[values_P2['Parameter']=='Realization per ton of HM','Value'].item(),
                                      spl_input_P2.loc[spl_input_P2['Parameter']=='Sinter','Actual'].item().value()*values_P2.loc[values_P2['Parameter']=='Realization per ton of HM','Value'].item()-lpSum(sinter_obj_P2).value()]
    scenario_P2_df['Blast Furnace(HM)']=scenario_P2_df['Blast Furnace(HM)'].apply(lambda x: '{:.1f}'.format(x))
    scenario_P2_df['Coke Oven']=scenario_P2_df['Coke Oven'].apply(lambda x: '{:.1f}'.format(x))
    scenario_P2_df['Sinter']=scenario_P2_df['Sinter'].apply(lambda x: '{:.1f}'.format(x))
    scenario_P2_df['DRI']=scenario_P2_df['DRI'].apply(lambda x: '{:.1f}'.format(x))
    scenario_P2_df


    # In[83]:


    contribution_dict={'Parameter':[],'P1':[],'P2':[]}
    contribution_df=pd.DataFrame(data=contribution_dict)
    contribution_df['Parameter']=['BF Contribution','DRI Contribution']
    contribution_df['P1']=[int_total_contribution_P1.value(),total_contribution_P1_dri.value()]
    contribution_df['P2']=[int_total_contribution_P2.value(),total_contribution_P2_dri.value()]
    contribution_df['P1'] = contribution_df['P1'].apply(lambda x: '{:.1f}'.format(x))
    contribution_df['P2'] = contribution_df['P2'].apply(lambda x: '{:.1f}'.format(x))
    contribution_df


    # In[84]:


    sinter_output_dict_P1={'Material':[]}
    sinter_output_df_P1=pd.DataFrame(data=sinter_output_dict_P1)
    sinter_output_df_P1['Material']=[i for i in var_P1]
    sinter_output_df_P1['Quantity(In Tonnes)']=[output(var_P1[i]) for i in var_P1]
    sinter_output_df_P1['Cost']=[output(var_P1[i])*materials.loc[materials['Material']==i,'Cost/MT'].item() for i in var_P1]
    sinter_output_df_P1=sinter_output_df_P1.loc[sinter_output_df_P1['Quantity(In Tonnes)']!=0].round(1).reset_index(drop=True)
    #sinter_output_df_P1=sinter_output_df_P1.round(1)
    #sinter_output_df_P1=sinter_output_df_P1.loc[sinter_output_df_P1['Quantity(In Tonnes)']!=0].reset_index(drop=True)
    sinter_output_df_P1


    # In[85]:


    sinter_output_dict_P2={'Material':[]}
    sinter_output_df_P2=pd.DataFrame(data=sinter_output_dict_P2)
    sinter_output_df_P2['Material']=[i for i in var_P2]
    sinter_output_df_P2['Quantity(In Tonnes)']=[output(var_P2[i]) for i in var_P2]
    sinter_output_df_P2['Cost']=[output(var_P2[i])*materials.loc[materials['Material']==i,'Cost/MT'].item() for i in var_P2]
    sinter_output_df_P2=sinter_output_df_P2.loc[sinter_output_df_P2['Quantity(In Tonnes)']!=0].round(1).reset_index(drop=True)
    #sinter_output_df_P2=sinter_output_df_P2.round(1)
    #sinter_output_df_P2=sinter_output_df_P2.loc[sinter_output_df_P2['Quantity(In Tonnes)']!=0].reset_index(drop=True)
    sinter_output_df_P2


    # In[86]:


    bmo_output_dict_P1={'Material':[]}
    bmo_output_df_P1=pd.DataFrame(data=bmo_output_dict_P1)
    bmo_output_df_P1['Material']=[i for i in var_bmo_P1]
    bmo_output_df_P1['Quantity(In Tonnes)']=[output(var_bmo_P1[i]) for i in var_bmo_P1]
    bmo_output_df_P1['Cost']=[output(var_bmo_P1[i])*bmo_materials.loc[bmo_materials['Material']==i,'Cost/MT'].item() for i in var_bmo_P1]
    bmo_output_df_P1=bmo_output_df_P1.loc[bmo_output_df_P1['Quantity(In Tonnes)']!=0].round(1).reset_index(drop=True)
    #bmo_output_df_P1=bmo_output_df_P1.round(1)
    #bmo_output_df_P1=bmo_output_df_P1.loc[bmo_output_df_P1['Quantity(In Tonnes)']!=0].reset_index(drop=True)
    bmo_output_df_P1


    # In[87]:


    bmo_output_dict_P2={'Material':[]}
    bmo_output_df_P2=pd.DataFrame(data=bmo_output_dict_P2)
    bmo_output_df_P2['Material']=[i for i in var_bmo_P2]
    bmo_output_df_P2['Quantity(In Tonnes)']=[output(var_bmo_P2[i]) for i in var_bmo_P2]
    bmo_output_df_P2['Cost']=[output(var_bmo_P2[i])*bmo_materials.loc[bmo_materials['Material']==i,'Cost/MT'].item() for i in var_bmo_P2]
    bmo_output_df_P2=bmo_output_df_P2.loc[bmo_output_df_P2['Quantity(In Tonnes)']!=0].round(1).reset_index(drop=True)
    #bmo_output_df_P2=bmo_output_df_P2.round(1)
    #bmo_output_df_P2=bmo_output_df_P2.loc[bmo_output_df_P2['Quantity(In Tonnes)']!=0].reset_index(drop=True)
    bmo_output_df_P2


    # In[329]:


    #sinter_P1_chemistry_output_df_P2['Actual']=
    fe_sinter_P1=output((lpSum(mines.loc[mines['Plant']=='P1','Fe(new)_fines'])+lpSum(materials['Fe(new)_P1']))*100)
    SiO2_sinter_P1=output((lpSum(mines.loc[mines['Plant']=='P1','SiO2(new)_fines'])+lpSum(materials['SiO2(new)_P1']))*100)
    al2o3_sinter_P1=output((lpSum(mines.loc[mines['Plant']=='P1','Al2O3(new)_fines'])+lpSum(materials['Al2O3(new)_P1']))*100)
    cao_sinter_P1=output((lpSum(mines.loc[mines['Plant']=='P1','CaO(new)_fines'])+lpSum(materials['CaO(new)_P1']))*100)
    mgo_sinter_P1=output((lpSum(mines.loc[mines['Plant']=='P1','MgO(new)_fines'])+lpSum(materials['MgO(new)_P1']))*100)
    basicity_P1=output((lpSum(mines.loc[mines['Plant']=='P1','CaO(new)_fines'])+lpSum(materials['CaO(new)_P1']))*100)
    feeding_rate_P1=output((lpSum(P1_fines_variables_list)+lpSum(var_P1)-var_P1['RF Internal'])/(no_of_days_in_month*24))
    sinter_P1_chemistry_dict_P1={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','Basicity','Residue Production','Feeding Rate'],
                                       'Actual':[fe_sinter_P1,SiO2_sinter_P1,al2o3_sinter_P1,cao_sinter_P1,mgo_sinter_P1,basicity_P1,(lpSum(mines.loc[mines['Plant']=='P1','Residue_fines'])+lpSum(materials['Residue_P1'])).value(),feeding_rate_P1]}
    sinter_P1_chemistry_output_df_P1=pd.DataFrame(data=sinter_P1_chemistry_dict_P1).round(1)
    sinter_P1_chemistry_output_df_P1


    # In[331]:


    fe_sinter_P2=output((lpSum(mines.loc[mines['Plant']=='P2','Fe(new)_fines'])+lpSum(materials['Fe(new)_P2']))*100)
    SiO2_sinter_P2=output((lpSum(mines.loc[mines['Plant']=='P2','SiO2(new)_fines'])+lpSum(materials['SiO2(new)_P2']))*100)
    al2o3_sinter_P2=output((lpSum(mines.loc[mines['Plant']=='P2','Al2O3(new)_fines'])+lpSum(materials['Al2O3(new)_P2']))*100)
    cao_sinter_P2=output((lpSum(mines.loc[mines['Plant']=='P2','CaO(new)_fines'])+lpSum(materials['CaO(new)_P2']))*100)
    mgo_sinter_P2=output((lpSum(mines.loc[mines['Plant']=='P2','MgO(new)_fines'])+lpSum(materials['MgO(new)_P2']))*100)
    basicity_P2=output((lpSum(mines.loc[mines['Plant']=='P2','CaO(new)_fines'])+lpSum(materials['CaO(new)_P2']))*100)
    feeding_rate_P2=output((lpSum(P2_fines_variables_list)+lpSum(var_P2)-var_P2['RF Internal'])/(no_of_days_in_month*24))
    sinter_P2_chemistry_dict_P2={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','Basicity','Residue Production','Feeding Rate'],
                                           'Actual':[fe_sinter_P2,SiO2_sinter_P2,al2o3_sinter_P2,cao_sinter_P2,mgo_sinter_P2,basicity_P2,(lpSum(mines.loc[mines['Plant']=='P2','Residue_fines'])+lpSum(materials['Residue_P2'])).value(),feeding_rate_P2]}
    sinter_P2_chemistry_output_df_P2=pd.DataFrame(data=sinter_P2_chemistry_dict_P2).round(1)
    sinter_P2_chemistry_output_df_P2


    spl_output_dict_P1={'Quantinty(In Tonnes)':[sinter_P1,pellet_P1,lump_P1]}
    spl_output_dict_P2={'Quantinty(In Tonnes)':[sinter_P2,pellet_P2,lump_P2]}
    spl_output_df_P1=pd.DataFrame(data=spl_output_dict_P1,index=['Sinter','Pellet','Lump']).round(1)
    spl_output_df_P2=pd.DataFrame(data=spl_output_dict_P2,index=['Sinter','Pellet','Lump']).round(1)
    spl_output_df_P1['Parameter']=['Sinter','Pellet','Lump']
    spl_output_df_P2['Parameter']=['Sinter','Pellet','Lump']
    #production_output
    #spl_output_df['Max(wtd) P1']=[sinter_P1_max_wtd,pellet_P1_max_wtd,lump_P1_max_wtd,0,0,0]
    #spl_output_df['Max(wtd) P2']=[sinter_P2_max_wtd,pellet_P2_max_wtd,lump_P2_max_wtd,0,0,0]
    spl_output_df_P1['% Mix']=spl_output_df_P1['Quantinty(In Tonnes)']/(sinter_P1+pellet_P1+lump_P1)*100
    spl_output_df_P2['% Mix']=spl_output_df_P2['Quantinty(In Tonnes)']/(sinter_P2+pellet_P2+lump_P2)*100
    spl_output_df_P1=spl_output_df_P1.round(1)
    spl_output_df_P2=spl_output_df_P2.round(1)
    spl_output_df_P1

    production_output_P1_dict={'Value':[hot_metal_prodction_P1,charge_capacity_P1_op]}#,dri_pellet_P1_model]}
    production_output_P2_dict={'Value':[hot_metal_prodction_P2,charge_capacity_P2_op]}#,dri_pellet_P2_model]}  
    production_output_df_P1=pd.DataFrame(data=production_output_P1_dict,index=['Hot Metal Production(In Tonnes)','Charge Capacity']).round(1)#,'Pellet DRI Model']).round(1)
    production_output_df_P2=pd.DataFrame(data=production_output_P2_dict,index=['Hot Metal Production(In Tonnes)','Charge Capacity']).round(1)#,'Pellet DRI Model']).round(1)    
    production_output_df_P2                                  


    # In[91]:


    dri_output_dict={'Parameter':[],'P1':[],'P2':[]}
    dri_output_df=pd.DataFrame(data=dri_output_dict)
    dri_output_df['Parameter']=['Pellet DRI Model','Pellet Consumption']
    dri_output_df.loc[dri_output_df['Parameter']=='Pellet DRI Model','P1']=dri_pellet_P1_model
    dri_output_df.loc[dri_output_df['Parameter']=='Pellet DRI Model','P2']=dri_pellet_P2_model
    dri_output_df.loc[dri_output_df['Parameter']=='Pellet Consumption','P1']=output(lpSum(var_fines_dri_P1))
    dri_output_df.loc[dri_output_df['Parameter']=='Pellet Consumption','P2']=output(lpSum(var_fines_dri_P2))
    dri_output_df=dri_output_df.round(1)
    dri_output_df
    #output(lpSum(var_fines_dri_P1))


    # In[92]:


    bmo_chemistry_dict_P1={'Value':[]}
    bmo_chemistry_output_df_P1=pd.DataFrame(data=bmo_chemistry_dict_P1)
    #bmo_chemistry_output_df_P1.index.name='Parameters'
    for i in bmo_chemistry_table_P1['Parameters']:
        #bmo_chemistry_output_df_P1.loc[i,'Parameters']=i
        #bmo_chemistry_output_df_P1.loc[i,'Actual']=output(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']==i,'Actual Values_P1'])
        #bmo_chemistry_output_df_P1.loc[i,'Min(wtd)']=output(bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1['Parameters']==i,'min(total)_P1'])
        #bmo_chemistry_output_df_P1.loc[i,'Max(wtd)']=output(bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1['Parameters']==i,'max(total)_P1'])
        bmo_chemistry_output_df_P1.loc[i,'Min']=bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']==i,'Min'].item()
        bmo_chemistry_output_df_P1.loc[i,'Max']=bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']==i,'Max'].item()
    bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index.values=='FE (Kg/THM)','Value']=output(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Actual Values_P1'])/output(Kg_by_Thm_P1)
    bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index.values=='Slag Amount','Value']=output(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Slag Amount','Actual Values_P1'])/output(Kg_by_Thm_P1)
    bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index.values=='Sulphur in HM','Value']=output(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Sulphur in HM','Actual Values_P1'])/output(Kg_by_Thm_P1)
    bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index.values=='Phosphorus','Value']=output(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Phosphorus','Actual Values_P1'])/output(Kg_by_Thm_P1)
    bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index.values=='MgO','Value']=output(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='MgO','Actual Values_P1'])/output((CaO_total_P1+MgO_total_P1+(SiO2_total_P1-(Si_total_P1/0.466))+Al2O3_total_P1)/(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Min'].item()/1000))
    bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index.values=='Al2O3','Value']=output(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='Al2O3','Actual Values_P1'].item())/output((CaO_total_P1+MgO_total_P1+(SiO2_total_P1-(Si_total_P1/0.466))+Al2O3_total_P1)/(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Min'].item()/1000))
    bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index.values=='CaO','Value']=output(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='CaO','Actual Values_P1'])/output(CaO_total_P1+MgO_total_P1+(SiO2_total_P1-(Si_total_P1/0.466))+Al2O3_total_P1)/(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Min'].item()/1000)
    bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index.values=='SiO2','Value']=output(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='SiO2','Actual Values_P1'])/output(CaO_total_P1+MgO_total_P1+(SiO2_total_P1-(Si_total_P1/0.466))+Al2O3_total_P1)/(bmo_chemistry_table_P1.loc[bmo_chemistry_table_P1['Parameters']=='FE (Kg/THM)','Min'].item()/1000)
    bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index.values=='Slag Basicity (B1)','Value']=bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index.values=='CaO','Value'].item()/bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index.values=='SiO2','Value'].item()
    for i in bmo_chemistry_output_df_P1.index[[0,1,2,4,5,6,7,8]]:
        bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index==i,'Value']=bmo_chemistry_output_df_P1.loc[bmo_chemistry_output_df_P1.index==i,'Value'].apply(lambda x: '{:.2f}'.format(x))
    bmo_chemistry_output_df_P1
        # In[93]:


    bmo_chemistry_dict_P2={'Value':[]}
    bmo_chemistry_output_df_P2=pd.DataFrame(data=bmo_chemistry_dict_P2)
    #bmo_chemistry_output_df_P2.index.name='Parameters'
    for i in bmo_chemistry_table_P2['Parameters']:
        #bmo_chemistry_output_df_P2.loc[i,'Parameters']=i
        #bmo_chemistry_output_df_P2.loc[i,'Actual']=output(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']==i,'Actual Values_P2'])
        #bmo_chemistry_output_df_P2.loc[i,'Min(wtd)']=output(bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2['Parameters']==i,'min(total)_P2'])
        #bmo_chemistry_output_df_P2.loc[i,'Max(wtd)']=output(bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2['Parameters']==i,'max(total)_P2'])
        bmo_chemistry_output_df_P2.loc[i,'Min']=bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']==i,'Min'].item()
        bmo_chemistry_output_df_P2.loc[i,'Max']=bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']==i,'Max'].item()

    bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index.values=='FE (Kg/THM)','Value']=output(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Actual Values_P2'])/output(Kg_by_Thm_P2)
    bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index.values=='Slag Amount','Value']=output(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Slag Amount','Actual Values_P2'])/output(Kg_by_Thm_P2)
    bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index.values=='Sulphur in HM','Value']=output(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Sulphur in HM','Actual Values_P2'])/output(Kg_by_Thm_P2)
    bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index.values=='Phosphorus','Value']=output(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Phosphorus','Actual Values_P2'])/output(Kg_by_Thm_P2)
    bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index.values=='MgO','Value']=output(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='MgO','Actual Values_P2'])/output((CaO_total_P2+MgO_total_P2+(SiO2_total_P2-(Si_total_P2/0.466))+Al2O3_total_P2)/(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Min'].item()/1000))
    bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index.values=='Al2O3','Value']=output(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='Al2O3','Actual Values_P2'].item())/output((CaO_total_P2+MgO_total_P2+(SiO2_total_P2-(Si_total_P2/0.466))+Al2O3_total_P2)/(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Min'].item()/1000))
    bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index.values=='CaO','Value']=output(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='CaO','Actual Values_P2'])/output(CaO_total_P2+MgO_total_P2+(SiO2_total_P2-(Si_total_P2/0.466))+Al2O3_total_P2)/(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Min'].item()/1000)
    bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index.values=='SiO2','Value']=output(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='SiO2','Actual Values_P2'])/output(CaO_total_P2+MgO_total_P2+(SiO2_total_P2-(Si_total_P2/0.466))+Al2O3_total_P2)/(bmo_chemistry_table_P2.loc[bmo_chemistry_table_P2['Parameters']=='FE (Kg/THM)','Min'].item()/1000)
    bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index.values=='Slag Basicity (B1)','Value']=bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index.values=='CaO','Value'].item()/bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index.values=='SiO2','Value'].item()
    for i in bmo_chemistry_output_df_P2.index[[0,1,2,4,5,6,7,8]]:
        bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index==i,'Value']=bmo_chemistry_output_df_P2.loc[bmo_chemistry_output_df_P2.index==i,'Value'].apply(lambda x: '{:.2f}'.format(x))
    bmo_chemistry_output_df_P2
    # In[94]:


    dri_pellet_P1_model=output(lpSum(var_fines_dri_P1)/Pellet_Sp_Cons_Ton_P1_model)
    dri_pellet_P2_model=output(lpSum(var_fines_dri_P2)/Pellet_Sp_Cons_Ton_P2_model)


    # In[97]:


    pellet_market=list()
    for v, coefficient in market_sale.items():
        pellet_market.append(coefficient*v.varValue)
    pellet_market_sale=lpSum(pellet_market).value()+pellet_production
    pellet_contribution_list=list()
    for v, coefficient in pellet_contribution.items():
        pellet_contribution_list.append(coefficient*v.varValue)
    pellet_contribution_op=(lpSum(pellet_contribution_list).value()+ pellet_production*values_P1.loc[values_P1['Parameter']=='Realization from pellet Sale','Value'].item())

    total_pellet_list=list()
    for v, coefficient in total_pellet.items():
        total_pellet_list.append(coefficient*v.varValue)
    total_pellet_op=(lpSum(total_pellet_list).value()+ pellet_production)

    pellet_used_in_bf_op=output(pellet_used_in_bf)
    barbil_cost=output(cost_objective_barbil_list)
    pellet_dict={'Value':[pellet_production,pellet_used_in_bf_op,output(DRI),pellet_market_sale,pellet_contribution_op,barbil_cost],'Limit Lower':['NA',0,0,0,'NA','NA'],'Limit Upper':[pellet_production,'NA','NA','NA','NA','NA']}
    pellet_df=pd.DataFrame(data=pellet_dict,index=['Total Pellet Demand','Used in BF','DRI','Market Sale','Pellet Realisation','Barbil Cost']).round(1)
    pellet_df


    # In[75]:


    coke_op_dict_P1={'Actual':[]}
    coke_op_output_df_P1=pd.DataFrame(data=coke_op_dict_P1)
    for i in coke_input_P1['Parameter']:
        if coke_input_P1.loc[coke_input_P1['Parameter']==i,'Parameter'].item()!='Coke (C1)':
            coke_op_output_df_P1.loc[i,'Actual']=output(var_bmo_P1[i])
            coke_op_output_df_P1.loc[i,'Limit(lower)']=output(Kg_by_Thm_P1*coke_input_P1.loc[coke_input_P1['Parameter']==i,'Limit'].item()/1000)
            coke_op_output_df_P1.loc[i,'Value']=coke_input_P1.loc[coke_input_P1['Parameter']==i,'Limit'].item()
    coke_op_output_df_P1.loc['Coke(C1)','Actual']=output(lpSum(var_coke_P1[i]*(((100-(coal_data_P1.loc[(coal_data_P1['category']==i),'tm'])-((100-(coal_data_P1.loc[(coal_data_P1['category']==i),'tm']))*((coal_data_P1.loc[(coal_data_P1['category']==i),'vm'])/100))-burning_loss_P1+end_vm_P1_P1+end_moisture_P1)/100).item()) for i in var_coke_P1))    
    coke_op_output_df_P1.loc['Coke(C1)','Limit(lower)']=output(Kg_by_Thm_P1*coke_input_P1.loc[coke_input_P1['Parameter']=='Coke (C1)','Limit'].item()/1000)
    coke_op_output_df_P1.loc['Coke(C1)','Value']=coke_input_P1.loc[coke_input_P1['Parameter']=='Coke (C1)','Limit'].item()
    coke_op_output_df_P1.loc[:,'Coke Rate']=coke_op_output_df_P1.loc[:,'Actual']*1000/(Kg_by_Thm_P1.value())
    coke_op_output_df_P1=coke_op_output_df_P1.round(1)
    coke_op_output_df_P1
        # In[99]:


    coke_op_dict_P2={'Actual':[]}
    coke_op_output_df_P2=pd.DataFrame(data=coke_op_dict_P2)
    for i in coke_input_P2['Parameter']:
        if coke_input_P2.loc[coke_input_P2['Parameter']==i,'Parameter'].item()!='Coke (C1)':
            coke_op_output_df_P2.loc[i,'Actual']=output(coke_input_P2.loc[coke_input_P2['Parameter']==i,'Actual'])
            coke_op_output_df_P2.loc[i,'Limit(lower)']=output(Kg_by_Thm_P2*coke_input_P2.loc[coke_input_P2['Parameter']==i,'Limit'].item()/1000)
            coke_op_output_df_P2.loc[i,'Value']=coke_input_P2.loc[coke_input_P2['Parameter']==i,'Limit'].item()
    coke_op_output_df_P2.loc['Coke(C1)','Actual']=output(lpSum(var_coke_P2[i]*(((100-(coal_data.loc[(coal_data['category']==i),'tm'])-((100-(coal_data.loc[(coal_data['category']==i),'tm']))*((coal_data.loc[(coal_data['category']==i),'vm'])/100))-burning_loss+end_vm+end_moisture)/100).item()) for i in var_coke_P2))    
    coke_op_output_df_P2.loc['Coke(C1)','Limit(lower)']=output(Kg_by_Thm_P2*coke_input_P2.loc[coke_input_P2['Parameter']=='Coke (C1)','Limit'].item()/1000)
    coke_op_output_df_P2.loc['Coke(C1)','Value']=coke_input_P2.loc[coke_input_P2['Parameter']=='Coke (C1)','Limit'].item()
    coke_op_output_df_P2.loc[:,'Coke Rate']=coke_op_output_df_P2.loc[:,'Actual']*1000/(Kg_by_Thm_P2.value())       
    coke_op_output_df_P2=coke_op_output_df_P2.round(1)
    coke_op_output_df_P2
        # In[100]:


    coal_output_P1_dict={'Coal':[i for i in var_coke_P1],
                            'Quantity in Tonnes':[output(var_coke_P1[i]) for i in var_coke_P1 ],
                            'Mix%':[output(var_coke_P1[i])/output(var_coke_P1)*100 for i in var_coke_P1]}
    coal_output_P1_df=pd.DataFrame(data=coal_output_P1_dict)
    coal_output_P1_df['Cost']=[output(var_coke_P1[i]*coals_P1.loc[coals_P1['coals_P1']==i,'input_price'].item()) for i in var_coke_P1]
    coal_output_P1_df['Category']=[coals_P1.loc[coals_P1['coals_P1']==i,'type'].item() for i in coal_output_P1_df['Coal']]
    coal_output_P1_df=coal_output_P1_df.loc[coal_output_P1_df['Quantity in Tonnes']!=0].round(1).reset_index(drop=True)
    pd.set_option('display.float_format', lambda x: '%.1f' % x)
    coal_output_P1_df


    # In[101]:


    coal_output_P2_dict={'Coal':[i for i in var_coke_P2],
                            'Quantity in Tonnes':[output(var_coke_P2[i]) for i in var_coke_P2 ],
                            'Mix%':[output(var_coke_P2[i])/output(var_coke_P2)*100 for i in var_coke_P2]}
    coal_output_P2_df=pd.DataFrame(data=coal_output_P2_dict).round(1)
    coal_output_P2_df['Category']=[coals.loc[coals['coals']==i,'type'].item() for i in coal_output_P2_df['Coal']]
    coal_output_P2_df['Cost']=[output(var_coke_P2[i]*coals.loc[coals['coals']==i,'input_price'].item()) for i in var_coke_P2]
    coal_output_P2_df=coal_output_P2_df.loc[coal_output_P2_df['Quantity in Tonnes']!=0].reset_index(drop=True)
    coal_output_P2_df


    # In[102]:


    #csr calculation
    csr_constraint=68.5*lpSum(var_coke_P1[i]for i in var_coke_P1)+0.512*mbi_P1+0.02308*lpSum((log10(coal_data_P1.loc[(coal_data_P1['category']==i),'Fluidity'])*100)*var_coke_P1[i] for i in var_coke_P1)-1.775*lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'mmr'])*var_coke_P1[i] for i in var_coke_P1)-0.075*lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'vitrinite'])*var_coke_P1[i] for i in var_coke_P1)
    csr_P1=output(csr_constraint)/output(var_coke_P1)
    coke_ash_P1=output((lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'Ash'])*var_coke_P1[i] for i in var_coke_P1))*inc_fact_coke_ash_P1)/output(var_coke_P1)
    coke_sulphur_P1=output((lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),'sulphur'])*var_coke_P1[i] for i in var_coke_P1))*sulphur_retention_P1)/output(var_coke_P1)
    coke_parameters_P1_dict={'Parameter':['CSR','Coke Ash','Coke Sulphur'],
                                'Value':[csr_P1,coke_ash_P1,coke_sulphur_P1]}
    coke_parameters_P1_df=pd.DataFrame(data=coke_parameters_P1_dict).round(1)
    coke_parameters_P1_df


    # In[103]:


    #csr calculation
    csr_constraint=68.5*lpSum(var_coke_P2[i]for i in var_coke_P2)+0.512*mbi+0.02308*lpSum((log10(coal_data.loc[(coal_data['category']==i),'Fluidity'])*100)*var_coke_P2[i] for i in var_coke_P2)-1.775*lpSum((coal_data.loc[(coal_data['category']==i),'mmr'])*var_coke_P2[i] for i in var_coke_P2)-0.075*lpSum((coal_data.loc[(coal_data['category']==i),'vitrinite'])*var_coke_P2[i] for i in var_coke_P2)
    csr_P2=output(csr_constraint)/output(var_coke_P2)
    coke_ash_P2=output((lpSum((coal_data.loc[(coal_data['category']==i),'Ash'])*var_coke_P2[i] for i in var_coke_P2))*inc_fact_coke_ash)/output(var_coke_P2)
    coke_sulphur_P2=output((lpSum((coal_data.loc[(coal_data['category']==i),'sulphur'])*var_coke_P2[i] for i in var_coke_P2))*sulphur_retention)/output(var_coke_P2)
    coke_parameters_P2_dict={'Parameter':['CSR','Coke Ash','Coke Sulphur'],
                                'Value':[csr_P2,coke_ash_P2,coke_sulphur_P2]}
    coke_parameters_P2_df=pd.DataFrame(data=coke_parameters_P2_dict).round(1)
    coke_parameters_P2_df


    # In[104]:


    #parameter table
    def pm_output_calculation_P2(j):
        op_list=list()
        for v, coefficient in (lpSum((coal_data.loc[(coal_data['category']==i),j])*var_coke_P2[i] for i in var_coke_P2)/lpSum(coal_output_P2_df.loc[:,'Quantity in Tonnes']).value()).items():
            op_list.append(coefficient*v.varValue)
        op_total=lpSum(op_list).value()
        return op_total
    vm_output_P2=pm_output_calculation_P2('vm')
    csn_output_P2=pm_output_calculation_P2('csn')
    mmr_output_P2=pm_output_calculation_P2('mmr')
    vitrinite_output_P2=pm_output_calculation_P2('vitrinite')
    v9_v14_output_P2=pm_output_calculation_P2('v9_v14')
    pm_output_calculation_P2('sulphur')
    ash_output_P2=(lpSum((coal_data.loc[(coal_data['category']==key),'Ash'])*value.varValue for key, value in var_coke_P2.items())/lpSum(value.varValue for key,value in var_coke_P2.items())).value()

    tm_op_list_P2=list()
    for v, coefficient in lpSum(coal_data.loc[coal_data['category']==i,'tm']*var_coke_P2[i] for i in var_coke_P2).items():
                tm_op_list_P2.append(coefficient*v.varValue)
    tm_output_P2=lpSum(tm_op_list_P2).value()/lpSum(coal_output_P2_df.loc[:,'Quantity in Tonnes']).value()
    tm_output_P2


    log_fluidity_op_list_P2=list()
    for v, coefficient in lpSum((log10(coal_data.loc[(coal_data['category']==i),'Fluidity'])*100)*var_coke_P2[i] for i in var_coke_P2).items():
                log_fluidity_op_list_P2.append(coefficient*v.varValue)
    log_fluidity_output_P2=lpSum(log_fluidity_op_list_P2).value()/lpSum(coal_output_P2_df.loc[:,'Quantity in Tonnes']).value()
    log_fluidity_output_P2


    fc_op_list_P2=list()
    for v, coefficient in lpSum(coal_data.loc[coal_data['category']==i,'FC ']*var_coke_P2[i] for i in var_coke_P2).items():
                fc_op_list_P2.append(coefficient*v.varValue)
    fc_output_P2=lpSum(fc_op_list_P2).value()/lpSum(coal_output_P2_df.loc[:,'Quantity in Tonnes']).value()
    fc_output_P2


    blended_coal_parameters_dict_P2=[['VM',vm_output_P2],['CSN',csn_output_P2],['MMR',mmr_output_P2],['Vitrinite',vitrinite_output_P2],
                            ['V9-V14',v9_v14_output_P2],['Sulphur',pm_output_calculation_P2('sulphur')],
                            ['Ash',ash_output_P2],['TM',tm_output_P2],['FC',fc_output_P2],
                            ['Log Fluidity',log_fluidity_output_P2]]
    blended_coal_parameters_df_P2=pd.DataFrame.from_records(blended_coal_parameters_dict_P2,columns=['Parameter','Value']).round(1)
    blended_coal_parameters_df_P2


    # In[105]:


    #parameter table
    def pm_output_calculation_P1(j):
        op_list=list()
        for v, coefficient in (lpSum((coal_data_P1.loc[(coal_data_P1['category']==i),j])*var_coke_P1[i] for i in var_coke_P1)/lpSum(coal_output_P1_df.loc[:,'Quantity in Tonnes']).value()).items():
            op_list.append(coefficient*v.varValue)
        op_total=lpSum(op_list).value()
        return op_total
    vm_output_P1=pm_output_calculation_P1('vm')
    csn_output_P1=pm_output_calculation_P1('csn')
    mmr_output_P1=pm_output_calculation_P1('mmr')
    vitrinite_output_P1=pm_output_calculation_P1('vitrinite')
    v9_v14_output_P1=pm_output_calculation_P1('v9_v14')
    pm_output_calculation_P1('sulphur')
    ash_output_P1=(lpSum((coal_data_P1.loc[(coal_data_P1['category']==key),'Ash'])*value.varValue for key, value in var_coke_P1.items())/lpSum(value.varValue for key,value in var_coke_P1.items())).value()

    tm_op_list_P1=list()
    for v, coefficient in lpSum(coal_data_P1.loc[coal_data_P1['category']==i,'tm']*var_coke_P1[i] for i in var_coke_P1).items():
                tm_op_list_P1.append(coefficient*v.varValue)
    tm_output_P1=lpSum(tm_op_list_P1).value()/lpSum(coal_output_P1_df.loc[:,'Quantity in Tonnes']).value()
    tm_output_P1


    log_fluidity_op_list_P1=list()
    for v, coefficient in lpSum((log10(coal_data_P1.loc[(coal_data_P1['category']==i),'Fluidity'])*100)*var_coke_P1[i] for i in var_coke_P1).items():
                log_fluidity_op_list_P1.append(coefficient*v.varValue)
    log_fluidity_output_P1=lpSum(log_fluidity_op_list_P1).value()/lpSum(coal_output_P1_df.loc[:,'Quantity in Tonnes']).value()
    log_fluidity_output_P1


    fc_op_list_P1=list()
    for v, coefficient in lpSum(coal_data_P1.loc[coal_data_P1['category']==i,'FC ']*var_coke_P1[i] for i in var_coke_P1).items():
                fc_op_list_P1.append(coefficient*v.varValue)
    fc_output_P1=lpSum(fc_op_list_P1).value()/lpSum(coal_output_P1_df.loc[:,'Quantity in Tonnes']).value()
    fc_output_P1


    blended_coal_parameters_dict_P1=[['VM',vm_output_P1],['CSN',csn_output_P1],['MMR',mmr_output_P1],['Vitrinite',vitrinite_output_P1],
                            ['V9-V14',v9_v14_output_P1],['Sulphur',pm_output_calculation_P1('sulphur')],
                            ['Ash',ash_output_P1],['TM',tm_output_P1],['FC',fc_output_P1],
                            ['Log Fluidity',log_fluidity_output_P1]]
    blended_coal_parameters_df_P1=pd.DataFrame.from_records(blended_coal_parameters_dict_P1,columns=['Parameter','Value']).round(1)
    blended_coal_parameters_df_P1


    # In[106]:


    #iron mines
    supply_mines_output_dict={'Mines':[]}
    supply_mines_output_df=pd.DataFrame(data=supply_mines_output_dict)
    supply_mines_output_df['Mines']=[j for j in supply_constraint_table['Mines'] ]
    for j in supply_constraint_table['Mines']:
        a=list()
        b=list()
        for i in var_fines:
            if mines.loc[mines['possibilities']==i,'Mine'].item()==j:
                a.append(var_fines[i])
        for i in var_fines_bf_P1:
            if mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Mine'].item()==j:
                a.append(var_fines_bf_P1[i])
        for i in var_fines_bf_P2:
            if mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Mine'].item()==j:
                a.append(var_fines_bf_P2[i])
        for i in var_fines_dri_P1:
            if mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Mine'].item()==j:
                a.append(var_fines_dri_P1[i])
        for i in var_fines_dri_P2:
            if mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Mine'].item()==j:
                a.append(var_fines_dri_P2[i])
        for i in var_fines_pellet_market:
            if mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Mine'].item()==j:
                a.append(var_fines_pellet_market[i])
        supply_mines_output_df.loc[supply_mines_output_df['Mines']==j,'Actual Fines']=output(a)
    supply_mines_output_df['Fines(Max)']=supply_constraint_table['Fines']
    for j in supply_constraint_table['Mines']:
        a=list()
        b=list()
        for i in var_lumps:
            if mines_lumps.loc[mines_lumps['possibilities']==i,'Mine'].item()==j:
                a.append(var_lumps[i])
        supply_mines_output_df.loc[supply_mines_output_df['Mines']==j,'Actual Lumps']=output(a)
    supply_mines_output_df['Lumps(Max)']=supply_constraint_table['Lumps']
    #supply_mines_output_df=supply_mines_output_df.loc[supply_mines_output_df['Actual Fines']!=0 and supply_mines_output_df['Actual Lumps']!=0  ]# and supply_mines_output_df.loc[supply_mines_output_df['Actual Lumps']!=0 ] #or supply_mines_output_df['Actual Lumps'].item()!=0].round(1) 
    supply_mines_output_df

    supply_mines_output_df_fines=supply_mines_output_df.iloc[:,:3].round(1)
    supply_mines_output_df_fines=supply_mines_output_df_fines[supply_mines_output_df_fines['Actual Fines']!=0].reset_index(drop=True)
    supply_mines_output_df_fines=supply_mines_output_df_fines.dropna()
    supply_mines_output_df_lumps=supply_mines_output_df.iloc[:,[0,3,4]]
    supply_mines_output_df_lumps=supply_mines_output_df_lumps[supply_mines_output_df_lumps['Actual Lumps']!=0].reset_index(drop=True)
    supply_mines_output_df_lumps=supply_mines_output_df_lumps.dropna()
    supply_mines_output_df_fines

    #Siding Constraints
    siding_output_dict={'Siding':[]}
    siding_output_df=pd.DataFrame(data=siding_output_dict)
    siding_output_df['Siding']=[j for j in siding_constraint_table['Siding'] ]
    for j in siding_constraint_table['Siding']:
        a=list()
        b=list()
        for i in var_fines:
            if mines.loc[mines['possibilities']==i,'Siding'].item()==j:
                a.append(var_fines[i])
        for i in var_fines_bf_P1:
            if mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Siding'].item()==j:
                a.append(var_fines_bf_P1[i])
        for i in var_fines_bf_P2:
            if mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Siding'].item()==j:
                a.append(var_fines_bf_P2[i])
        for i in var_fines_dri_P1:
            if mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Siding'].item()==j:
                a.append(var_fines_dri_P1[i])
        for i in var_fines_dri_P2:
            if mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Siding'].item()==j:
                a.append(var_fines_dri_P2[i])
        for i in var_fines_pellet_market:
            if mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Siding'].item()==j:
                a.append(var_fines_pellet_market[i])
        for i in var_lumps:
            if mines_lumps.loc[mines_lumps['possibilities']==i,'Siding'].item()==j:
                a.append(var_lumps[i])
        siding_output_df.loc[siding_output_df['Siding']==j,'Actual']=output(a)

    siding_output_df['Max']=siding_constraint_table['Limit']
    siding_output_df=siding_output_df[siding_output_df['Actual']!=0].round(1)
    siding_output_df    


    logistics_output_dict={'Plant':[]}
    logistics_output_df=pd.DataFrame(data=logistics_output_dict)
    logistics_output_df['Plant']=[j for j in logistics_constraint_table.loc[:,'Plant']]
    logistics_output_df['Type']='Road' 
    #P1 & Raiagrh
    #raod
    for j in logistics_constraint_table.loc[:1,'Plant']:
        a=list()
        b=list()
        for i in var_fines:
            if mines.loc[mines['possibilities']==i,'Plant'].item()==j and (mines.loc[mines['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_fines[i])
        for i in var_lumps:
            if (mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()==j) and (mines_lumps.loc[mines_lumps['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_lumps[i])
        logistics_output_df.loc[logistics_output_df['Plant']==j,'Actual(Road)']=output(a)
    #Barbil
    #raod
    for j in logistics_constraint_table.loc[2:,'Plant']:
        a=list()
        b=list()
        for i in var_fines_bf_P1:
            if mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Plant'].item()==j and (mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_fines_bf_P1[i])
        for i in var_fines_bf_P2:
            if mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Plant'].item()==j and (mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_fines_bf_P2[i])
        for i in var_fines_dri_P1:
            if mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Plant'].item()==j and (mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_fines_dri_P1[i])
        for i in var_fines_dri_P2:
            if mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Plant'].item()==j and (mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_fines_dri_P2[i])
        for i in var_fines_pellet_market:
            if mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Plant'].item()==j and (mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_fines_pellet_market[i]) 
        for i in var_lumps:
            if (mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()==j) and (mines_lumps.loc[mines_lumps['possibilities']==i,'Siding'].item()=='Road'):
                a.append(var_lumps[i])
        logistics_output_df.loc[logistics_output_df['Plant']=='Barbil','Actual(Road)']=output(a)    

    logistics_output_df['Max']=logistics_constraint_table['Road']

    #rail
    logistics_output_df['Type ']='Rail'
    for j in logistics_constraint_table.loc[:1,'Plant']:
        a=list()
        b=list()
        for i in var_fines:
            if mines.loc[mines['possibilities']==i,'Plant'].item()==j and (mines.loc[mines['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_fines[i])
        for i in var_lumps:
            if (mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()==j) and (mines_lumps.loc[mines_lumps['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_lumps[i])
        logistics_output_df.loc[logistics_output_df['Plant']==j,'Actual(Rail)']=output(a)    
    #Rail Barbil
    for j in logistics_constraint_table.loc[2:,'Plant']:
        a=list()
        b=list()
        for i in var_fines_bf_P1:
            if mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Plant'].item()==j and (mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_fines_bf_P1[i])
        for i in var_fines_bf_P2:
            if mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Plant'].item()==j and (mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_fines_bf_P2[i])
        for i in var_fines_dri_P1:
            if mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Plant'].item()==j and (mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_fines_dri_P1[i])
        for i in var_fines_dri_P2:
            if mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Plant'].item()==j and (mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_fines_dri_P2[i])
        for i in var_fines_pellet_market:
            if mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Plant'].item()==j and (mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_fines_pellet_market[i]) 
        for i in var_lumps:
            if (mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item()==j) and (mines_lumps.loc[mines_lumps['possibilities']==i,'Siding'].item()!='Road'):
                a.append(var_lumps[i])
        logistics_output_df.loc[logistics_output_df['Plant']=='Barbil','Actual(Rail)']=[output(a)]
    logistics_output_df['Max ']=logistics_constraint_table ['Rail']
    logistics_output_df=logistics_output_df.round(1)
    logistics_output_df


    #ltc
    ltc_output_dict={'Mine Owners':[]}
    ltc_output_df=pd.DataFrame(data=ltc_output_dict)
    ltc_output_df['Mine Owners']=[j for j in ltc_constraint_table['Mine Owners'] ]
    for j in ltc_constraint_table['Mine Owners']:
        a=list()
        b=list()
        for i in var_fines:
            if mines.loc[mines['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_fines[i])
        for i in var_fines_bf_P1:
            if mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_fines_bf_P1[i])
        for i in var_fines_bf_P2:
            if mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_fines_bf_P2[i])
        for i in var_fines_dri_P1:
            if mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_fines_dri_P1[i])
        for i in var_fines_dri_P2:
            if mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_fines_dri_P2[i])
        for i in var_fines_pellet_market:
            if mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_fines_pellet_market[i])
        for i in var_lumps:
            if mines_lumps.loc[mines_lumps['possibilities']==i,'Mine Owners'].item()==j:
                a.append(var_lumps[i])
        ltc_output_df.loc[ltc_output_df['Mine Owners']==j,'Actual']=output(a)
    ltc_output_df=ltc_output_df.round(1)


    # In[108]:


    fines_mines_dict={'Possibility':[]}
    fines_mines_df=pd.DataFrame(data=fines_mines_dict)
    fines_mines_df['Possibility']=[i for i in var_fines]
    fines_mines_df['Mine']=[mines.loc[mines['possibilities']==i,'Mine'].item() for i in fines_mines_df['Possibility']]
    fines_mines_df['Mine Owner']=[mines.loc[mines['possibilities']==i,'Mine Owners'].item() for i in fines_mines_df['Possibility']]
    fines_mines_df['Plant']=[mines.loc[mines['possibilities']==i,'Plant'].item() for i in fines_mines_df['Possibility']]
    fines_mines_df['Siding']=[mines.loc[mines['possibilities']==i,'Siding'].item() for i in fines_mines_df['Possibility']]
    fines_mines_df['Quantity(Tonnes)']=[output(var_fines[i]) for i in var_fines]
    fines_mines_df['Cost']=[output(var_fines[i]*mines.loc[mines['possibilities']==i,'Total Cost Fines'].item()) for i in var_fines]
    fines_mines_df=fines_mines_df.loc[fines_mines_df['Quantity(Tonnes)']!=0].round(1).reset_index(drop=True)
    fines_mines_df=fines_mines_df.iloc[:,1:]
    fines_mines_df


    # In[109]:


    lumps_mines_dict={'Possibility':[]}
    lumps_mines_df=pd.DataFrame(data=lumps_mines_dict)
    lumps_mines_df['Possibility']=[i for i in var_lumps]
    #for i in lumps_mines_df['Possibility']:
    lumps_mines_df['Mine']=[mines_lumps.loc[mines_lumps['possibilities']==i,'Mine'].item() for i in lumps_mines_df['Possibility']]
    lumps_mines_df['Mine Owner']=[mines_lumps.loc[mines_lumps['possibilities']==i,'Mine Owners'].item() for i in lumps_mines_df['Possibility']]
    lumps_mines_df['Plant']=[mines_lumps.loc[mines_lumps['possibilities']==i,'Plant'].item() for i in lumps_mines_df['Possibility']]
    lumps_mines_df['Siding']=[mines_lumps.loc[mines_lumps['possibilities']==i,'Siding'].item() for i in lumps_mines_df['Possibility']]
    lumps_mines_df['Quantity(Tonnes)']=[output(var_lumps[i]) for i in var_lumps]
    lumps_mines_df['Cost']=[output(var_lumps[i]*mines_lumps.loc[mines_lumps['possibilities']==i,'Total Cost Lumps'].item()) for i in var_lumps]
    lumps_mines_df=lumps_mines_df.loc[lumps_mines_df['Quantity(Tonnes)']!=0].round(1).reset_index(drop=True)
    lumps_mines_df=lumps_mines_df.iloc[:,1:]
    lumps_mines_df


    # In[110]:


    pellet_to_P1_bf_dict={'Possibility':[]}
    pellet_to_P1_bf_df=pd.DataFrame(data=pellet_to_P1_bf_dict)
    pellet_to_P1_bf_df['Possibility']=[i for i in var_fines_bf_P1]
    #for i in pellet_to_P1_bf_df['Possibility']:
    pellet_to_P1_bf_df['Mine']=[mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Mine'].item() for i in pellet_to_P1_bf_df['Possibility']]
    pellet_to_P1_bf_df['Mine Owner']=[mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Mine Owners'].item() for i in pellet_to_P1_bf_df['Possibility']]
    pellet_to_P1_bf_df['Plant']=[mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Plant'].item() for i in pellet_to_P1_bf_df['Possibility']]
    pellet_to_P1_bf_df['Siding']=[mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Siding'].item() for i in pellet_to_P1_bf_df['Possibility']]
    pellet_to_P1_bf_df['Quantity(In Tonnes)']=[output(var_fines_bf_P1[i]) for i in var_fines_bf_P1]
    pellet_to_P1_bf_df['Cost']=[output(var_fines_bf_P1[i]*mines_bf_P1.loc[mines_bf_P1['possibilities']==i,'Total Cost Fines'].item()) for i in var_fines_bf_P1]
    pellet_to_P1_bf_df=pellet_to_P1_bf_df.loc[pellet_to_P1_bf_df['Quantity(In Tonnes)']!=0].round(1).reset_index(drop=True)
    pellet_to_P1_bf_df=pellet_to_P1_bf_df.iloc[:,1:]
    pellet_to_P1_bf_df


    # In[111]:


    pellet_to_P2_bf_dict={'Possibility':[]}
    pellet_to_P2_bf_df=pd.DataFrame(data=pellet_to_P2_bf_dict)
    pellet_to_P2_bf_df['Possibility']=[i for i in var_fines_bf_P2]
    #for i in pellet_to_P2_bf_df['Possibility']:
    pellet_to_P2_bf_df['Mine']=[mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Mine'].item() for i in pellet_to_P2_bf_df['Possibility']]
    pellet_to_P2_bf_df['Mine Owner']=[mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Mine Owners'].item() for i in pellet_to_P2_bf_df['Possibility']]
    pellet_to_P2_bf_df['Plant']=[mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Plant'].item() for i in pellet_to_P2_bf_df['Possibility']]
    pellet_to_P2_bf_df['Siding']=[mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Siding'].item() for i in pellet_to_P2_bf_df['Possibility']]
    pellet_to_P2_bf_df['Quantity(In Tonnes)']=[output(var_fines_bf_P2[i]) for i in var_fines_bf_P2]
    pellet_to_P2_bf_df['Cost']=[output(var_fines_bf_P2[i]*mines_bf_P2.loc[mines_bf_P2['possibilities']==i,'Total Cost Fines'].item()) for i in var_fines_bf_P2]
    pellet_to_P2_bf_df=pellet_to_P2_bf_df.loc[pellet_to_P2_bf_df['Quantity(In Tonnes)']!=0].round(1).reset_index(drop=True)
    pellet_to_P2_bf_df=pellet_to_P2_bf_df.iloc[:,1:]
    pellet_to_P2_bf_df


    # In[112]:


    pellet_to_P1_dri_dict={'Possibility':[]}
    pellet_to_P1_dri_df=pd.DataFrame(data=pellet_to_P1_dri_dict)
    pellet_to_P1_dri_df['Possibility']=[i for i in var_fines_dri_P1]
    #for i in pellet_to_P1_dri_df['Possibility']:
    pellet_to_P1_dri_df['Mine']=[mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Mine'].item() for i in pellet_to_P1_dri_df['Possibility']]
    pellet_to_P1_dri_df['Mine Owner']=[mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Mine Owners'].item() for i in pellet_to_P1_dri_df['Possibility']]
    pellet_to_P1_dri_df['Plant']=[mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Plant'].item() for i in pellet_to_P1_dri_df['Possibility']]
    pellet_to_P1_dri_df['Siding']=[mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Siding'].item() for i in pellet_to_P1_dri_df['Possibility']]
    pellet_to_P1_dri_df['Quantity(In Tonnes)']=[output(var_fines_dri_P1[i]) for i in var_fines_dri_P1]
    pellet_to_P1_dri_df['Cost']=[output(var_fines_dri_P1[i]*mines_dri_P1.loc[mines_dri_P1['possibilities']==i,'Total Cost Fines'].item()) for i in var_fines_dri_P1]
    pellet_to_P1_dri_df=pellet_to_P1_dri_df.loc[pellet_to_P1_dri_df['Quantity(In Tonnes)']!=0].round(1).reset_index(drop=True)
    pellet_to_P1_dri_df=pellet_to_P1_dri_df.iloc[:,1:]
    pellet_to_P1_dri_df


    # In[113]:


    pellet_to_P2_dri_dict={'Possibility':[]}
    pellet_to_P2_dri_df=pd.DataFrame(data=pellet_to_P2_dri_dict)
    pellet_to_P2_dri_df['Possibility']=[i for i in var_fines_dri_P2]
    #for i in pellet_to_P2_dri_df['Possibility']:
    pellet_to_P2_dri_df['Mine']=[mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Mine'].item() for i in pellet_to_P2_dri_df['Possibility']]
    pellet_to_P2_dri_df['Mine Owner']=[mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Mine Owners'].item() for i in pellet_to_P2_dri_df['Possibility']]
    pellet_to_P2_dri_df['Plant']=[mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Plant'].item() for i in pellet_to_P2_dri_df['Possibility']]
    pellet_to_P2_dri_df['Siding']=[mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Siding'].item() for i in pellet_to_P2_dri_df['Possibility']]
    pellet_to_P2_dri_df['Quantity(In Tonnes)']=[output(var_fines_dri_P2[i]) for i in var_fines_dri_P2]
    pellet_to_P2_dri_df['Cost']=[output(var_fines_dri_P2[i]*mines_dri_P2.loc[mines_dri_P2['possibilities']==i,'Total Cost Fines'].item()) for i in var_fines_dri_P2]
    pellet_to_P2_dri_df=pellet_to_P2_dri_df.loc[pellet_to_P2_dri_df['Quantity(In Tonnes)']!=0].round(1).reset_index(drop=True)
    pellet_to_P2_dri_df=pellet_to_P2_dri_df.iloc[:,1:]
    pellet_to_P2_dri_df


    # In[114]:


    pellet_to_market_dict={'Possibility':[]}
    pellet_to_market_df=pd.DataFrame(data=pellet_to_market_dict)
    pellet_to_market_df['Possibility']=[i for i in var_fines_pellet_market]
    #for i in pellet_to_market_df['Possibility']:
    pellet_to_market_df['Mine']=[mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Mine'].item() for i in pellet_to_market_df['Possibility']]
    pellet_to_market_df['Mine Owner']=[mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Mine Owners'].item() for i in pellet_to_market_df['Possibility']]
    pellet_to_market_df['Plant']=[mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Plant'].item() for i in pellet_to_market_df['Possibility']]
    pellet_to_market_df['Siding']=[mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Siding'].item() for i in pellet_to_market_df['Possibility']]
    pellet_to_market_df['Quantity(In Tonnes)']=[output(var_fines_pellet_market[i]) for i in var_fines_pellet_market]
    pellet_to_market_df['Cost']=[output(var_fines_pellet_market[i]*mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Total Cost Fines'].item()) for i in var_fines_pellet_market]
    pellet_to_market_df=pellet_to_market_df.loc[pellet_to_market_df['Quantity(In Tonnes)']!=0].round(1).reset_index(drop=True)
    pellet_to_market_df=pellet_to_market_df.iloc[:,1:]
    pellet_to_market_df


    # In[115]:


    #Purchase Plan
    purchase_plan_dict={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df=pd.DataFrame(data=purchase_plan_dict)
    purchase_plan_df['Output']=[i for i in fines_mines_df['Mine']]
    purchase_plan_df['Quantity(Tonnes)']=[i for i in fines_mines_df['Quantity(Tonnes)']]
    purchase_plan_df['Cost']=fines_mines_df['Cost']
    purchase_plan_df['Main Plant']=fines_mines_df['Plant']
    purchase_plan_df['Serving Plant']=fines_mines_df['Plant']
    purchase_plan_df['Mode']=fines_mines_df['Siding']
    purchase_plan_df['Category 1']='Iron Ore Fines'
    purchase_plan_df['Category 2']='Iron Ore Fines'
    purchase_plan_df['Company']='C1'
    purchase_plan_df['Mode'].append(fines_mines_df['Plant'])#[i for i in fines_mines_df['Mine']]
    purchase_plan_df


    # In[116]:


    #Purchase Plan
    purchase_plan_dict1={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df_1=pd.DataFrame(data=purchase_plan_dict1)
    purchase_plan_df_1['Output']=[i for i in lumps_mines_df['Mine']]
    purchase_plan_df_1['Quantity(Tonnes)']=[i for i in lumps_mines_df['Quantity(Tonnes)']]
    purchase_plan_df_1['Cost']=lumps_mines_df['Cost']
    purchase_plan_df_1['Main Plant']=lumps_mines_df['Plant']
    purchase_plan_df_1['Serving Plant']=lumps_mines_df['Plant']
    purchase_plan_df_1['Mode']=lumps_mines_df['Siding']
    purchase_plan_df_1['Category 1']='Iron Ore Lumps'
    purchase_plan_df_1['Category 2']='Iron Ore Lumps'
    purchase_plan_df_1['Company']='C1'
    #purchase_plan_df_1['Mode'].append(fines_mines_df['Plant'])#[i for i in fines_mines_df['Mine']]
    purchase_plan_df_1


    # In[117]:


    #Purchase Plan
    purchase_plan_dict2={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df_2=pd.DataFrame(data=purchase_plan_dict2)
    purchase_plan_df_2['Output']=[i for i in pellet_to_P1_bf_df['Mine']]
    purchase_plan_df_2['Quantity(Tonnes)']=[i for i in pellet_to_P1_bf_df['Quantity(In Tonnes)']]
    purchase_plan_df_2['Cost']=pellet_to_P1_bf_df['Cost']
    purchase_plan_df_2['Main Plant']=pellet_to_P1_bf_df['Plant']
    purchase_plan_df_2['Serving Plant']='P1'
    purchase_plan_df_2['Mode']=pellet_to_P1_bf_df['Siding']
    purchase_plan_df_2['Category 1']='Iron Ore Pellet'
    purchase_plan_df_2['Category 2']='Iron Ore Pellet'
    purchase_plan_df_2['Company']='C1'
    #purchase_plan_df_2['Mode'].append(fines_mines_df['Plant'])#[i for i in fines_mines_df['Mine']]
    purchase_plan_df_2


    # In[118]:


    #Purchase Plan
    purchase_plan_dict3={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df_3=pd.DataFrame(data=purchase_plan_dict3)
    purchase_plan_df_3['Output']=[i for i in pellet_to_P2_bf_df['Mine']]
    purchase_plan_df_3['Quantity(Tonnes)']=[i for i in pellet_to_P2_bf_df['Quantity(In Tonnes)']]
    purchase_plan_df_3['Cost']=pellet_to_P2_bf_df['Cost']
    purchase_plan_df_3['Main Plant']=pellet_to_P2_bf_df['Plant']
    purchase_plan_df_3['Serving Plant']='P2'
    purchase_plan_df_3['Mode']=pellet_to_P2_bf_df['Siding']
    purchase_plan_df_3['Category 1']='Iron Ore Pellet'
    purchase_plan_df_3['Category 2']='Iron Ore Pellet'
    purchase_plan_df_3['Company']='C1'
    #purchase_plan_df_3['Mode'].append(fines_mines_df['Plant'])#[i for i in fines_mines_df['Mine']]
    purchase_plan_df_3


    # In[119]:


    #Purchase Plan
    purchase_plan_dict4={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df_4=pd.DataFrame(data=purchase_plan_dict4)
    purchase_plan_df_4['Output']=[i for i in pellet_to_P1_dri_df['Mine']]
    purchase_plan_df_4['Quantity(Tonnes)']=[i for i in pellet_to_P1_dri_df['Quantity(In Tonnes)']]
    purchase_plan_df_4['Cost']=pellet_to_P1_dri_df['Cost']
    purchase_plan_df_4['Main Plant']=pellet_to_P1_dri_df['Plant']
    purchase_plan_df_4['Serving Plant']='P1'
    purchase_plan_df_4['Mode']=pellet_to_P1_dri_df['Siding']
    purchase_plan_df_4['Category 1']='Iron Ore Pellet'
    purchase_plan_df_4['Category 2']='Iron Ore Pellet'
    purchase_plan_df_4['Company']='C1'
    #purchase_plan_df_4['Mode'].append(fines_mines_df['Plant'])#[i for i in fines_mines_df['Mine']]
    purchase_plan_df_4


    # In[120]:


    #Purchase Plan
    purchase_plan_dict5={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df_5=pd.DataFrame(data=purchase_plan_dict5)
    purchase_plan_df_5['Output']=[i for i in pellet_to_P2_dri_df['Mine']]
    purchase_plan_df_5['Quantity(Tonnes)']=[i for i in pellet_to_P2_dri_df['Quantity(In Tonnes)']]
    purchase_plan_df_5['Cost']=pellet_to_P2_dri_df['Cost']
    purchase_plan_df_5['Main Plant']=pellet_to_P2_dri_df['Plant']
    purchase_plan_df_5['Serving Plant']='P2'
    purchase_plan_df_5['Mode']=pellet_to_P2_dri_df['Siding']
    purchase_plan_df_5['Category 1']='Iron Ore Pellet'
    purchase_plan_df_5['Category 2']='Iron Ore Pellet'
    purchase_plan_df_5['Company']='C1'
    #purchase_plan_df_5['Mode'].append(fines_mines_df['Plant'])#[i for i in fines_mines_df['Mine']]
    purchase_plan_df_5


    # In[121]:


    #Purchase Plan
    purchase_plan_dict6={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df_6=pd.DataFrame(data=purchase_plan_dict6)
    purchase_plan_df_6['Output']=[i for i in pellet_to_market_df['Mine']]
    purchase_plan_df_6['Quantity(Tonnes)']=[i for i in pellet_to_market_df['Quantity(In Tonnes)']]
    purchase_plan_df_6['Cost']=pellet_to_market_df['Cost']
    purchase_plan_df_6['Main Plant']=pellet_to_market_df['Plant']
    purchase_plan_df_6['Serving Plant']='Market'
    purchase_plan_df_6['Mode']=pellet_to_market_df['Siding']
    purchase_plan_df_6['Category 1']='Iron Ore Pellet'
    purchase_plan_df_6['Category 2']='Iron Ore Pellet'
    purchase_plan_df_6['Company']='C1'
    #purchase_plan_df_6['Mode'].append(fines_mines_df['Plant'])#[i for i in fines_mines_df['Mine']]
    purchase_plan_df_6


    # In[122]:


    #Purchase Plan
    purchase_plan_dict7={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df_7=pd.DataFrame(data=purchase_plan_dict7)
    purchase_plan_df_7['Output']=coal_output_P1_df['Coal']
    purchase_plan_df_7['Quantity(Tonnes)']=coal_output_P1_df['Quantity in Tonnes']
    purchase_plan_df_7['Cost']=coal_output_P1_df['Cost']
    purchase_plan_df_7['Company']='C1'
    purchase_plan_df_7['Main Plant']='P1'
    purchase_plan_df_7['Serving Plant']='P1'
    purchase_plan_df_7['Mode']='None'
    purchase_plan_df_7['Category 2']='Coking Coal'
    purchase_plan_df_7['Category 1']=[coals_P1.loc[coals_P1['coals_P1']==i,'type'].item() for i in coal_output_P1_df['Coal']]
    purchase_plan_df_7


    # In[123]:


    #Purchase Plan
    purchase_plan_dict8={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df_8=pd.DataFrame(data=purchase_plan_dict8)
    purchase_plan_df_8['Output']=coal_output_P2_df['Coal']
    purchase_plan_df_8['Quantity(Tonnes)']=coal_output_P2_df['Quantity in Tonnes']
    purchase_plan_df_8['Cost']=coal_output_P2_df['Cost']
    purchase_plan_df_8['Company']='C1'
    purchase_plan_df_8['Main Plant']='P2'
    purchase_plan_df_8['Serving Plant']='P2'
    purchase_plan_df_8['Mode']='None'
    purchase_plan_df_8['Category 2']='Coking Coal'
    purchase_plan_df_8['Category 1']=[coals.loc[coals['coals']==i,'type'].item() for i in coal_output_P2_df['Coal']]
    purchase_plan_df_8


    # In[124]:


    #Purchase Plan
    purchase_plan_dict9={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df_9=pd.DataFrame(data=purchase_plan_dict9)
    purchase_plan_df_9['Output']=sinter_output_df_P1['Material']
    purchase_plan_df_9['Quantity(Tonnes)']=sinter_output_df_P1['Quantity(In Tonnes)']
    purchase_plan_df_9['Cost']=sinter_output_df_P1['Cost']
    purchase_plan_df_9['Company']='C1'
    purchase_plan_df_9['Main Plant']='P1'
    purchase_plan_df_9['Serving Plant']='P1'
    purchase_plan_df_9['Mode']='None'
    purchase_plan_df_9['Category 2']='Flux'
    purchase_plan_df_9['Category 1']=sinter_output_df_P1['Material']
    purchase_plan_df_9


    # In[125]:


    #Purchase Plan
    purchase_plan_dict10={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df_10=pd.DataFrame(data=purchase_plan_dict10)
    purchase_plan_df_10['Output']=sinter_output_df_P2['Material']
    purchase_plan_df_10['Quantity(Tonnes)']=sinter_output_df_P2['Quantity(In Tonnes)']
    purchase_plan_df_10['Cost']=sinter_output_df_P2['Cost']
    purchase_plan_df_10['Company']='C1'
    purchase_plan_df_10['Main Plant']='P2'
    purchase_plan_df_10['Serving Plant']='P2'
    purchase_plan_df_10['Mode']='None'
    purchase_plan_df_10['Category 2']='Flux'
    purchase_plan_df_10['Category 1']=sinter_output_df_P2['Material']
    purchase_plan_df_10


    # In[126]:


    #Purchase Plan
    purchase_plan_dict11={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df_11=pd.DataFrame(data=purchase_plan_dict11)
    purchase_plan_df_11['Output']=bmo_output_df_P1['Material']
    purchase_plan_df_11['Quantity(Tonnes)']=bmo_output_df_P1['Quantity(In Tonnes)']
    purchase_plan_df_11['Cost']=bmo_output_df_P1['Cost']
    purchase_plan_df_11['Company']='C1'
    purchase_plan_df_11['Main Plant']='P1'
    purchase_plan_df_11['Serving Plant']='P1'
    purchase_plan_df_11['Mode']='None'
    purchase_plan_df_11['Category 2']='Flux'
    purchase_plan_df_11['Category 1']=bmo_output_df_P1['Material']
    purchase_plan_df_11
    #bmo_output_df_P1


    # In[127]:


    #Purchase Plan
    purchase_plan_dict12={'Company':[],'Main Plant':[],'Serving Plant':[],'Output':[],'Mode':[],'Category 1':[],'Category 2':[],
                       'Quantity(Tonnes)':[],'Cost':[]}
    purchase_plan_df_12=pd.DataFrame(data=purchase_plan_dict12)
    purchase_plan_df_12['Output']=bmo_output_df_P2['Material']
    purchase_plan_df_12['Quantity(Tonnes)']=bmo_output_df_P2['Quantity(In Tonnes)']
    purchase_plan_df_12['Cost']=bmo_output_df_P2['Cost']
    purchase_plan_df_12['Company']='C1'
    purchase_plan_df_12['Main Plant']='P2'
    purchase_plan_df_12['Serving Plant']='P2'
    purchase_plan_df_12['Mode']='None'
    purchase_plan_df_12['Category 2']='Flux'
    purchase_plan_df_12['Category 1']=bmo_output_df_P2['Material']
    purchase_plan_df_12
    #bmo_output_df_P2


    # In[534]:


    purchase_plan_df_final=pd.concat([purchase_plan_df,purchase_plan_df_1,purchase_plan_df_2,purchase_plan_df_3
                                      ,purchase_plan_df_4,purchase_plan_df_5,purchase_plan_df_6,purchase_plan_df_7
                                      ,purchase_plan_df_8,purchase_plan_df_9,purchase_plan_df_10,purchase_plan_df_11,purchase_plan_df_12],ignore_index=True)
    purchase_plan_df_final['Cost']=purchase_plan_df_final['Cost']/10000000
    purchase_plan_df_final=purchase_plan_df_final.rename(columns={'Cost':'Cost(In Cr)'})
    
    #Appending Excel Files
    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False, 
                           **to_excel_kwargs):
        from openpyxl import load_workbook

        import pandas as pd

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError


        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()    
    purchase_plan_df_final['No. of Planned Days']=no_of_days_input
    append_df_to_excel("Output/Purchase Plan Output.xlsx",purchase_plan_df_final,index=False)
    #purchase_plan_df_final.to_excel("Output/Purchase Plan Output.xlsx",index=False)
    purchase_plan_df_final 


    # In[480]:


    #output ratio
    output_ratio_dict1={'Company':[],'Main Plant':[],'Serving Plant':[],'Category':[],'% Mix':[]}
    output_ratio_bf_df1=pd.DataFrame(data=output_ratio_dict1)
    output_ratio_bf_df1['Category']=spl_output_df_P1.index
    output_ratio_bf_df1['% Mix']=[spl_output_df_P1.loc[spl_output_df_P1['Parameter']==i,'% Mix'].item() for i in spl_output_df_P1.index]
    output_ratio_bf_df1['Company']='C1'
    output_ratio_bf_df1['Main Plant']='P1'
    output_ratio_bf_df1['Serving Plant']='Blast Furnace'
    #output_ratio_bf_df1


    # In[479]:


    #output ratio
    output_ratio_dict2={'Company':[],'Main Plant':[],'Serving Plant':[],'Category':[],'% Mix':[]}
    output_ratio_bf_df2=pd.DataFrame(data=output_ratio_dict2)
    output_ratio_bf_df2['Category']=spl_output_df_P2.index
    output_ratio_bf_df2['% Mix']=[spl_output_df_P2.loc[spl_output_df_P2['Parameter']==i,'% Mix'].item() for i in spl_output_df_P2.index]
    output_ratio_bf_df2['Company']='C1'
    output_ratio_bf_df2['Main Plant']='P2'
    output_ratio_bf_df2['Serving Plant']='Blast Furnace'
    #output_ratio_bf_df2


    # In[478]:


    #output ratio
    output_ratio_dict3={'Company':[],'Main Plant':[],'Serving Plant':[],'Category':[],'% Mix':[]}
    output_ratio_coke_P1_df3=pd.DataFrame(data=output_ratio_dict3)
    output_ratio_coke_P1_df3['Category']=coal_output_P1_df['Category'].unique()
    output_ratio_coke_P1_df3['% Mix']=[lpSum(coal_output_P1_df.loc[coal_output_P1_df['Category']==i,'Mix%']).value() for i in coal_output_P1_df['Category'].unique()]
    output_ratio_coke_P1_df3['Company']='C1'
    output_ratio_coke_P1_df3['Main Plant']='P1'
    output_ratio_coke_P1_df3['Serving Plant']='Coke Oven'
    #output_ratio_coke_P1_df3


    # In[477]:


    #output ratio
    output_ratio_dict4={'Company':[],'Main Plant':[],'Serving Plant':[],'Category':[],'% Mix':[]}
    output_ratio_coke_P2_df4=pd.DataFrame(data=output_ratio_dict4)
    output_ratio_coke_P2_df4['Category']=coal_output_P2_df['Category'].unique()
    output_ratio_coke_P2_df4['% Mix']=[lpSum(coal_output_P2_df.loc[coal_output_P2_df['Category']==i,'Mix%']).value() for i in coal_output_P2_df['Category'].unique()]
    output_ratio_coke_P2_df4['Company']='C1'
    output_ratio_coke_P2_df4['Main Plant']='P2'
    output_ratio_coke_P2_df4['Serving Plant']='Coke Oven'
    #output_ratio_coke_P2_df4


    # In[535]:


    output_ratio_final=pd.concat([output_ratio_bf_df1,output_ratio_bf_df2,output_ratio_coke_P1_df3,output_ratio_coke_P2_df4],ignore_index=True)
    output_ratio_final['No. of Planned Days']=no_of_days_input
    append_df_to_excel("Output/Ratio Output.xlsx",output_ratio_final,index=False)
 
    #output_ratio_final.to_excel("Output/Ratio Output.xlsx",index=False)
    output_ratio_final


    # In[539]:


    #Contribution Output
    pellet_sales_revenue=lpSum(var_fines_pellet_market[i]*(values_P2.loc[values_P2['Parameter']=='Realization from pellet Sale','Value'].item()+mines_pellet_market.loc[mines_pellet_market['possibilities']==i,'Total Cost Fines'].item()/pellet_conversion_market) for i in var_fines_pellet_market).value()
    contribution_output_dict={'Company':['C1' for i in range(10)],'Main Plant':['P1' for i in range(10)],
                              'Material':['Blast Furnace','Sinter','Lumps','Limestone','Dolomite','Quartz','Coke (Purchase)','Coke (Nut)','PCI','Coke (C1)'],
                             'Plant':['Blast Furnace','Sinter','Lumps','Flux','Flux','Flux','Coke','Coke','Coke','Coke']}

    contribution_output_df=pd.DataFrame(data=contribution_output_dict)
    contribution_output_df['Serving Plant']='Blast Furnace'
    contribution_output_df.loc[contribution_output_df['Material']=='Blast Furnace','Revenue']=int_total_realization_P1.value()
    contribution_output_df.loc[contribution_output_df['Material']=='Blast Furnace','Cost']=int_total_cost_P1.value()
    contribution_output_df.loc[contribution_output_df['Material']=='Blast Furnace','Contribution']=int_total_contribution_P1.value()
    contribution_output_df['Cost'][3:9]=[(var_bmo_P1[i]*bmo_materials.loc[bmo_materials['Material']==i,'Cost/MT'].item()).value() for i in contribution_output_df['Material'][3:9]]
    contribution_output_df.loc[contribution_output_df['Material']=='Lumps','Cost']=lpSum(smo_obj_lumps_P1).value()
    contribution_output_df.loc[contribution_output_df['Material']=='Coke (C1)','Cost']=Total_Cost_of_Usable_Coke_at_Blast_Furnace_req_P1.value()
    contribution_output_df.loc[contribution_output_df['Material']=='Sinter','Cost']=Total_Cost_P1.value()
    contribution_output_df.loc[contribution_output_df['Revenue'].isna(),'Revenue']=''
    contribution_output_df.loc[contribution_output_df['Contribution'].isna(),'Contribution']=''
    contribution_output_df
    #Contribution Output
    contribution_output_dict1={'Company':['C1' for i in range(10)],'Main Plant':['P2' for i in range(10)],
                              'Material':['Blast Furnace','Sinter','Lumps','Limestone','Dolomite','Quartz','Coke (Purchase)','Coke (Nut)','PCI','Coke (C1)'],
                             'Plant':['Blast Furnace','Sinter','Lumps','Flux','Flux','Flux','Coke','Coke','Coke','Coke']}

    contribution_output_df1=pd.DataFrame(data=contribution_output_dict1)
    contribution_output_df1['Serving Plant']='Blast Furnace'
    contribution_output_df1.loc[contribution_output_df1['Material']=='Blast Furnace','Revenue']=int_total_realization_P2.value()
    contribution_output_df1.loc[contribution_output_df1['Material']=='Blast Furnace','Cost']=int_total_cost_P2.value()
    contribution_output_df1.loc[contribution_output_df1['Material']=='Blast Furnace','Contribution']=int_total_contribution_P2.value()
    contribution_output_df1['Cost'][3:9]=[(var_bmo_P2[i]*bmo_materials.loc[bmo_materials['Material']==i,'Cost/MT'].item()).value() for i in contribution_output_df1['Material'][3:9]]
    contribution_output_df1.loc[contribution_output_df1['Material']=='Lumps','Cost']=lpSum(smo_obj_lumps_P2).value()
    contribution_output_df1.loc[contribution_output_df1['Material']=='Coke (C1)','Cost']=Total_Cost_of_Usable_Coke_at_Blast_Furnace_req.value()
    contribution_output_df1.loc[contribution_output_df1['Material']=='Sinter','Cost']=Total_Cost_P2.value()
    contribution_output_df1.loc[contribution_output_df1['Revenue'].isna(),'Revenue']=''
    contribution_output_df1.loc[contribution_output_df1['Contribution'].isna(),'Contribution']=''
    contribution_output_df1
    #Contribution Output
    contribution_output_dict2={'Company':['C1' for i in range(5)],'Main Plant':['Barbil' for i in range(5)],
                              'Material':['Pellet to P1','Pellet to P2','Pellet to P1','Pellet to P2','Pellet Sales']}

    contribution_output_df2=pd.DataFrame(data=contribution_output_dict2)
    contribution_output_df2['Plant']='Pellet'
    contribution_output_df2['Serving Plant']=['Blast Furnace','Blast Furnace','DRI','DRI','Market']
    contribution_output_df2['Revenue']=['','',total_realisation_P1_dri.value(),total_realisation_P2_dri.value(),pellet_sales_revenue]
    contribution_output_df2['Cost']=[cost_objective_P1_bf.value(),cost_objective_P2_bf.value(),total_cost_P1_dri.value(),total_cost_P2_dri.value(),lpSum(cost_objective_barbil_list).value()]
    contribution_output_df2['Contribution']=['','',total_contribution_P1_dri.value(),total_contribution_P2_dri.value(),(pellet_sales_revenue-lpSum(cost_objective_barbil_list).value())]
    contribution_output_df2
    contribution_output_final=pd.concat([contribution_output_df,contribution_output_df1,contribution_output_df2],ignore_index=True)
    contribution_output_final['No. of Planned Days']=no_of_days_input
    append_df_to_excel("Output/Contribution.xlsx",contribution_output_final,index=False)
     #contribution_output_final.to_excel("Output/Contribution Output.xlsx",index=False)
    


    # In[540]:


    #Quality output
    coke_bf_P1_chemistry_dict_P1={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','P','S'],
                                           'Actual':[0 for i in range(0,7)]}
    coke_bf_P1_chemistry_output_df_P1=pd.DataFrame(data=coke_bf_P1_chemistry_dict_P1)
    if lpSum(var_coke_P1).value()!=0:
        coke_bf_P1_chemistry_output_df_P1.loc[coke_bf_P1_chemistry_output_df_P1['Parameter']=='Fe','Actual']=(lpSum(coals_P1['Fe(new)_P1']*100)).value()/(lpSum(var_coke_P1).value()-lpSum(coals_P1['feeding_dmt_P1']).value())
        coke_bf_P1_chemistry_output_df_P1.loc[coke_bf_P1_chemistry_output_df_P1['Parameter']=='SiO2','Actual']=(lpSum(coals_P1['SiO2(new)_P1']*100)).value()/(lpSum(var_coke_P1).value()-lpSum(coals_P1['feeding_dmt_P1']).value())
        coke_bf_P1_chemistry_output_df_P1.loc[coke_bf_P1_chemistry_output_df_P1['Parameter']=='Al2O3','Actual']=(lpSum(coals_P1['Al2O3(new)_P1']*100)).value()/(lpSum(var_coke_P1).value()-lpSum(coals_P1['feeding_dmt_P1']).value())
        coke_bf_P1_chemistry_output_df_P1.loc[coke_bf_P1_chemistry_output_df_P1['Parameter']=='CaO','Actual']=(lpSum(coals_P1['CaO(new)_P1']*100)).value()/(lpSum(var_coke_P1).value()-lpSum(coals_P1['feeding_dmt_P1']).value())
        coke_bf_P1_chemistry_output_df_P1.loc[coke_bf_P1_chemistry_output_df_P1['Parameter']=='MgO','Actual']=(lpSum(coals_P1['MgO(new)_P1']*100)).value()/(lpSum(var_coke_P1).value()-lpSum(coals_P1['feeding_dmt_P1']).value())
        coke_bf_P1_chemistry_output_df_P1.loc[coke_bf_P1_chemistry_output_df_P1['Parameter']=='P','Actual']=(lpSum(coals_P1['P(new)_P1']*100)).value()/(lpSum(var_coke_P1).value()-lpSum(coals_P1['feeding_dmt_P1']).value())
        coke_bf_P1_chemistry_output_df_P1.loc[coke_bf_P1_chemistry_output_df_P1['Parameter']=='S','Actual']=(lpSum(coals_P1['S(new)_P1']*100)).value()/(lpSum(var_coke_P1).value()-lpSum(coals_P1['feeding_dmt_P1']).value())
        coke_bf_P1_chemistry_output_df_P1=coke_bf_P1_chemistry_output_df_P1.round()
    coke_bf_P1_chemistry_output_df_P1
    #
    coke_bf_P2_chemistry_dict_P2={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','P','S'],
                                           'Actual':[0 for i in range(0,7)]}
    coke_bf_P2_chemistry_output_df_P2=pd.DataFrame(data=coke_bf_P2_chemistry_dict_P2)
    if lpSum(var_coke_P2).value()!=0:
        coke_bf_P2_chemistry_output_df_P2.loc[coke_bf_P2_chemistry_output_df_P2['Parameter']=='Fe','Actual']=(lpSum(coals['Fe(new)_P2']*100)).value()/(lpSum(var_coke_P2).value()-lpSum(coals['feeding_dmt_P2']).value())
        coke_bf_P2_chemistry_output_df_P2.loc[coke_bf_P2_chemistry_output_df_P2['Parameter']=='SiO2','Actual']=(lpSum(coals['SiO2(new)_P2']*100)).value()/(lpSum(var_coke_P2).value()-lpSum(coals['feeding_dmt_P2']).value())
        coke_bf_P2_chemistry_output_df_P2.loc[coke_bf_P2_chemistry_output_df_P2['Parameter']=='Al2O3','Actual']=(lpSum(coals['Al2O3(new)_P2']*100)).value()/(lpSum(var_coke_P2).value()-lpSum(coals['feeding_dmt_P2']).value())
        coke_bf_P2_chemistry_output_df_P2.loc[coke_bf_P2_chemistry_output_df_P2['Parameter']=='CaO','Actual']=(lpSum(coals['CaO(new)_P2']*100)).value()/(lpSum(var_coke_P2).value()-lpSum(coals['feeding_dmt_P2']).value())
        coke_bf_P2_chemistry_output_df_P2.loc[coke_bf_P2_chemistry_output_df_P2['Parameter']=='MgO','Actual']=(lpSum(coals['MgO(new)_P2']*100)).value()/(lpSum(var_coke_P2).value()-lpSum(coals['feeding_dmt_P2']).value())
        coke_bf_P2_chemistry_output_df_P2.loc[coke_bf_P2_chemistry_output_df_P2['Parameter']=='P','Actual']=(lpSum(coals['P(new)_P2']*100)).value()/(lpSum(var_coke_P2).value()-lpSum(coals['feeding_dmt_P2']).value())
        coke_bf_P2_chemistry_output_df_P2.loc[coke_bf_P2_chemistry_output_df_P2['Parameter']=='S','Actual']=(lpSum(coals['S(new)_P2']*100)).value()/(lpSum(var_coke_P2).value()-lpSum(coals['feeding_dmt_P2']).value())
        coke_bf_P2_chemistry_output_df_P2=coke_bf_P2_chemistry_output_df_P2.round()


    # In[541]:


    #pellet_bf_P1_chemistry_output_df_P2['Actual']=
    fe_pellet_bf_P1=output((lpSum(mines_bf_P1['Fe(new)_fines']))*100)
    SiO2_pellet_bf_P1=output((lpSum(mines_bf_P1['SiO2(new)_fines']))*100)
    al2o3_pellet_bf_P1=output((lpSum(mines_bf_P1['Al2O3(new)_fines']))*100)
    cao_pellet_bf_P1=output((lpSum(mines_bf_P1['CaO(new)_fines']))*100)
    mgo_pellet_bf_P1=output((lpSum(mines_bf_P1['MgO(new)_fines']))*100)
    basicity_P1=output((lpSum(mines_bf_P1['CaO(new)_fines']))*100)
    pellet_bf_P1_chemistry_dict_P1={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','Basicity','Residue Production'],
                                       'Actual':[fe_pellet_bf_P1,SiO2_pellet_bf_P1,al2o3_pellet_bf_P1,cao_pellet_bf_P1,mgo_pellet_bf_P1,basicity_P1,lpSum(mines_bf_P1['Residue_fines']).value()]}
    pellet_bf_P1_chemistry_output_df_P1=pd.DataFrame(data=pellet_bf_P1_chemistry_dict_P1).round(1)
    pellet_bf_P1_chemistry_output_df_P1
    #lump_bf_P1_chemistry_output_df_P2['Actual']=
    fe_lump_bf_P1=output((lpSum(mines_lumps['Fe(new)_lumps']))*100)
    SiO2_lump_bf_P1=output((lpSum(mines_lumps['SiO2(new)_lumps']))*100)
    al2o3_lump_bf_P1=output((lpSum(mines_lumps['Al2O3(new)_lumps']))*100)
    cao_lump_bf_P1=output((lpSum(mines_lumps['CaO(new)_lumps']))*100)
    mgo_lump_bf_P1=output((lpSum(mines_lumps['MgO(new)_lumps']))*100)
    basicity_P1=output((lpSum(mines_lumps['CaO(new)_lumps']))*100)
    lump_bf_P1_chemistry_dict_P1={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','Basicity','Residue Production'],
                                       'Actual':[fe_lump_bf_P1,SiO2_lump_bf_P1,al2o3_lump_bf_P1,cao_lump_bf_P1,mgo_lump_bf_P1,basicity_P1,lpSum(mines_lumps['Residue_lumps']).value()]}
    lump_bf_P1_chemistry_output_df_P1=pd.DataFrame(data=lump_bf_P1_chemistry_dict_P1).round(1)
    lump_bf_P1_chemistry_output_df_P1
    #Limestone_bf_P1_chemistry_output_df_P2['Actual']=
    Limestone_bf_P1_chemistry_dict_P1={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','P','S'],
                                           'Actual':[0 for i in range(0,7)]}
    Limestone_bf_P1_chemistry_output_df_P1=pd.DataFrame(data=Limestone_bf_P1_chemistry_dict_P1)
    if lpSum(var_bmo_P1['Limestone']).value()!=0:
        Limestone_bf_P1_chemistry_output_df_P1.loc[Limestone_bf_P1_chemistry_output_df_P1['Parameter']=='Fe','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','Fe(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P1']).value())
        Limestone_bf_P1_chemistry_output_df_P1.loc[Limestone_bf_P1_chemistry_output_df_P1['Parameter']=='SiO2','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','SiO2(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P1']).value())
        Limestone_bf_P1_chemistry_output_df_P1.loc[Limestone_bf_P1_chemistry_output_df_P1['Parameter']=='Al2O3','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','Al2O3(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P1']).value())
        Limestone_bf_P1_chemistry_output_df_P1.loc[Limestone_bf_P1_chemistry_output_df_P1['Parameter']=='CaO','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','CaO(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P1']).value())
        Limestone_bf_P1_chemistry_output_df_P1.loc[Limestone_bf_P1_chemistry_output_df_P1['Parameter']=='MgO','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','MgO(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P1']).value())
        Limestone_bf_P1_chemistry_output_df_P1.loc[Limestone_bf_P1_chemistry_output_df_P1['Parameter']=='P','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','P(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P1']).value())
        Limestone_bf_P1_chemistry_output_df_P1.loc[Limestone_bf_P1_chemistry_output_df_P1['Parameter']=='S','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','S(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P1']).value())
        Limestone_bf_P1_chemistry_output_df_P1=Limestone_bf_P1_chemistry_output_df_P1.round()
    Limestone_bf_P1_chemistry_output_df_P1
    #Dolomite_bf_P1_chemistry_output_df_P2['Actual']=
    Dolomite_bf_P1_chemistry_dict_P1={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','P','S'],
                                           'Actual':[0 for i in range(0,7)]}
    Dolomite_bf_P1_chemistry_output_df_P1=pd.DataFrame(data=Dolomite_bf_P1_chemistry_dict_P1)
    if lpSum(var_bmo_P1['Dolomite']).value()!=0:
        Dolomite_bf_P1_chemistry_output_df_P1.loc[Dolomite_bf_P1_chemistry_output_df_P1['Parameter']=='Fe','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','Fe(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P1']).value())
        Dolomite_bf_P1_chemistry_output_df_P1.loc[Dolomite_bf_P1_chemistry_output_df_P1['Parameter']=='SiO2','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','SiO2(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P1']).value())
        Dolomite_bf_P1_chemistry_output_df_P1.loc[Dolomite_bf_P1_chemistry_output_df_P1['Parameter']=='Al2O3','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','Al2O3(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P1']).value())
        Dolomite_bf_P1_chemistry_output_df_P1.loc[Dolomite_bf_P1_chemistry_output_df_P1['Parameter']=='CaO','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','CaO(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P1']).value())
        Dolomite_bf_P1_chemistry_output_df_P1.loc[Dolomite_bf_P1_chemistry_output_df_P1['Parameter']=='MgO','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','MgO(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P1']).value())
        Dolomite_bf_P1_chemistry_output_df_P1.loc[Dolomite_bf_P1_chemistry_output_df_P1['Parameter']=='P','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','P(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P1']).value())
        Dolomite_bf_P1_chemistry_output_df_P1.loc[Dolomite_bf_P1_chemistry_output_df_P1['Parameter']=='S','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','S(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P1']).value())
        Dolomite_bf_P1_chemistry_output_df_P1=Dolomite_bf_P1_chemistry_output_df_P1.round(2)
    Dolomite_bf_P1_chemistry_output_df_P1
    #Quartz_bf_P1_chemistry_output_df_P2['Actual']=
    Quartz_bf_P1_chemistry_dict_P1={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','P','S'],
                                           'Actual':[0 for i in range(0,7)]}
    Quartz_bf_P1_chemistry_output_df_P1=pd.DataFrame(data=Quartz_bf_P1_chemistry_dict_P1)
    if lpSum(var_bmo_P1['Quartz']).value()!=0:
        Quartz_bf_P1_chemistry_output_df_P1.loc[Quartz_bf_P1_chemistry_output_df_P1['Parameter']=='Fe','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','Fe(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P1']).value())
        Quartz_bf_P1_chemistry_output_df_P1.loc[Quartz_bf_P1_chemistry_output_df_P1['Parameter']=='SiO2','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','SiO2(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P1']).value())
        Quartz_bf_P1_chemistry_output_df_P1.loc[Quartz_bf_P1_chemistry_output_df_P1['Parameter']=='Al2O3','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','Al2O3(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P1']).value())
        Quartz_bf_P1_chemistry_output_df_P1.loc[Quartz_bf_P1_chemistry_output_df_P1['Parameter']=='CaO','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','CaO(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P1']).value())
        Quartz_bf_P1_chemistry_output_df_P1.loc[Quartz_bf_P1_chemistry_output_df_P1['Parameter']=='MgO','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','MgO(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P1']).value())
        Quartz_bf_P1_chemistry_output_df_P1.loc[Quartz_bf_P1_chemistry_output_df_P1['Parameter']=='P','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','P(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P1']).value())
        Quartz_bf_P1_chemistry_output_df_P1.loc[Quartz_bf_P1_chemistry_output_df_P1['Parameter']=='S','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','S(new)_P1']*100)).value()/(lpSum(var_bmo_P1['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P1']).value())
        Quartz_bf_P1_chemistry_output_df_P1=Quartz_bf_P1_chemistry_output_df_P1.round(1)
    Quartz_bf_P1_chemistry_output_df_P1
    #Quality Output
    quality_output_P1_dict={'Company':['C1' for i in range(7)],'Main Plant':['P1' for i in range(7)],
                              'Serving Plant':['Blast Furnace'for i in range(7)],
                             'Category 1':['Coke','Sinter','Pellet','Lump','Limestone','Dolomite','Quartz'],
                         'Category 2':['Coke','Sinter','Pellet','Lump','Flux','Flux','Flux'],'CSR':['None' for i in range(7)],'Ash':['None' for i in range(7)],
                         'Moi':['None' for i in range(7)],'Fe':['None' for i in range(7)],'SiO2':['None' for i in range(7)],'Al2O3':['None' for i in range(7)],'CaO':['None' for i in range(7)]
                         ,'MgO':['None' for i in range(7)],'Basicity':['None' for i in range(7)],'FC':['None' for i in range(7)],'P':['None' for i in range(7)],'S':['None' for i in range(7)]}#,
                         #'VM':['None' for i in range(7) ],'CSN':['None' for i in range(7)],'MMR':['None' for i in range(7)],'Vitrinite':['None' for i in range(7)],
                         #'V9-V14':['None' for i in range(7)],'Sulphur':['None' for i in range(7)],'Ash':['None' for i in range(7)],
                         #'TM':['None' for i in range(7)],'FC':['None' for i in range(7)],'Log Fluidity':['None' for i in range(7)]}

    quality_output_P1_df=pd.DataFrame(data=quality_output_P1_dict)
    quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Coke','CSR']=[coke_parameters_P1_df.loc[coke_parameters_P1_df['Parameter']=='CSR','Value'].item()]
    quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Coke','Ash']=[coke_parameters_P1_df.loc[coke_parameters_P1_df['Parameter']=='Coke Ash','Value'].item()]
    quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Coke','S']=[coke_parameters_P1_df.loc[coke_parameters_P1_df['Parameter']=='Coke Sulphur','Value'].item()]
    for i in quality_output_P1_df.columns[8:13]: 
        quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Sinter',i]=[sinter_P1_chemistry_output_df_P1.loc[sinter_P1_chemistry_output_df_P1['Parameter']==i,'Actual'].item()/sinter_P1_chemistry_output_df_P1.loc[sinter_P1_chemistry_output_df_P1['Parameter']=='Residue Production','Actual'].item()]
    quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Sinter','Basicity']=quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Sinter','CaO'].item()/quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Sinter','SiO2'].item()

    for i in quality_output_P1_df.columns[8:13]: 
        quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Pellet',i]=[pellet_bf_P1_chemistry_output_df_P1.loc[pellet_bf_P1_chemistry_output_df_P1['Parameter']==i,'Actual'].item()/pellet_bf_P1_chemistry_output_df_P1.loc[pellet_bf_P1_chemistry_output_df_P1['Parameter']=='Residue Production','Actual'].item()]
    quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Pellet','Basicity']=quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Pellet','CaO'].item()/quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Pellet','SiO2'].item()
    quality_output_P1_df

    for i in quality_output_P1_df.columns[8:13]: 
        quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Lump',i]=[lump_bf_P1_chemistry_output_df_P1.loc[lump_bf_P1_chemistry_output_df_P1['Parameter']==i,'Actual'].item()/lump_bf_P1_chemistry_output_df_P1.loc[lump_bf_P1_chemistry_output_df_P1['Parameter']=='Residue Production','Actual'].item()]
    quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Lump','Basicity']=quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Lump','CaO'].item()/quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Lump','SiO2'].item()
    quality_output_P1_df

    for i in quality_output_P1_df['Category 1'][4:7]:
        quality_output_P1_df.loc[quality_output_P1_df['Category 1']==i,'Moi']=[(var_bmo_P1[i]*bmo_materials.loc[bmo_materials['Material']==i,'Moi'].item()/100).value()]

    for i in quality_output_P1_df.columns[pd.np.r_[8:13, 15:17]]: 
        quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Limestone',i]=[Limestone_bf_P1_chemistry_output_df_P1.loc[Limestone_bf_P1_chemistry_output_df_P1['Parameter']==i,'Actual'].item()]
    if var_bmo_P1['Limestone'].value()!=0:
        quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Limestone','Basicity']=quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Limestone','CaO'].item()/quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Limestone','SiO2'].item()

    for i in quality_output_P1_df.columns[pd.np.r_[8:13, 15:17]]:
        quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Dolomite',i]=[Dolomite_bf_P1_chemistry_output_df_P1.loc[Dolomite_bf_P1_chemistry_output_df_P1['Parameter']==i,'Actual'].item()]
    if var_bmo_P1['Dolomite'].value()!=0:
        quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Dolomite','Basicity']=quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Dolomite','CaO'].item()/quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Dolomite','SiO2'].item()

    for i in quality_output_P1_df.columns[pd.np.r_[8:13, 15:17]]: 
        quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Quartz',i]=[Quartz_bf_P1_chemistry_output_df_P1.loc[Quartz_bf_P1_chemistry_output_df_P1['Parameter']==i,'Actual'].item()]
    if var_bmo_P1['Quartz'].value()!=0:
        quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Quartz','Basicity']=quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Quartz','CaO'].item()/quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Quartz','SiO2'].item()

    for i in quality_output_P1_df.columns[pd.np.r_[8:13, 15]]:
        quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Coke',i]=[coke_bf_P1_chemistry_output_df_P1.loc[coke_bf_P1_chemistry_output_df_P1['Parameter']==i,'Actual'].item()]
    if lpSum(var_coke_P1).value()!=0:
        quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Coke','Basicity']=quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Coke','CaO'].item()/quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Coke','SiO2'].item()
    quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Coke','FC']=blended_coal_parameters_df_P1.loc[blended_coal_parameters_df_P1['Parameter']=='FC','Value'].item()

    #for i in quality_output_P1_df.columns[8:18]:
    #    quality_output_P1_df.loc[quality_output_P1_df['Category 1']=='Coke',i]=blended_coal_parameters_df_P1.loc[blended_coal_parameters_df_P1['Parameter']==i,'Value'].item()


    # In[542]:


    #pellet_bf_P2_chemistry_output_df_P2['Actual']=
    fe_pellet_bf_P2=output((lpSum(mines_bf_P2['Fe(new)_fines']))*100)
    SiO2_pellet_bf_P2=output((lpSum(mines_bf_P2['SiO2(new)_fines']))*100)
    al2o3_pellet_bf_P2=output((lpSum(mines_bf_P2['Al2O3(new)_fines']))*100)
    cao_pellet_bf_P2=output((lpSum(mines_bf_P2['CaO(new)_fines']))*100)
    mgo_pellet_bf_P2=output((lpSum(mines_bf_P2['MgO(new)_fines']))*100)
    basicity_P2=output((lpSum(mines_bf_P2['CaO(new)_fines']))*100)
    pellet_bf_P2_chemistry_dict_P2={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','Basicity','Residue Production'],
                                       'Actual':[fe_pellet_bf_P2,SiO2_pellet_bf_P2,al2o3_pellet_bf_P2,cao_pellet_bf_P2,mgo_pellet_bf_P2,basicity_P2,lpSum(mines_bf_P2['Residue_fines']).value()]}
    pellet_bf_P2_chemistry_output_df_P2=pd.DataFrame(data=pellet_bf_P2_chemistry_dict_P2).round(1)
    pellet_bf_P2_chemistry_output_df_P2
    #lump_bf_P2_chemistry_output_df_P2['Actual']=
    fe_lump_bf_P2=output((lpSum(mines_lumps['Fe(new)_lumps']))*100)
    SiO2_lump_bf_P2=output((lpSum(mines_lumps['SiO2(new)_lumps']))*100)
    al2o3_lump_bf_P2=output((lpSum(mines_lumps['Al2O3(new)_lumps']))*100)
    cao_lump_bf_P2=output((lpSum(mines_lumps['CaO(new)_lumps']))*100)
    mgo_lump_bf_P2=output((lpSum(mines_lumps['MgO(new)_lumps']))*100)
    basicity_P2=output((lpSum(mines_lumps['CaO(new)_lumps']))*100)
    lump_bf_P2_chemistry_dict_P2={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','Basicity','Residue Production'],
                                       'Actual':[fe_lump_bf_P2,SiO2_lump_bf_P2,al2o3_lump_bf_P2,cao_lump_bf_P2,mgo_lump_bf_P2,basicity_P2,lpSum(mines_lumps['Residue_lumps']).value()]}
    lump_bf_P2_chemistry_output_df_P2=pd.DataFrame(data=lump_bf_P2_chemistry_dict_P2).round(1)
    lump_bf_P2_chemistry_output_df_P2
    #Limestone_bf_P2_chemistry_output_df_P2['Actual']=
    Limestone_bf_P2_chemistry_dict_P2={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','P','S'],
                                           'Actual':[0 for i in range(0,7)]}
    Limestone_bf_P2_chemistry_output_df_P2=pd.DataFrame(data=Limestone_bf_P2_chemistry_dict_P2)
    if lpSum(var_bmo_P2['Limestone']).value()!=0:
        Limestone_bf_P2_chemistry_output_df_P2.loc[Limestone_bf_P2_chemistry_output_df_P2['Parameter']=='Fe','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','Fe(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P2']).value())
        Limestone_bf_P2_chemistry_output_df_P2.loc[Limestone_bf_P2_chemistry_output_df_P2['Parameter']=='SiO2','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','SiO2(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P2']).value())
        Limestone_bf_P2_chemistry_output_df_P2.loc[Limestone_bf_P2_chemistry_output_df_P2['Parameter']=='Al2O3','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','Al2O3(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P2']).value())
        Limestone_bf_P2_chemistry_output_df_P2.loc[Limestone_bf_P2_chemistry_output_df_P2['Parameter']=='CaO','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','CaO(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P2']).value())
        Limestone_bf_P2_chemistry_output_df_P2.loc[Limestone_bf_P2_chemistry_output_df_P2['Parameter']=='MgO','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','MgO(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P2']).value())
        Limestone_bf_P2_chemistry_output_df_P2.loc[Limestone_bf_P2_chemistry_output_df_P2['Parameter']=='P','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','P(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P2']).value())
        Limestone_bf_P2_chemistry_output_df_P2.loc[Limestone_bf_P2_chemistry_output_df_P2['Parameter']=='S','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','S(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Limestone']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Limestone','feeding_dmt_P2']).value())
        Limestone_bf_P2_chemistry_output_df_P2=Limestone_bf_P2_chemistry_output_df_P2.round()
    Limestone_bf_P2_chemistry_output_df_P2
    #Dolomite_bf_P2_chemistry_output_df_P2['Actual']=
    Dolomite_bf_P2_chemistry_dict_P2={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','P','S'],
                                           'Actual':[0 for i in range(0,7)]}
    Dolomite_bf_P2_chemistry_output_df_P2=pd.DataFrame(data=Dolomite_bf_P2_chemistry_dict_P2)
    if lpSum(var_bmo_P2['Dolomite']).value()!=0:
        Dolomite_bf_P2_chemistry_output_df_P2.loc[Dolomite_bf_P2_chemistry_output_df_P2['Parameter']=='Fe','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','Fe(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P2']).value())
        Dolomite_bf_P2_chemistry_output_df_P2.loc[Dolomite_bf_P2_chemistry_output_df_P2['Parameter']=='SiO2','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','SiO2(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P2']).value())
        Dolomite_bf_P2_chemistry_output_df_P2.loc[Dolomite_bf_P2_chemistry_output_df_P2['Parameter']=='Al2O3','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','Al2O3(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P2']).value())
        Dolomite_bf_P2_chemistry_output_df_P2.loc[Dolomite_bf_P2_chemistry_output_df_P2['Parameter']=='CaO','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','CaO(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P2']).value())
        Dolomite_bf_P2_chemistry_output_df_P2.loc[Dolomite_bf_P2_chemistry_output_df_P2['Parameter']=='MgO','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','MgO(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P2']).value())
        Dolomite_bf_P2_chemistry_output_df_P2.loc[Dolomite_bf_P2_chemistry_output_df_P2['Parameter']=='P','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','P(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P2']).value())
        Dolomite_bf_P2_chemistry_output_df_P2.loc[Dolomite_bf_P2_chemistry_output_df_P2['Parameter']=='S','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','S(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Dolomite']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Dolomite','feeding_dmt_P2']).value())
        Dolomite_bf_P2_chemistry_output_df_P2=Dolomite_bf_P2_chemistry_output_df_P2.round(2)
    Dolomite_bf_P2_chemistry_output_df_P2
    #Quartz_bf_P2_chemistry_output_df_P2['Actual']=
    Quartz_bf_P2_chemistry_dict_P2={'Parameter':['Fe','SiO2','Al2O3','CaO','MgO','P','S'],
                                           'Actual':[0 for i in range(0,7)]}
    Quartz_bf_P2_chemistry_output_df_P2=pd.DataFrame(data=Quartz_bf_P2_chemistry_dict_P2)
    if lpSum(var_bmo_P2['Quartz']).value()!=0:
        Quartz_bf_P2_chemistry_output_df_P2.loc[Quartz_bf_P2_chemistry_output_df_P2['Parameter']=='Fe','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','Fe(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P2']).value())
        Quartz_bf_P2_chemistry_output_df_P2.loc[Quartz_bf_P2_chemistry_output_df_P2['Parameter']=='SiO2','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','SiO2(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P2']).value())
        Quartz_bf_P2_chemistry_output_df_P2.loc[Quartz_bf_P2_chemistry_output_df_P2['Parameter']=='Al2O3','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','Al2O3(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P2']).value())
        Quartz_bf_P2_chemistry_output_df_P2.loc[Quartz_bf_P2_chemistry_output_df_P2['Parameter']=='CaO','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','CaO(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P2']).value())
        Quartz_bf_P2_chemistry_output_df_P2.loc[Quartz_bf_P2_chemistry_output_df_P2['Parameter']=='MgO','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','MgO(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P2']).value())
        Quartz_bf_P2_chemistry_output_df_P2.loc[Quartz_bf_P2_chemistry_output_df_P2['Parameter']=='P','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','P(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P2']).value())
        Quartz_bf_P2_chemistry_output_df_P2.loc[Quartz_bf_P2_chemistry_output_df_P2['Parameter']=='S','Actual']=(lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','S(new)_P2']*100)).value()/(lpSum(var_bmo_P2['Quartz']).value()-lpSum(bmo_materials.loc[bmo_materials['Material']=='Quartz','feeding_dmt_P2']).value())
        Quartz_bf_P2_chemistry_output_df_P2=Quartz_bf_P2_chemistry_output_df_P2.round(1)
    Quartz_bf_P2_chemistry_output_df_P2
    #Quality Output
    quality_output_P2_dict={'Company':['C1' for i in range(7)],'Main Plant':['P2' for i in range(7)],
                              'Serving Plant':['Blast Furnace'for i in range(7)],
                             'Category 1':['Coke','Sinter','Pellet','Lump','Limestone','Dolomite','Quartz'],
                         'Category 2':['Coke','Sinter','Pellet','Lump','Flux','Flux','Flux'],'CSR':['None' for i in range(7)],'Ash':['None' for i in range(7)],
                         'Moi':['None' for i in range(7)],'Fe':['None' for i in range(7)],'SiO2':['None' for i in range(7)],'Al2O3':['None' for i in range(7)],'CaO':['None' for i in range(7)]
                         ,'MgO':['None' for i in range(7)],'Basicity':['None' for i in range(7)],'FC':['None' for i in range(7)],'P':['None' for i in range(7)],'S':['None' for i in range(7)]}#,
                         #'VM':['None' for i in range(7) ],'CSN':['None' for i in range(7)],'MMR':['None' for i in range(7)],'Vitrinite':['None' for i in range(7)],
                         #'V9-V14':['None' for i in range(7)],'Sulphur':['None' for i in range(7)],'Ash':['None' for i in range(7)],
                         #'TM':['None' for i in range(7)],'FC':['None' for i in range(7)],'Log Fluidity':['None' for i in range(7)]}

    quality_output_P2_df=pd.DataFrame(data=quality_output_P2_dict)
    quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Coke','CSR']=[coke_parameters_P2_df.loc[coke_parameters_P2_df['Parameter']=='CSR','Value'].item()]
    quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Coke','Ash']=[coke_parameters_P2_df.loc[coke_parameters_P2_df['Parameter']=='Coke Ash','Value'].item()]
    quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Coke','S']=[coke_parameters_P2_df.loc[coke_parameters_P2_df['Parameter']=='Coke Sulphur','Value'].item()]
    for i in quality_output_P2_df.columns[8:13]: 
        quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Sinter',i]=[sinter_P2_chemistry_output_df_P2.loc[sinter_P2_chemistry_output_df_P2['Parameter']==i,'Actual'].item()/sinter_P2_chemistry_output_df_P2.loc[sinter_P2_chemistry_output_df_P2['Parameter']=='Residue Production','Actual'].item()]
    quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Sinter','Basicity']=quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Sinter','CaO'].item()/quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Sinter','SiO2'].item()

    for i in quality_output_P2_df.columns[8:13]: 
        quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Pellet',i]=[pellet_bf_P2_chemistry_output_df_P2.loc[pellet_bf_P2_chemistry_output_df_P2['Parameter']==i,'Actual'].item()/pellet_bf_P2_chemistry_output_df_P2.loc[pellet_bf_P2_chemistry_output_df_P2['Parameter']=='Residue Production','Actual'].item()]
    quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Pellet','Basicity']=quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Pellet','CaO'].item()/quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Pellet','SiO2'].item()
    quality_output_P2_df

    for i in quality_output_P2_df.columns[8:13]: 
        quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Lump',i]=[lump_bf_P2_chemistry_output_df_P2.loc[lump_bf_P2_chemistry_output_df_P2['Parameter']==i,'Actual'].item()/lump_bf_P2_chemistry_output_df_P2.loc[lump_bf_P2_chemistry_output_df_P2['Parameter']=='Residue Production','Actual'].item()]
    quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Lump','Basicity']=quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Lump','CaO'].item()/quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Lump','SiO2'].item()
    quality_output_P2_df

    for i in quality_output_P2_df['Category 1'][4:7]:
        quality_output_P2_df.loc[quality_output_P2_df['Category 1']==i,'Moi']=[(var_bmo_P2[i]*bmo_materials.loc[bmo_materials['Material']==i,'Moi'].item()/100).value()]

    for i in quality_output_P2_df.columns[pd.np.r_[8:13, 15:17]]: 
        quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Limestone',i]=[Limestone_bf_P2_chemistry_output_df_P2.loc[Limestone_bf_P2_chemistry_output_df_P2['Parameter']==i,'Actual'].item()]
    if var_bmo_P2['Limestone'].value()!=0:
        quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Limestone','Basicity']=quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Limestone','CaO'].item()/quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Limestone','SiO2'].item()

    for i in quality_output_P2_df.columns[pd.np.r_[8:13, 15:17]]:
        quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Dolomite',i]=[Dolomite_bf_P2_chemistry_output_df_P2.loc[Dolomite_bf_P2_chemistry_output_df_P2['Parameter']==i,'Actual'].item()]
    if var_bmo_P2['Dolomite'].value()!=0:
        quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Dolomite','Basicity']=quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Dolomite','CaO'].item()/quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Dolomite','SiO2'].item()

    for i in quality_output_P2_df.columns[pd.np.r_[8:13, 15:17]]: 
        quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Quartz',i]=[Quartz_bf_P2_chemistry_output_df_P2.loc[Quartz_bf_P2_chemistry_output_df_P2['Parameter']==i,'Actual'].item()]
    if var_bmo_P2['Quartz'].value()!=0:
        quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Quartz','Basicity']=quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Quartz','CaO'].item()/quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Quartz','SiO2'].item()

    for i in quality_output_P2_df.columns[pd.np.r_[8:13, 15]]:
        quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Coke',i]=[coke_bf_P2_chemistry_output_df_P2.loc[coke_bf_P2_chemistry_output_df_P2['Parameter']==i,'Actual'].item()]
    if lpSum(var_coke_P2).value()!=0:
        quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Coke','Basicity']=quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Coke','CaO'].item()/quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Coke','SiO2'].item()
    quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Coke','FC']=blended_coal_parameters_df_P2.loc[blended_coal_parameters_df_P2['Parameter']=='FC','Value'].item()
    if quality_output_P2_df.columns[7]=='Moi':
        del quality_output_P2_df['Moi']
    if quality_output_P1_df.columns[7]=='Moi':
        del quality_output_P1_df['Moi']
    #for i in quality_output_P2_df.columns[8:18]:
    #    quality_output_P2_df.loc[quality_output_P2_df['Category 1']=='Coke',i]=blended_coal_parameters_df_P2.loc[blended_coal_parameters_df_P2['Parameter']==i,'Value'].item()


    # In[547]:


    quality_output_final_df=pd.concat([quality_output_P1_df,quality_output_P2_df],ignore_index=True)
    quality_output_final_df=quality_output_final_df.drop('FC',axis=1)
    quality_output_final_df=quality_output_final_df.round(1)
    quality_output_final_df['No. of Planned Days']=no_of_days_input
    append_df_to_excel("Output/Quality Output.xlsx",quality_output_final_df,index=False)

    #quality_output_final_df.to_excel("Output/Quality Output.xlsx",index=False)
    quality_output_final_df
    
    bf_parameters_output_dict_P1={'Company':['C1' for i in range(9)],'Main Plant':['P1' for i in range(9)],
                          'Serving Plant':['Blast Furnace' for i in range(9)],'Parameter':['Slag' for i in range(9)]}

    bf_parameters_output_df_P1=pd.DataFrame(data=bf_parameters_output_dict_P1)
    bf_parameters_output_df_P1['Category']=bmo_chemistry_output_df_P1.index
    bf_parameters_output_df_P1['Value']=bmo_chemistry_output_df_P1['Value'].values

    bf_coke_parameters_output_dict_P1={'Company':['C1' for i in range(4)],'Main Plant':['P1' for i in range(4)],
                              'Serving Plant':['Blast Furnace' for i in range(4)],'Parameter':['Coke Rate' for i in range(4)]}
    bf_coke_parameters_output_df_P1=pd.DataFrame(data=bf_coke_parameters_output_dict_P1)
    bf_coke_parameters_output_df_P1['Category']=coke_op_output_df_P1.index
    bf_coke_parameters_output_df_P1['Value']=coke_op_output_df_P1['Coke Rate'].values

    bf_parameters_output_dict_P2={'Company':['C1' for i in range(9)],'Main Plant':['P2' for i in range(9)],
                              'Serving Plant':['Blast Furnace' for i in range(9)],'Parameter':['Slag' for i in range(9)]}

    bf_parameters_output_df_P2=pd.DataFrame(data=bf_parameters_output_dict_P2)
    bf_parameters_output_df_P2['Category']=bmo_chemistry_output_df_P2.index
    bf_parameters_output_df_P2['Value']=bmo_chemistry_output_df_P2['Value'].values

    bf_coke_parameters_output_dict_P2={'Company':['C1' for i in range(4)],'Main Plant':['P2' for i in range(4)],
                              'Serving Plant':['Blast Furnace' for i in range(4)],'Parameter':['Coke Rate' for i in range(4)]}
    bf_coke_parameters_output_df_P2=pd.DataFrame(data=bf_coke_parameters_output_dict_P2)
    bf_coke_parameters_output_df_P2['Category']=coke_op_output_df_P2.index
    bf_coke_parameters_output_df_P2['Value']=coke_op_output_df_P2['Coke Rate'].values

    bf_parameters_final=pd.concat([bf_parameters_output_df_P1,bf_coke_parameters_output_df_P1,bf_parameters_output_df_P2,bf_coke_parameters_output_df_P2],ignore_index=True)
    bf_parameters_final['No. of Planned Days']=no_of_days_input
    append_df_to_excel("Output/Blast Furnace Parameters.xlsx",bf_parameters_final,index=False)

    #bf_parameters_final.to_excel("Output/Blast Furnace Parameters.xlsx",index=False)
    bf_parameters_final
    messagebox.showinfo("Message","Output Folder Updated")
    coke_test=(lpSum(var_coke_P2[i]*(((100-(coal_data.loc[(coal_data['category']==i),'tm'])-((100-(coal_data.loc[(coal_data['category']==i),'tm']))*((coal_data.loc[(coal_data['category']==i),'vm'])/100))-burning_loss+end_vm+end_moisture)/100).item()) for i in var_coke_P2)).value()
    # In[252]:




button1=tk.Button(text="Run", command=model,background='gray95')
button1.grid(column=1,row=28)
window.mainloop()    
#window.mainloop()l

